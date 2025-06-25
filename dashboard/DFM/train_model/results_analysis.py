# -*- coding: utf-8 -*-
"""
ç»“æœåˆ†æã€ä¿å­˜å’Œç»˜å›¾ç›¸å…³å‡½æ•°
"""
import pandas as pd
import numpy as np
import os
import matplotlib
matplotlib.use('Agg') # Ensure Agg backend is used before importing pyplot
import matplotlib.pyplot as plt
import matplotlib.dates as mdates # <-- Add this import
import seaborn as sns
import unicodedata
from typing import Tuple, List, Dict, Union, Any, Optional # Added Optional
from sklearn.metrics import mean_squared_error, mean_absolute_error
import logging # Import logging
import pickle # <<< æ–°å¢ï¼šå¯¼å…¥ pickle æ¨¡å— (ç”¨äºé‡æ–°åŠ è½½å…ƒæ•°æ®)
# from .excel_utils import write_df_to_excel # Import the utility function <-- REMOVED
import unicodedata # ç”¨äºè§„èŒƒåŒ–å­—ç¬¦ä¸²
from collections import Counter # ç”¨äºå› å­è§£é‡Š
from statsmodels.tsa.stattools import adfuller # <--- æ–°å¢ ADF æ£€éªŒå¯¼å…¥ (ä¸ºäº† apply_stationarity_transforms)
from statsmodels.tsa.statespace.dynamic_factor_mq import DynamicFactorMQ
import datetime
import statsmodels.api as sm # ç”¨äº R2 è®¡ç®—
from openpyxl.utils import get_column_letter # <<< æ–°å¢å¯¼å…¥
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill # <<< æ‰©å±•å¯¼å…¥
try:
    from . import config # <--- æ–°å¢å¯¼å…¥ config
except ImportError:
    try:
        import config
    except ImportError:
        # å¦‚æœconfigæ¨¡å—ä¸å­˜åœ¨ï¼Œåˆ›å»ºä¸€ä¸ªç©ºçš„configå¯¹è±¡
        class Config:
            pass
        config = Config()
# <<< æ–°å¢ï¼šå¯¼å…¥ tune_dfm ä¸­çš„æ˜ å°„åŠ è½½å‡½æ•° (å¦‚æœéœ€è¦) >>>
# from .tune_dfm import load_mappings # ä¿æŒæ³¨é‡Šï¼Œå› ä¸ºæ˜ å°„åº”è¯¥ä»å…ƒæ•°æ®åŠ è½½
# --- ä¿®æ”¹ï¼šå¯¼å…¥æ–°å¢çš„ Dominance Analysis å‡½æ•° ---
try:
    from .analysis_utils import (
        calculate_metrics_with_lagged_target,
        calculate_factor_contributions,
        calculate_individual_variable_r2,
        calculate_industry_r2,
        calculate_factor_industry_r2,
        calculate_factor_type_r2,
        #calculate_dominance_analysis # <<< ç§»é™¤å¯¼å…¥
    )
except ImportError:
    try:
        from analysis_utils import (
            calculate_metrics_with_lagged_target,
            calculate_factor_contributions,
            calculate_individual_variable_r2,
            calculate_industry_r2,
            calculate_factor_industry_r2,
            calculate_factor_type_r2,
            #calculate_dominance_analysis # <<< ç§»é™¤å¯¼å…¥
        )
    except ImportError as e:
        print(f"ä¸¥é‡é”™è¯¯: æ— æ³•å¯¼å…¥analysis_utilsæ¨¡å—: {e}")
        print("è¿™å°†å¯¼è‡´æŒ‡æ ‡è®¡ç®—å¤±è´¥ï¼Œè¯·æ£€æŸ¥æ¨¡å—è·¯å¾„å’Œä¾èµ–")
        # ğŸ”¥ ä¿®å¤ï¼šä¸å†æä¾›è¿”å›N/Açš„æ¨¡æ‹Ÿå‡½æ•°ï¼Œè€Œæ˜¯æŠ›å‡ºé”™è¯¯
        # è¿™æ ·å¯ä»¥å¼ºåˆ¶è§£å†³å¯¼å…¥é—®é¢˜è€Œä¸æ˜¯éšè—å®ƒ
        raise ImportError(f"æ— æ³•å¯¼å…¥å…³é”®çš„analysis_utilsæ¨¡å—: {e}. è¯·æ£€æŸ¥æ¨¡å—è·¯å¾„å’Œä¾èµ–å…³ç³»ã€‚")
# --- ç»“æŸä¿®æ”¹ ---
from scipy.optimize import linear_sum_assignment # <<< æ–°å¢ï¼šå¯¼å…¥åŒˆç‰™åˆ©ç®—æ³•

# Get logger for this module
logger = logging.getLogger(__name__)

# å°è¯•å¯¼å…¥ openpyxlï¼Œå¦‚æœ Excel å†™å…¥éœ€è¦å®ƒ (è™½ç„¶ pandas ä¼šå¤„ç†)
# try:
#     import openpyxl
# except ImportError:
#     print("è­¦å‘Š: openpyxl æœªå®‰è£…ï¼ŒExcel (.xlsx) æ–‡ä»¶å†™å…¥å¯èƒ½å¤±è´¥ã€‚è¯·è¿è¡Œ 'pip install openpyxl'")

# --- æ–°å¢: å°è¯•è®¾ç½®ä¸­æ–‡æ”¯æŒ ---
try:
    plt.rcParams['font.sans-serif'] = ['SimHei'] # æˆ–è€… 'Microsoft YaHei', 'WenQuanYi Micro Hei' ç­‰
    plt.rcParams['axes.unicode_minus'] = False  # è§£å†³è´Ÿå·'-'æ˜¾ç¤ºä¸ºæ–¹å—çš„é—®é¢˜
except Exception as e:
    print(f"è®¾ç½® Matplotlib ä¸­æ–‡å­—ä½“å¤±è´¥: {e}")
# --- ç»“æŸæ–°å¢ ---


# <<< æ·»åŠ ä» tune_dfm.py ç§»åŠ¨è¿‡æ¥çš„å‡½æ•° >>>

def write_r2_tables_to_excel(
    r2_results: Optional[Dict[str, pd.DataFrame]],
    excel_writer: pd.ExcelWriter,
    sheet_name: str = "Factor R2 Analysis",
    industry_r2: Optional[pd.Series] = None,
    factor_industry_r2: Optional[Dict[str, pd.Series]] = None,
    factor_type_r2: Optional[Dict[str, pd.Series]] = None,
    # dominance_industry_summary: Optional[pd.DataFrame] = None # <<< ç§»é™¤ Dominance å‚æ•°
):
    """
    å°†å„ç§ R2 è¡¨æ ¼å†™å…¥ Excel æ–‡ä»¶ã€‚
    - å•å› å­å¯¹å„å˜é‡çš„ R2 (å¹¶æ’)
    - æ•´ä½“å› å­å¯¹å„è¡Œä¸šçš„ R2 (å•ä¸ªè¡¨æ ¼)
    - å•å› å­å¯¹å„è¡Œä¸šçš„ Pooled R2 (å•ä¸ªè¡¨æ ¼)
    - å•å› å­å¯¹å„ç±»å‹çš„ Pooled R2 (å•ä¸ªè¡¨æ ¼)
    # - Dominance Analysis: è¡Œä¸šå¯¹å› å­è´¡çŒ®æ±‡æ€» (å•ä¸ªè¡¨æ ¼) <-- ç§»é™¤ Dominance

    Args:
        r2_results: å•å› å­å¯¹å„å˜é‡ R2 çš„å­—å…¸ã€‚
        excel_writer: pandas ExcelWriter å¯¹è±¡ã€‚
        sheet_name: è¦å†™å…¥çš„ Sheet åç§°ã€‚
        industry_r2: æ•´ä½“å› å­å¯¹å„è¡Œä¸šçš„ R2 Seriesã€‚
        factor_industry_r2: å•å› å­å¯¹å„è¡Œä¸šçš„ Pooled R2 å­—å…¸ã€‚
        factor_type_r2: å•å› å­å¯¹å„ç±»å‹çš„ Pooled R2 å­—å…¸ã€‚
        # dominance_industry_summary: è¡Œä¸šå¯¹å› å­è´¡çŒ®çš„æ±‡æ€» DataFrameã€‚ <-- ç§»é™¤ Dominance å‚æ•°
    """
    
    # --- 1. å†™å…¥å•å› å­å¯¹å„å˜é‡çš„ R2 (å¹¶æ’) --- 
    max_row_indiv_r2 = 0
    if r2_results is not None and r2_results:
        print(f"\n--- å°†å› å­å¯¹å•å˜é‡ R2 åˆ†æè¡¨å†™å…¥ Excel Sheet: {sheet_name} (å¹¶æ’å¸ƒå±€) ---")
        try:
            workbook = excel_writer.book
            if sheet_name in excel_writer.sheets:
                worksheet = excel_writer.sheets[sheet_name]
            else:
                worksheet = workbook.create_sheet(title=sheet_name)
                excel_writer.sheets[sheet_name] = worksheet

            bold_font = Font(bold=True)
            start_col = 1
            max_row_written = 0 

            for factor_name in sorted(r2_results.keys()):
                factor_df = r2_results.get(factor_name)
                if factor_df is None or factor_df.empty:
                    continue
                all_vars_df = factor_df.copy()
                all_vars_df.insert(0, '#', range(1, len(all_vars_df) + 1))
                num_cols = len(all_vars_df.columns)
                current_row = 1 # Start writing from row 1
                
                # Write title
                title_cell = worksheet.cell(row=current_row, column=start_col, value=f"All variables explained by {factor_name}")
                title_cell.font = bold_font
                try:
                    worksheet.merge_cells(start_row=current_row, start_column=start_col, end_row=current_row, end_column=start_col + num_cols - 1)
                except ValueError:
                    pass # Ignore merge cell errors if cells are already merged or invalid range
                current_row += 1
                
                # Write header
                header_row = current_row
                for c_idx, value in enumerate(all_vars_df.columns.values):
                    header_cell = worksheet.cell(row=header_row, column=start_col + c_idx, value=value)
                    header_cell.font = bold_font
                current_row += 1
                
                # Write data
                for r_idx, row_data in enumerate(all_vars_df.itertuples(index=False)):
                    data_current_row = current_row + r_idx
                    for c_idx, value in enumerate(row_data):
                        cell = worksheet.cell(row=data_current_row, column=start_col + c_idx)
                        if isinstance(value, (float, np.number)) and (np.isnan(value) or np.isinf(value)): cell.value = None
                        else: cell.value = value
                        if all_vars_df.columns[c_idx] == 'R2': cell.number_format = '0.0000'
                        elif all_vars_df.columns[c_idx] == '#': cell.alignment = cell.alignment.copy(horizontal='center')
                    if data_current_row > max_row_written: max_row_written = data_current_row
                    
                # Adjust column widths
                for c_idx, col_name in enumerate(all_vars_df.columns):
                    col_letter = get_column_letter(start_col + c_idx)
                    try:
                        col_data_str = all_vars_df.iloc[:, c_idx].astype(str)
                        max_len_data = col_data_str.map(len).max()
                        if pd.isna(max_len_data): max_len_data = 4
                        header_len = len(str(col_name))
                        adjusted_width = max(max_len_data, header_len) + 2
                        if col_name == 'Variable': adjusted_width = max(adjusted_width, 35)
                        elif col_name == 'R2': adjusted_width = max(adjusted_width, 10)
                        elif col_name == '#': adjusted_width = 5
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                    except Exception: # Catch potential errors like empty data
                        worksheet.column_dimensions[col_letter].width = 15 # Default width
                        
                start_col += num_cols + 1
                
            max_row_indiv_r2 = max_row_written # è®°å½•è¿™éƒ¨åˆ†å†™å…¥çš„æœ€å¤§è¡Œ
            print(f"--- å› å­å¯¹å•å˜é‡ R2 åˆ†æè¡¨å†™å…¥å®Œæˆ --- (Max Row: {max_row_indiv_r2})")
        except Exception as e_indiv:
            print(f"å†™å…¥å•å› å­å¯¹å˜é‡ R2 (å¹¶æ’å¸ƒå±€) æ—¶å‘ç”Ÿé”™è¯¯: {e_indiv}")
            # traceback.print_exc() # ç§»é™¤ traceback æ‰“å°ä»¥å‡å°‘æ—¥å¿—å™ªéŸ³
            max_row_indiv_r2 = 1 # å‡è®¾è‡³å°‘å†™äº†æ ‡é¢˜è¡Œ
    else:
        print(f"  æœªæä¾›å•å˜é‡ R2 ç»“æœ (r2_results)ï¼Œè·³è¿‡å¹¶æ’è¡¨æ ¼ã€‚")
        max_row_indiv_r2 = 0 # æ²¡æœ‰å†™å…¥è¡Œ
        # Ensure sheet exists if other tables need to be written
        try:
            if sheet_name not in excel_writer.sheets:
                 workbook = excel_writer.book
                 worksheet = workbook.create_sheet(title=sheet_name)
                 excel_writer.sheets[sheet_name] = worksheet
        except Exception as e_sheet_create:
             print(f"  è­¦å‘Š: å°è¯•ä¸ºåç»­è¡¨æ ¼åˆ›å»º Sheet '{sheet_name}' å¤±è´¥: {e_sheet_create}")


    # --- 2. å†™å…¥è¡Œä¸šå’Œ Pooled R2 è¡¨æ ¼ (ä»¥åŠ Dominance ç»“æœ) --- 
    current_row = max_row_indiv_r2 + 3 # åœ¨ä¹‹å‰è¡¨æ ¼ä¸‹æ–¹ç•™å‡º 2 è¡Œç©ºç™½
    start_col_combined = 1

    # è·å– worksheet (å¯èƒ½åœ¨ä¸Šé¢å·²åˆ›å»ºæˆ–è·å–)
    worksheet = excel_writer.sheets.get(sheet_name)
    if worksheet is None:
        print(f"é”™è¯¯ï¼šæ— æ³•è·å– Worksheet '{sheet_name}' ä»¥å†™å…¥åˆå¹¶çš„ R2 è¡¨æ ¼ã€‚")
        return # æ— æ³•ç»§ç»­
    bold_font = Font(bold=True)

    # --- Helper function to write a DataFrame table (ä¿æŒä¸å˜) ---
    def write_single_table(ws, df, title, start_r, start_c, bold_f, number_format='0.0000'): # Add number_format
        print(f"  æ­£åœ¨å†™å…¥è¡¨æ ¼: '{title}' (å¼€å§‹äº R{start_r}C{start_c})...")
        max_c_written = start_c -1 # Track max column written for this table
        # Write title
        try:
            title_cell = ws.cell(row=start_r, column=start_c, value=title)
            title_cell.font = bold_f
            current_r = start_r + 1
        except Exception as e_title:
            print(f"    é”™è¯¯: å†™å…¥è¡¨æ ¼æ ‡é¢˜ '{title}' æ—¶å‡ºé”™: {e_title}")
            return start_r # Return original row if title fails

        # Write header
        try:
            index_header = df.index.name if df.index.name else "Index"
            ws.cell(row=current_r, column=start_c, value=index_header).font = bold_f
            max_c_written = start_c
            for c_idx, col_name in enumerate(df.columns):
                col_c = start_c + 1 + c_idx
                ws.cell(row=current_r, column=col_c, value=col_name).font = bold_f
                max_c_written = col_c
            current_r += 1
        except Exception as e_header:
            print(f"    é”™è¯¯: å†™å…¥è¡¨æ ¼ '{title}' çš„è¡¨å¤´æ—¶å‡ºé”™: {e_header}")
            return start_r + 1 # Return row after attempted title

        # Write data
        try:
            for r_idx, index_val in enumerate(df.index):
                data_r = current_r + r_idx
                ws.cell(row=data_r, column=start_c, value=index_val)
                for c_idx, col_name in enumerate(df.columns):
                    col_c = start_c + 1 + c_idx
                    value = df.iloc[r_idx, c_idx]
                    cell = ws.cell(row=data_r, column=col_c)
                    if isinstance(value, (float, np.number)) and (np.isnan(value) or np.isinf(value)): cell.value = None
                    else: cell.value = value
                    cell.number_format = number_format # Apply specified format
            final_row = current_r + len(df) - 1
        except Exception as e_data:
            print(f"    é”™è¯¯: å†™å…¥è¡¨æ ¼ '{title}' çš„æ•°æ®æ—¶å‡ºé”™: {e_data}")
            return current_r # Return row where data writing started

        # Adjust column widths for this table
        try:
            # Index column
            col_letter = get_column_letter(start_c)
            index_header = df.index.name if df.index.name else "Index"
            max_len_index = max(len(str(index_header)), df.index.astype(str).map(len).max()) + 2
            ws.column_dimensions[col_letter].width = max(max_len_index, 15) # Min width 15 for index
            # Data columns
            for c_idx, col_name in enumerate(df.columns):
                col_c = start_c + 1 + c_idx
                col_letter = get_column_letter(col_c)
                # Format number as string with target format to estimate width
                if number_format.endswith('%'):
                     col_data_str = df.iloc[:, c_idx].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")
                elif '0' in number_format:
                     num_decimals = number_format.count('0', number_format.find('.')) if '.' in number_format else 0
                     col_data_str = df.iloc[:, c_idx].apply(lambda x: f"{x:.{num_decimals}f}" if pd.notna(x) else "")
                else:
                     col_data_str = df.iloc[:, c_idx].astype(str)

                max_len_data = col_data_str.map(len).max()
                if pd.isna(max_len_data): max_len_data = 6 # 'None' + buffer
                header_len = len(str(col_name))
                adjusted_width = max(max_len_data, header_len) + 2
                ws.column_dimensions[col_letter].width = max(adjusted_width, 12) # Min width 12 for R2/Contrib cols
        except Exception as e_width:
             print(f"    è­¦å‘Š: è°ƒæ•´è¡¨æ ¼ '{title}' åˆ—å®½æ—¶å‡ºé”™: {e_width}")

        return final_row # Return the last row written by this table
    # --- End Helper ---
    
    last_row_written = current_row - 1 # Initialize from starting point
    
    # --- Write Industry R2 --- 
    if industry_r2 is not None and not industry_r2.empty:
        df_industry_r2 = industry_r2.to_frame(name="Industry R2 (All Factors)")
        df_industry_r2.index.name = "Industry"
        last_row_written = write_single_table(worksheet, df_industry_r2, "Industry R2 (All Factors)", current_row, start_col_combined, bold_font)
        current_row = last_row_written + 3 # Update for next table
    else:
        print("  æœªæä¾›è¡Œä¸š R2 ç»“æœ (industry_r2)ï¼Œè·³è¿‡è¡¨æ ¼ã€‚")
        
    # --- Write Factor-Industry Pooled R2 --- 
    if factor_industry_r2 is not None and factor_industry_r2:
        try:
            df_factor_industry = pd.DataFrame(factor_industry_r2)
            df_factor_industry.index.name = "Industry"
            df_factor_industry = df_factor_industry.sort_index() # Sort by industry name
            # Ensure only factor columns, handling potential non-factor keys
            factor_cols = sorted([col for col in df_factor_industry.columns if col.startswith('Factor')], 
                                key=lambda x: int(x.replace('Factor','')))
            df_factor_industry = df_factor_industry[factor_cols] 
            last_row_written = write_single_table(worksheet, df_factor_industry, "Factor-Industry Pooled RÂ²", current_row, start_col_combined, bold_font)
            current_row = last_row_written + 3
        except Exception as e_fi_prep:
            print(f"  å‡†å¤‡æˆ–å†™å…¥ Factor-Industry Pooled R2 è¡¨æ ¼æ—¶å‡ºé”™: {e_fi_prep}")
    else:
         print("  æœªæä¾›å› å­-è¡Œä¸š Pooled R2 ç»“æœ (factor_industry_r2)ï¼Œè·³è¿‡è¡¨æ ¼ã€‚")

    # --- Write Factor-Type Pooled R2 --- 
    # ... (ä»£ç ä¿æŒä¸å˜) ...
    logger.info(f"[Debug Write R2] Received 'factor_type_r2'. Type: {type(factor_type_r2)}. Is None or Empty: {factor_type_r2 is None or (isinstance(factor_type_r2, dict) and not factor_type_r2)}")
    if isinstance(factor_type_r2, dict) and factor_type_r2:
        logger.info(f"[Debug Write R2] 'factor_type_r2' keys: {list(factor_type_r2.keys())}")

    if factor_type_r2 is not None and factor_type_r2:
        try:
            df_factor_type = pd.DataFrame(factor_type_r2)
            df_factor_type.index.name = "Type" # Set index name to Type
            df_factor_type = df_factor_type.sort_index() # Sort by type name
            # Ensure only factor columns and sort them
            factor_cols = sorted([col for col in df_factor_type.columns if col.startswith('Factor')],
                                key=lambda x: int(x.replace('Factor','')))
            df_factor_type = df_factor_type[factor_cols]
            last_row_written = write_single_table(worksheet, df_factor_type, "Factor-Type Pooled RÂ²", current_row, start_col_combined, bold_font)
            current_row = last_row_written + 3 # Update row counter
        except Exception as e_ft_prep:
            print(f"  å‡†å¤‡æˆ–å†™å…¥ Factor-Type Pooled R2 è¡¨æ ¼æ—¶å‡ºé”™: {e_ft_prep}")
            # traceback.print_exc() # Removed for brevity
    else:
        print("  æœªæä¾›å› å­-ç±»å‹ Pooled R2 ç»“æœ (factor_type_r2)ï¼Œè·³è¿‡è¡¨æ ¼ã€‚")
        
    # --- <<< ç§»é™¤ï¼šå†™å…¥ Dominance Analysis è¡Œä¸šæ±‡æ€» >>> ---
    # if dominance_industry_summary is not None and not dominance_industry_summary.empty:
    #     logger.info("  æ­£åœ¨å†™å…¥ Dominance Analysis è¡Œä¸šè´¡çŒ®æ±‡æ€»...")
    #     try:
    #         # ç¡®ä¿å› å­åˆ—æŒ‰ Factor1, Factor2... æ’åº
    #         factor_cols = sorted([col for col in dominance_industry_summary.columns if col.startswith('Factor')],
    #                             key=lambda x: int(x.replace('Factor','')))
    #         df_dominance = dominance_industry_summary[factor_cols].copy()
    #         df_dominance = df_dominance.sort_index() # æŒ‰è¡Œä¸šæ’åº
    #         df_dominance.index.name = "Industry"
    #         
    #         # ä½¿ç”¨è¾…åŠ©å‡½æ•°å†™å…¥è¡¨æ ¼ï¼ŒæŒ‡å®šç™¾åˆ†æ¯”æ ¼å¼
    #         last_row_written = write_single_table(
    #             worksheet, 
    #             df_dominance, 
    #             "Dominance Analysis: Industry Contribution to Factor RÂ² (%)", 
    #             current_row, 
    #             start_col_combined, 
    #             bold_font,
    #             number_format='0.00' # Display as percentage with 2 decimals (e.g., 12.34)
    #         )
    #         current_row = last_row_written + 3
    #         logger.info("  Dominance Analysis è¡¨æ ¼å†™å…¥å®Œæˆã€‚")
    #     except Exception as e_dom_write:
    #         print(f"  å†™å…¥ Dominance Analysis è¡Œä¸šæ±‡æ€»è¡¨æ ¼æ—¶å‡ºé”™: {e_dom_write}")
    # else:
    #     print("  æœªæä¾› Dominance Analysis è¡Œä¸šæ±‡æ€»ç»“æœ (dominance_industry_summary)ï¼Œè·³è¿‡è¡¨æ ¼ã€‚")
    # --- <<< ç»“æŸç§»é™¤ >>> ---

    print(f"--- æ‰€æœ‰ R2 åˆ†æè¡¨æ ¼å†™å…¥å®Œæˆ (Sheet: {sheet_name}) ---") # <-- æ›´æ–°ç»“æŸæ¶ˆæ¯
# <<< ç»“æŸæ·»åŠ å‡½æ•° >>>

# <<< æ–°å¢: åˆ›å»ºå¯¹é½çš„ Nowcast vs Target è¡¨æ ¼çš„è¾…åŠ©å‡½æ•° >>>
def create_aligned_nowcast_target_table(
    nowcast_weekly_orig: pd.Series,
    target_orig: pd.Series,
    target_variable_name: str = "Target (Original Scale)" # å…è®¸è‡ªå®šä¹‰ç›®æ ‡åˆ—å
) -> pd.DataFrame:
    """
    å°†å‘¨åº¦ Nowcast ä¸æœˆåº¦ Target å¯¹é½ã€‚

    å¯¹é½è§„åˆ™ï¼š
    1. é€‰å–æ¯ä¸ªæœˆæœ€åä¸€ä¸ªå‘¨äº”çš„ Nowcast å€¼ã€‚
    2. å°†è¯¥ Nowcast å€¼ä¸ä¸‹ä¸€ä¸ªæœˆçš„ Target å€¼è¿›è¡ŒåŒ¹é…ã€‚
    3. è¿”å›çš„ DataFrame ä»¥æ¯æœˆæœ€åä¸€ä¸ªå‘¨äº”ä½œä¸ºç´¢å¼•ã€‚

    Args:
        nowcast_weekly_orig: å‘¨åº¦é¢‘ç‡çš„ Nowcast æ—¶é—´åºåˆ— (åŸå§‹å°ºåº¦)ã€‚
        target_orig: åŸå§‹é¢‘ç‡çš„ç›®æ ‡å˜é‡æ—¶é—´åºåˆ— (åŸå§‹å°ºåº¦)ã€‚
        target_variable_name: åœ¨è¾“å‡º DataFrame ä¸­ç”¨äºç›®æ ‡åˆ—çš„åç§°ã€‚

    Returns:
        åŒ…å«å¯¹é½åçš„ Nowcast å’Œ Target çš„ DataFrameã€‚
        ç´¢å¼•ï¼šæ¯æœˆæœ€åä¸€ä¸ªå‘¨äº”çš„æ—¥æœŸã€‚
        åˆ—ï¼š['Nowcast (Original Scale)', target_variable_name]
    """
    print("ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: å¼€å§‹åˆ›å»ºå¯¹é½çš„ Nowcast vs Target è¡¨æ ¼...")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcast_weekly_origç±»å‹: {type(nowcast_weekly_orig)}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcast_weekly_origé•¿åº¦: {len(nowcast_weekly_orig) if nowcast_weekly_orig is not None else 'None'}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_origç±»å‹: {type(target_orig)}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_origé•¿åº¦: {len(target_orig) if target_orig is not None else 'None'}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_variable_name: {target_variable_name}")

    logger.info("å¼€å§‹åˆ›å»ºå¯¹é½çš„ Nowcast vs Target è¡¨æ ¼...")

    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: å¼€å§‹å¤„ç†ç´¢å¼•...")
    if not isinstance(nowcast_weekly_orig.index, pd.DatetimeIndex):
        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: è½¬æ¢nowcastç´¢å¼•ä¸ºDatetimeIndex")
        nowcast_weekly_orig.index = pd.to_datetime(nowcast_weekly_orig.index)
    if not isinstance(target_orig.index, pd.DatetimeIndex):
        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: è½¬æ¢targetç´¢å¼•ä¸ºDatetimeIndex")
        target_orig.index = pd.to_datetime(target_orig.index)

    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcastæ—¶é—´èŒƒå›´: {nowcast_weekly_orig.index.min()} åˆ° {nowcast_weekly_orig.index.max()}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: targetæ—¶é—´èŒƒå›´: {target_orig.index.min()} åˆ° {target_orig.index.max()}")

    # 1. ğŸ”¥ ä¿®å¤ï¼šä¿ç•™æ‰€æœ‰å‘¨åº¦ Nowcast æ•°æ®ï¼Œä¸åªæ˜¯æ¯æœˆæœ€åä¸€ä¸ªå‘¨äº”
    # ç­›é€‰å‡ºæ‰€æœ‰å‘¨äº”
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: ç­›é€‰å‘¨äº”æ•°æ®...")
    fridays_index = nowcast_weekly_orig[nowcast_weekly_orig.index.dayofweek == 4].index
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: æ‰¾åˆ° {len(fridays_index)} ä¸ªå‘¨äº”")
    if fridays_index.empty:
        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: âŒ æ²¡æœ‰æ‰¾åˆ°å‘¨äº”æ•°æ®ï¼")
        logger.warning("åœ¨ Nowcast åºåˆ—ä¸­æœªæ‰¾åˆ°ä»»ä½•å‘¨äº”ï¼Œæ— æ³•åˆ›å»ºå¯¹é½è¡¨æ ¼ã€‚")
        return pd.DataFrame(columns=['Nowcast (Original Scale)', target_variable_name])

    # ğŸ”¥ ä¿®å¤ï¼šä½¿ç”¨æ‰€æœ‰å‘¨äº”çš„ Nowcast å€¼ï¼Œè€Œä¸æ˜¯åªé€‰æ‹©æ¯æœˆæœ€åä¸€ä¸ªå‘¨äº”
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: æå–æ‰€æœ‰å‘¨äº”çš„nowcastæ•°æ®...")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: fridays_indexé•¿åº¦: {len(fridays_index)}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: fridays_indexå‰5ä¸ª: {fridays_index[:5].tolist()}")

    nowcast_all_fridays = nowcast_weekly_orig.loc[fridays_index].copy()
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcast_all_fridaysé•¿åº¦: {len(nowcast_all_fridays)}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcast_all_fridayså‰3ä¸ªå€¼:")
    print(nowcast_all_fridays.head(3))

    nowcast_all_fridays.name = 'Nowcast (Original Scale)'

    # åŒæ—¶è®¡ç®—æ¯æœˆæœ€åä¸€ä¸ªå‘¨äº”ï¼ˆç”¨äºå¯¹é½çœŸå®å€¼ï¼‰
    try:
        last_fridays = fridays_index.to_series().groupby(fridays_index.to_period('M')).max()
    except Exception as e:
         logger.error(f"æŒ‰æœˆæŸ¥æ‰¾æœ€åä¸€ä¸ªå‘¨äº”æ—¶å‡ºé”™: {e}. Nowcast index type: {type(nowcast_weekly_orig.index)}, first few: {nowcast_weekly_orig.index[:5]}")
         return pd.DataFrame(columns=['Nowcast (Original Scale)', target_variable_name])

    # 2. å‡†å¤‡ Target æ•°æ®
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: å‡†å¤‡Targetæ•°æ®...")
    # ğŸ”¥ ä¿®å¤ï¼šæ­£ç¡®ç†è§£ç›®æ ‡æ•°æ®çš„æ—¶é—´å«ä¹‰
    # ç›®æ ‡æ•°æ®çš„ç´¢å¼•æ˜¯å‘å¸ƒæ—¥æœŸï¼Œå‘å¸ƒæ—¥æœŸå°±ä»£è¡¨æ•°æ®æ‰€å±çš„æœˆä»½
    # ä¾‹å¦‚ï¼š2023-01-31å‘å¸ƒçš„æ•°æ®ä»£è¡¨2023å¹´1æœˆçš„æ•°æ®
    target_df = target_orig.dropna().to_frame(target_variable_name)
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_dfå½¢çŠ¶: {target_df.shape}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_dféç©ºæ•°æ®ç‚¹: {len(target_df)}")
    if len(target_df) > 0:
        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_dfå‰3è¡Œ:")
        print(target_df.head(3))

    # ğŸ”¥ ä¿®å¤ï¼šå‘å¸ƒæ—¥æœŸå°±æ˜¯æ•°æ®æ‰€å±æœˆä»½ï¼Œä¸éœ€è¦å‡ä¸€ä¸ªæœˆ
    # æ­£ç¡®é€»è¾‘ï¼šå‘å¸ƒæ—¥æœŸç›´æ¥è½¬æ¢ä¸ºPeriod
    # ä¾‹å¦‚ï¼š2023-01-31å‘å¸ƒçš„æ•°æ® -> 2023-01 Period
    target_df['TargetPeriod'] = target_df.index.to_period('M')
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: æ·»åŠ TargetPeriodåˆ—å®Œæˆ")
    # ğŸ”¥ ç»“æŸä¿®å¤

    # 3. ğŸ”¥ ä¿®å¤ï¼šåˆ›å»ºå®Œæ•´çš„å‘¨åº¦å¯¹é½ DataFrame
    # åˆ›å»ºåŸºç¡€è¡¨æ ¼ï¼ŒåŒ…å«æ‰€æœ‰å‘¨äº”çš„ Nowcast æ•°æ®
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: åˆ›å»ºfinal_aligned_table...")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcast_all_fridaysç±»å‹: {type(nowcast_all_fridays)}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: nowcast_all_fridaysæ˜¯å¦ä¸ºç©º: {nowcast_all_fridays.empty if hasattr(nowcast_all_fridays, 'empty') else 'N/A'}")

    final_aligned_table = nowcast_all_fridays.to_frame()
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: final_aligned_tableåˆå§‹å½¢çŠ¶: {final_aligned_table.shape}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: final_aligned_tableåˆå§‹åˆ—å: {list(final_aligned_table.columns)}")

    final_aligned_table[target_variable_name] = np.nan
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: æ·»åŠ targetåˆ—åå½¢çŠ¶: {final_aligned_table.shape}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: æ·»åŠ targetåˆ—ååˆ—å: {list(final_aligned_table.columns)}")
    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_variable_name: {target_variable_name}")

    # 4. ğŸ”¥ ä¿®å¤ï¼šæ­£ç¡®çš„å¯¹é½é€»è¾‘ - æ¯ä¸ªæœˆæœ€åä¸€ä¸ªå‘¨äº”å¯¹åº”ä¸‹ä¸ªæœˆçš„çœŸå®å€¼
    # å¯¹äºæ¯ä¸ªçœŸå®å€¼ï¼Œæ‰¾åˆ°å¯¹åº”çš„ä¸Šä¸ªæœˆæœ€åä¸€ä¸ªå‘¨äº”
    logger.info(f"å¼€å§‹å¯¹é½çœŸå®å€¼æ•°æ®ï¼Œå…±æœ‰ {len(target_df)} ä¸ªçœŸå®å€¼")
    logger.info(f"çœŸå®å€¼æ—¶é—´èŒƒå›´: {target_df.index.min()} åˆ° {target_df.index.max()}")
    logger.info(f"å¯ç”¨çš„æœˆåº¦æœ€åå‘¨äº”: {len(last_fridays)} ä¸ª")

    aligned_count = 0
    for target_date, target_value in target_df[target_variable_name].items():
        target_period = target_df.loc[target_date, 'TargetPeriod']
        logger.debug(f"å¤„ç†çœŸå®å€¼: {target_value:.2f} (å‘å¸ƒäº{target_date.strftime('%Y-%m-%d')}, æœŸé—´: {target_period})")

        # ğŸ”¥ å…³é”®ä¿®å¤ï¼šæ‰¾åˆ°ä¸Šä¸ªæœˆçš„æœ€åä¸€ä¸ªå‘¨äº”
        prev_month = target_period - 1
        logger.debug(f"  æŸ¥æ‰¾ä¸Šä¸ªæœˆ: {prev_month}")

        if prev_month in last_fridays.index:
            last_friday_date = last_fridays.loc[prev_month]
            logger.debug(f"  æ‰¾åˆ°ä¸Šä¸ªæœˆæœ€åä¸€ä¸ªå‘¨äº”: {last_friday_date.strftime('%Y-%m-%d')}")

            if last_friday_date in final_aligned_table.index:
                final_aligned_table.loc[last_friday_date, target_variable_name] = target_value
                aligned_count += 1
                logger.debug(f"  âœ… æˆåŠŸå¯¹é½: çœŸå®å€¼ {target_value:.2f} -> å‘¨äº” {last_friday_date.strftime('%Y-%m-%d')}")
            else:
                logger.warning(f"  âš ï¸ å‘¨äº”æ—¥æœŸ {last_friday_date.strftime('%Y-%m-%d')} ä¸åœ¨æœ€ç»ˆè¡¨æ ¼ç´¢å¼•ä¸­")
        else:
            logger.warning(f"  âš ï¸ ä¸Šä¸ªæœˆ {prev_month} æ²¡æœ‰æ‰¾åˆ°å¯¹åº”çš„æœ€åå‘¨äº”")

    logger.info(f"å¯¹é½å®Œæˆï¼ŒæˆåŠŸå¯¹é½äº† {aligned_count} ä¸ªçœŸå®å€¼")

    # æ¸…ç†å’Œæ’åºç»“æœ
    final_aligned_table = final_aligned_table.sort_index()
    logger.info(f"æˆåŠŸåˆ›å»ºå®Œæ•´çš„å‘¨åº¦å¯¹é½è¡¨æ ¼ï¼ŒåŒ…å« {len(final_aligned_table)} è¡Œæ•°æ®ã€‚")
    logger.info(f"å…¶ä¸­åŒ…å« {final_aligned_table['Nowcast (Original Scale)'].notna().sum()} ä¸ªNowcastå€¼")
    logger.info(f"å…¶ä¸­åŒ…å« {final_aligned_table[target_variable_name].notna().sum()} ä¸ªçœŸå®å€¼")

    return final_aligned_table
# <<< ç»“æŸæ–°å¢ >>>

# --- <<< æ–°å¢ï¼šé‡æ–°ç”Ÿæˆ plot_final_nowcast å‡½æ•° >>> ---
def plot_final_nowcast(
    final_nowcast_series: pd.Series,
    target_for_plot: pd.Series, # åŸå§‹ç›®æ ‡åºåˆ— (dropna å)
    validation_start: Union[str, pd.Timestamp],
    validation_end: Union[str, pd.Timestamp],
    title: str,
    filename: str
):
    """
    (é‡æ–°ç”Ÿæˆ) ç»˜åˆ¶æœ€ç»ˆçš„å‘¨åº¦ nowcast ä¸å®é™…è§‚æµ‹å€¼çš„å¯¹æ¯”å›¾ï¼ˆåŸå§‹æ°´å¹³ï¼‰ã€‚
    åŠŸèƒ½å°½é‡æ¨¡æ‹ŸåŸå§‹æ„å›¾ï¼Œç»˜åˆ¶å®Œæ•´ nowcastï¼Œå±è”½ 1/2 æœˆå®é™…å€¼ã€‚
    """
    logger.info(f"\n[ç»˜å›¾å‡½æ•°] ç”Ÿæˆæœ€ç»ˆ Nowcasting å›¾: {filename}...")
    try:
        # --- ç¡®ä¿ç´¢å¼•ä¸º DatetimeIndex ---
        try:
            if not isinstance(final_nowcast_series.index, pd.DatetimeIndex):
                final_nowcast_series.index = pd.to_datetime(final_nowcast_series.index)
            if not isinstance(target_for_plot.index, pd.DatetimeIndex):
                target_for_plot.index = pd.to_datetime(target_for_plot.index)
        except Exception as e_index_conv:
             logger.warning(f"è­¦å‘Š: å°†ç´¢å¼•è½¬æ¢ä¸º DatetimeIndex æ—¶å‡ºé”™: {e_index_conv}")

        nowcast_col_name = 'Nowcast_Orig'
        target_col_name = target_for_plot.name if target_for_plot.name is not None else 'Actual'
        if target_col_name == nowcast_col_name: target_col_name = 'Observed_Value'

        # --- ğŸ”¥ ä¿®å¤ï¼šå‡†å¤‡ç»˜å›¾æ•°æ®ï¼Œç¡®ä¿æ˜¾ç¤ºå®Œæ•´çš„çœŸå®å€¼æ—¶é—´èŒƒå›´ ---
        # åˆ›å»ºåŒ…å«å®Œæ•´æ—¶é—´èŒƒå›´çš„DataFrameï¼Œä»¥çœŸå®å€¼çš„æ—¶é—´èŒƒå›´ä¸ºåŸºå‡†
        target_clean = target_for_plot.dropna()

        if not target_clean.empty:
            # æ‰©å±•æ—¶é—´èŒƒå›´ï¼šä»çœŸå®å€¼æœ€æ—©æ—¥æœŸåˆ°nowcastæœ€æ™šæ—¥æœŸ
            start_date = target_clean.index.min()
            end_date = max(target_clean.index.max(), final_nowcast_series.index.max())

            # åˆ›å»ºå®Œæ•´çš„å‘¨åº¦æ—¶é—´èŒƒå›´
            full_time_range = pd.date_range(start=start_date, end=end_date, freq='W-FRI')

            # ä»¥å®Œæ•´æ—¶é—´èŒƒå›´ä¸ºåŸºç¡€åˆ›å»ºç»˜å›¾DataFrame
            plot_df = pd.DataFrame(index=full_time_range)

            # æ·»åŠ nowcastæ•°æ®
            plot_df[nowcast_col_name] = final_nowcast_series.reindex(full_time_range)

            # æ·»åŠ çœŸå®å€¼æ•°æ®ï¼ˆä½¿ç”¨outer joinç¡®ä¿ä¸ä¸¢å¤±ä»»ä½•çœŸå®å€¼ï¼‰
            plot_df[target_col_name] = target_clean.reindex(full_time_range)

            logger.info(f"  âœ… ç»˜å›¾æ•°æ®èŒƒå›´: {start_date} åˆ° {end_date} ({len(full_time_range)} ä¸ªæ—¶é—´ç‚¹)")
            logger.info(f"  âœ… çœŸå®å€¼æ•°æ®ç‚¹: {target_clean.notna().sum()} ä¸ª")
            logger.info(f"  âœ… Nowcastæ•°æ®ç‚¹: {final_nowcast_series.notna().sum()} ä¸ª")
        else:
            # å›é€€åˆ°åŸæœ‰é€»è¾‘
            plot_df = final_nowcast_series.to_frame(name=nowcast_col_name)
            plot_df[target_col_name] = target_for_plot.rename(target_col_name)
            logger.warning("  âš ï¸ çœŸå®å€¼æ•°æ®ä¸ºç©ºï¼Œä½¿ç”¨å›é€€ç»˜å›¾é€»è¾‘")

        # ğŸ”¥ ä¿®å¤ï¼šç§»é™¤å±è”½1/2æœˆçœŸå®å€¼çš„é€»è¾‘ï¼Œæ˜¾ç¤ºæ‰€æœ‰çœŸå®å€¼
        # åŸæ¥çš„å±è”½é€»è¾‘å·²ç§»é™¤ï¼Œç°åœ¨æ˜¾ç¤ºæ‰€æœ‰æœˆä»½çš„çœŸå®å€¼
        logger.info("  æ˜¾ç¤ºæ‰€æœ‰æœˆä»½çš„çœŸå®è§‚æµ‹å€¼ç”¨äºç»˜å›¾ã€‚")

        if not plot_df.empty:
            plt.figure(figsize=(14, 7))
            nowcast_label = 'å‘¨åº¦ Nowcast (åŸå§‹æ°´å¹³)'
            actual_label = 'è§‚æµ‹å€¼ (åŸå§‹æ°´å¹³)'  # ğŸ”¥ ä¿®å¤ï¼šç§»é™¤"å±è”½1/2æœˆ"è¯´æ˜
            ylabel = 'å€¼ (åŸå§‹æ°´å¹³)'

            plt.plot(plot_df.index, plot_df[nowcast_col_name], label=nowcast_label, linestyle='-', alpha=0.8, color='blue')

            if target_col_name in plot_df.columns:
                target_to_plot_filtered = plot_df[target_col_name].dropna()
                plt.plot(target_to_plot_filtered.index, target_to_plot_filtered.values, label=actual_label, marker='o', linestyle='None', markersize=4, color='red')

            # --- æ ‡è®°éªŒè¯æœŸ --- 
            try:
                if isinstance(plot_df.index, pd.DatetimeIndex):
                    plot_start_date = plot_df.index.min()
                    plot_end_date = plot_df.index.max()
                    val_start_dt = pd.to_datetime(validation_start)
                    val_end_dt = pd.to_datetime(validation_end)
                    span_start = max(plot_start_date, val_start_dt)
                    span_end = min(plot_end_date, val_end_dt)
                    if span_start < span_end:
                        plt.axvspan(span_start, span_end, color='yellow', alpha=0.2, label='éªŒè¯æœŸ')
                    else:
                        plt.axvspan(val_start_dt, val_end_dt, color='yellow', alpha=0.2, label='éªŒè¯æœŸ (è¶…å‡ºèŒƒå›´)')
                else:
                     logger.warning("  è­¦å‘Š: ç»˜å›¾æ•°æ®ç´¢å¼•ä¸æ˜¯ DatetimeIndexï¼Œæ— æ³•æ ‡è®°éªŒè¯æœŸã€‚")
            except Exception as date_err:
                logger.warning(f"  è­¦å‘Šï¼šæ ‡è®°éªŒè¯æœŸæ—¶å‡ºé”™ - {date_err}")

            plt.title(title)
            plt.xlabel('æ—¥æœŸ')
            plt.ylabel(ylabel)
            plt.legend()
            plt.grid(True, linestyle='--', alpha=0.6)
            ax = plt.gca()
            ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=8, maxticks=15))
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()
            plt.savefig(filename)
            plt.close()
            logger.info(f"æœ€ç»ˆ Nowcasting å›¾å·²ä¿å­˜åˆ°: {filename}")
        else:
             logger.error("é”™è¯¯ï¼šæ— æ³•å‡†å¤‡ç”¨äºç»˜å›¾çš„æ•°æ® (plot_df ä¸ºç©º)ã€‚")
    except Exception as e:
        logger.error(f"[ç»˜å›¾å‡½æ•°] ç”Ÿæˆæˆ–ä¿å­˜æœ€ç»ˆ Nowcasting å›¾æ—¶å‡ºé”™: {e}", exc_info=True)
        try:
            plt.close() # å°è¯•å…³é—­ä»¥é˜²ä¸‡ä¸€
        except Exception: pass
# --- ç»“æŸé‡æ–°ç”Ÿæˆ --- 

def analyze_and_save_final_results(
    run_output_dir: str,
    timestamp_str: str,
    excel_output_path: str,
    # --- ä¿®æ”¹ï¼šæ·»åŠ /ç¡®è®¤æ‰€éœ€æ•°æ® --- 
    all_data_full: pd.DataFrame, # åŒ…å«åŸå§‹ç›®æ ‡å˜é‡
    final_data_processed: pd.DataFrame, # DFM è¾“å…¥æ•°æ® (å¤„ç†å, åŒ…å«éªŒè¯/é¢„æµ‹æœŸ) <--- ç¡®è®¤åŒ…å«éªŒè¯æœŸ
    final_target_mean_rescale: Optional[float], # ç”¨äº Nowcast åæ ‡å‡†åŒ–
    final_target_std_rescale: Optional[float], # ç”¨äº Nowcast åæ ‡å‡†åŒ–
    # --- ç»“æŸä¿®æ”¹ ---
    target_variable: str,
    final_dfm_results: Any, # DFM ç»“æœå¯¹è±¡, å‡å®šä¸º statsmodels results wrapper OR DFMEMResultsWrapper
    best_variables: List[str],
    best_params: Dict[str, Any],
    var_type_map: Dict[str, str], # ç”¨äºå˜é‡ç±»å‹æ³¨é‡Š (key éœ€è¦è§„èŒƒåŒ–!)
    total_runtime_seconds: float,
    training_start_date: str,   # <-- æ–°å¢å‚æ•°
    validation_start_date: str, # éªŒè¯æœŸå¼€å§‹æ—¥æœŸ
    validation_end_date: str,   # éªŒè¯æœŸç»“æŸæ—¥æœŸ
    train_end_date: str,        # è®­ç»ƒæœŸç»“æŸæ—¥æœŸ (ç”¨äºæŒ‡æ ‡è®¡ç®—åˆ†å‰²)
    factor_contributions: Optional[Dict[str, float]] = None, # å› å­å¯¹ç›®æ ‡æ–¹å·®è´¡çŒ®åº¦
    final_transform_log: Optional[Dict[str, str]] = None,   # æœ€ç»ˆå˜é‡è½¬æ¢æ—¥å¿—
    pca_results_df: Optional[pd.DataFrame] = None,          # PCA ç»“æœ DataFrame
    contribution_results_df: Optional[pd.DataFrame] = None, # å› å­è´¡çŒ®åº¦ç»“æœ DataFrame (Note: Appears unused in provided snippet, maybe used later?)
    var_industry_map: Optional[Dict[str, str]] = None,      # å˜é‡è¡Œä¸šæ˜ å°„ (key éœ€è¦è§„èŒƒåŒ–!)
    individual_r2_results: Optional[Dict[str, pd.DataFrame]] = None, # <<< æ·»åŠ æ–°å‚æ•°
    industry_r2_results: Optional[pd.Series] = None, # <<< æ·»åŠ è¡Œä¸š R2 å‚æ•°
    factor_industry_r2_results: Optional[Dict[str, pd.Series]] = None, # <<< æ·»åŠ å› å­-è¡Œä¸š R2 å‚æ•°
    factor_type_r2_results: Optional[Dict[str, pd.Series]] = None, # <<< æ–°å¢ï¼šæ·»åŠ å› å­-ç±»å‹ R2 å‚æ•°
    final_eigenvalues: Optional[np.ndarray] = None # <<< æ–°å¢ï¼šæ·»åŠ æœ€ç»ˆç‰¹å¾æ ¹å€¼å‚æ•°
) -> Tuple[Optional[pd.Series], Dict[str, Any]]: # <<< ä¿®æ”¹è¿”å›ç±»å‹æ³¨è§£
    """
    åˆ†ææœ€ç»ˆ DFM æ¨¡å‹ç»“æœå¹¶å°†å…¶ä¿å­˜åˆ° Excel æ–‡ä»¶ã€‚
    (ä¿®æ”¹ï¼šè®¡ç®—å®æ—¶æ»¤æ³¢ Nowcast y(t|t) å¹¶ç”¨äºåˆ†æ)
    (ä¿®æ”¹ï¼šå…¼å®¹ DFMEMResultsWrapper ç±»å‹)
    """
    logger.info(f"ğŸ”¥ å¼€å§‹åˆ†ææœ€ç»ˆç»“æœå¹¶å†™å…¥ Excel: {excel_output_path}")
    logger.info(f"ğŸ”¥ è¾“å…¥å‚æ•°æ£€æŸ¥:")
    logger.info(f"  - all_data_full: {type(all_data_full)} {getattr(all_data_full, 'shape', 'N/A')}")
    logger.info(f"  - final_data_processed: {type(final_data_processed)} {getattr(final_data_processed, 'shape', 'N/A')}")
    logger.info(f"  - target_variable: {target_variable}")
    logger.info(f"  - best_variables: {len(best_variables) if best_variables else 0} ä¸ªå˜é‡")
    logger.info(f"  - final_dfm_results: {type(final_dfm_results)}")
    logger.info(f"  - var_type_map: {len(var_type_map) if var_type_map else 0} ä¸ªæ˜ å°„")

    calculated_nowcast_orig = None # This will hold the FILTERED nowcast
    calculated_nowcast_smoothed_orig = None # Optional: To store the smoothed one
    metrics = {} # Initialize metrics dictionary
    aligned_df_for_metrics = None # Initialize aligned df
    loadings_df_final = None
    filtered_state = None # Initialize filtered_state

    try:
        # --- 1. æå–å¿…è¦ä¿¡æ¯ --- 
        # --- ä¿®æ”¹ï¼šæ£€æŸ¥ DFMEMResultsWrapper çš„å±æ€§ ---
        if not final_dfm_results or not hasattr(final_dfm_results, 'x_sm') or not hasattr(final_dfm_results, 'Lambda'):
             raise ValueError("æä¾›çš„ DFM ç»“æœå¯¹è±¡æ— æ•ˆæˆ–ç¼ºå°‘å¿…è¦çš„å±æ€§ (å¦‚ 'x_sm', 'Lambda')ã€‚")
        # --- ç»“æŸä¿®æ”¹ ---

        # --- ç§»é™¤ filter æ–¹æ³•æ£€æŸ¥ ---
        # # æ£€æŸ¥ filter æ–¹æ³•
        # # Adjusted check: Some results objects might have filter on the model object
        # results_filter_method = getattr(final_dfm_results, 'filter', None)
        # model_filter_method = getattr(getattr(final_dfm_results, 'model', None), 'filter', None)
        # if not callable(results_filter_method) and not callable(model_filter_method):
        #      raise AttributeError("DFM ç»“æœå¯¹è±¡åŠå…¶å…³è”çš„ model å¯¹è±¡éƒ½ç¼ºå°‘å¯è°ƒç”¨çš„ 'filter' æ–¹æ³•ã€‚")
        # --- ç»“æŸç§»é™¤ ---

        # --- ä¿®æ”¹ï¼šè°ƒæ•´è½½è·æå–é€»è¾‘ (Lambda æ˜¯å¤§å†™) ---
        if hasattr(final_dfm_results, 'Lambda') and final_dfm_results.Lambda is not None:
            loadings = final_dfm_results.Lambda # ä½¿ç”¨å¤§å†™ Lambda
        else:
            raise ValueError("æ— æ³•ä» DFM ç»“æœå¯¹è±¡è·å–è½½è·çŸ©é˜µ ('Lambda' å±æ€§)ã€‚")
        # --- ç»“æŸä¿®æ”¹ ---


        # æ£€æŸ¥è½½è· (Lambda) æ˜¯å¦æ˜¯ DataFrame æˆ– ndarray
        if not isinstance(loadings, pd.DataFrame):
            logger.warning(f"Loadings (Lambda) ä¸æ˜¯ DataFrame (ç±»å‹: {type(loadings)}), å°è¯•è½¬æ¢...")
            try:
                # å‡è®¾ loadings æ˜¯ ndarray [n_endog, k_factors]
                # éœ€è¦ endog_names å’Œ factor names
                # å°è¯•ä» best_params è·å–å› å­æ•°ï¼Œä» best_variables è·å–å˜é‡å
                k_factors = best_params.get('k_factors_final', None) # ä» best_params è·å–
                endog_names = best_variables # ä½¿ç”¨ best_variables

                if k_factors is None and isinstance(loadings, np.ndarray):
                    k_factors = loadings.shape[1] # ä» ndarray å½¢çŠ¶æ¨æ–­
                    logger.warning(f"æ— æ³•ä» best_params è·å–å› å­æ•°ï¼Œä» Lambda å½¢çŠ¶æ¨æ–­ä¸º {k_factors}")

                if k_factors is not None and endog_names is not None:
                    if isinstance(loadings, np.ndarray): # Check if it's an ndarray first
                        if loadings.shape == (len(endog_names), k_factors):
                            loadings = pd.DataFrame(loadings, index=endog_names, columns=[f'Factor{i+1}' for i in range(k_factors)])
                            logger.info("Loadings æˆåŠŸä» ndarray è½¬æ¢ä¸º DataFrameã€‚")
                        else: # <<< This else handles the shape mismatch for ndarray
                            raise ValueError(f"Loadings (ndarray) ç»´åº¦ ({getattr(loadings, 'shape', 'N/A')}) ä¸ endog_names ({len(endog_names)}) æˆ– k_factors ({k_factors}) ä¸åŒ¹é…ã€‚")
                    # Removed the problematic else: here, as it doesn't pair with the try
                else: # This else pairs with 'if k_factors is not None and endog_names is not None:'
                    raise ValueError("æ— æ³•è·å– k_factors æˆ– endog_names (best_variables) ä»¥è½¬æ¢ loadings ndarrayã€‚")
            # Correctly paired except block for the try block above
            except Exception as e_load_conv:
                logger.error(f"è½¬æ¢ Loadings ä¸º DataFrame å¤±è´¥: {e_load_conv}", exc_info=True)
                raise TypeError(f"è½¬æ¢ Loadings ä¸º DataFrame å¤±è´¥: {e_load_conv}")

        # Ensure loadings_df_final is assigned even if conversion wasn't needed or failed initially
        if not isinstance(loadings, pd.DataFrame):
            logger.warning("Loadings æ—¢ä¸æ˜¯é¢„æœŸçš„ DataFrame ä¹Ÿæœªèƒ½æˆåŠŸä» ndarray è½¬æ¢ã€‚")
            # Assign None or raise, depending on desired behavior if conversion fails.
            loadings_df_final = None
        else:
            loadings_df_final = loadings # Store the final loadings DataFrame

        # Add a check here to ensure loadings_df_final is usable AFTER the try-except block
        if loadings_df_final is None or not isinstance(loadings_df_final, pd.DataFrame):
            raise ValueError("æœªèƒ½è·å–æˆ–ç”Ÿæˆæœ‰æ•ˆçš„ Loadings DataFrameã€‚")


        # æ£€æŸ¥ç›®æ ‡å˜é‡æ˜¯å¦åœ¨è½½è·çŸ©é˜µä¸­
        if target_variable not in loadings_df_final.index:
            # Attempt to normalize target_variable name if it exists in logs but not loadings
            if final_transform_log and target_variable in final_transform_log:
                transformed_target_name = final_transform_log[target_variable]
                if transformed_target_name in loadings_df_final.index:
                    logger.warning(f"ç›®æ ‡å˜é‡ '{target_variable}' ä¸åœ¨è½½è·ç´¢å¼•ä¸­ï¼Œä½†è½¬æ¢åçš„åç§° '{transformed_target_name}' å­˜åœ¨ã€‚å°†ä½¿ç”¨è½¬æ¢åçš„åç§°ã€‚")
                    target_variable_in_loadings = transformed_target_name
                else: # <<< Correct indentation for this else
                    raise ValueError(f"DFM è½½è·çŸ©é˜µç´¢å¼•ä¸­ä¸åŒ…å«ç›®æ ‡å˜é‡ '{target_variable}' æˆ–å…¶è½¬æ¢ååç§° '{transformed_target_name}'ã€‚Index: {loadings_df_final.index}")
            else:
                raise ValueError(f"DFM è½½è·çŸ©é˜µ (Loadings DataFrame) çš„ç´¢å¼•ä¸­ä¸åŒ…å«ç›®æ ‡å˜é‡ '{target_variable}'ã€‚Index: {loadings_df_final.index}")
        else:
            target_variable_in_loadings = target_variable # Use original name
        
        target_loadings = loadings_df_final.loc[target_variable_in_loadings]

        # è·å–åæ ‡å‡†åŒ–å‚æ•°
        target_mean = final_target_mean_rescale
        target_std = final_target_std_rescale
        if target_mean is None or target_std is None:
            raise ValueError("ç¼ºå°‘ç”¨äºåæ ‡å‡†åŒ–çš„ç›®æ ‡å˜é‡å‡å€¼æˆ–æ ‡å‡†å·®ã€‚")


        # --- 2. è®¡ç®—å®æ—¶æ»¤æ³¢ Nowcast y(t|t) ---
        # --- ä¿®æ”¹ï¼šç›´æ¥ä½¿ç”¨ .x å±æ€§ä½œä¸ºæ»¤æ³¢çŠ¶æ€ ---
        # logger.info("è®¡ç®—å®æ—¶æ»¤æ³¢ Nowcast åºåˆ— y(t|t) (ä½¿ç”¨ final_dfm_results.x)...") #<-- æ—§æ—¥å¿—
        # --- ä¿®æ”¹ï¼šæ”¹ä¸ºè®¡ç®— y_{T(t)|t} ---
        logger.info("è®¡ç®—å¯¹æœˆåº•ç›®æ ‡çš„ Nowcast åºåˆ— y_{T(t)|t} (ä½¿ç”¨ final_dfm_results.x å’Œ A)...")
        try:
            # æ£€æŸ¥ final_dfm_results æ˜¯å¦æœ‰ .x å’Œ .A å±æ€§
            if not hasattr(final_dfm_results, 'x') or final_dfm_results.x is None:
                raise AttributeError("DFM ç»“æœå¯¹è±¡ç¼ºå°‘ 'x' å±æ€§ (å‡å®šä¸ºæ»¤æ³¢çŠ¶æ€) æˆ–å…¶å€¼ä¸º Noneã€‚")
            if not hasattr(final_dfm_results, 'A') or final_dfm_results.A is None:
                 raise AttributeError("DFM ç»“æœå¯¹è±¡ç¼ºå°‘ 'A' å±æ€§ (çŠ¶æ€è½¬ç§»çŸ©é˜µ) æˆ–å…¶å€¼ä¸º Noneã€‚")


            filtered_state_raw = final_dfm_results.x # ç›´æ¥è·å–æ»¤æ³¢çŠ¶æ€
            A = final_dfm_results.A # è·å–çŠ¶æ€è½¬ç§»çŸ©é˜µ

            # ç¡®ä¿ A æ˜¯ NumPy array
            if not isinstance(A, np.ndarray):
                raise TypeError(f"çŠ¶æ€è½¬ç§»çŸ©é˜µ A ä¸æ˜¯ NumPy array (Type: {type(A)})")
            n_factors = A.shape[0] # è·å–å› å­æ•°é‡

            # ğŸ”¥ å…³é”®ä¿®å¤ï¼šæ‰©å±•filtered_stateçš„ç´¢å¼•èŒƒå›´ä»¥åŒ…å«2025å¹´æ•°æ®
            # ç¡®ä¿ filtered_state æ˜¯ DataFrame å¹¶å…·æœ‰æ­£ç¡®çš„ç´¢å¼•
            if not isinstance(filtered_state_raw, pd.DataFrame):
                 if isinstance(filtered_state_raw, np.ndarray):
                     logger.warning("Filtered state (x) æ˜¯ ndarrayï¼Œå°è¯•ä½¿ç”¨ final_data_processed çš„ç´¢å¼•è½¬æ¢ä¸º DataFrameã€‚")
                     try:
                         # ğŸ”¥ ä¿®å¤ï¼šä½¿ç”¨all_data_fullçš„ç´¢å¼•èŒƒå›´è€Œä¸æ˜¯final_data_processed
                         # è¿™æ ·nowcastè®¡ç®—å¯ä»¥è¦†ç›–åˆ°2025å¹´
                         if all_data_full is not None and not all_data_full.empty:
                             # ä½¿ç”¨å®Œæ•´æ•°æ®çš„ç´¢å¼•èŒƒå›´
                             full_index = all_data_full.index
                             logger.info(f"ğŸ”¥ ä½¿ç”¨all_data_fullçš„å®Œæ•´ç´¢å¼•èŒƒå›´: {full_index.min()} åˆ° {full_index.max()}")

                             # æ£€æŸ¥çŠ¶æ€æ•°ç»„é•¿åº¦æ˜¯å¦åŒ¹é…è®­ç»ƒæ•°æ®
                             training_data_length = len(final_data_processed)
                             state_length = filtered_state_raw.shape[0]

                             if state_length == training_data_length:
                                 # çŠ¶æ€æ•°ç»„å¯¹åº”è®­ç»ƒæ•°æ®èŒƒå›´ï¼Œéœ€è¦æ‰©å±•åˆ°å®Œæ•´èŒƒå›´
                                 training_index = final_data_processed.index
                                 logger.info(f"çŠ¶æ€æ•°ç»„é•¿åº¦ {state_length} åŒ¹é…è®­ç»ƒæ•°æ®é•¿åº¦ {training_data_length}")
                                 logger.info(f"è®­ç»ƒæ•°æ®èŒƒå›´: {training_index.min()} åˆ° {training_index.max()}")

                                 # åˆ›å»ºæ‰©å±•çš„çŠ¶æ€æ•°ç»„
                                 extended_state = np.full((len(full_index), filtered_state_raw.shape[1]), np.nan)

                                 # å°†è®­ç»ƒæœŸçš„çŠ¶æ€å¡«å…¥å¯¹åº”ä½ç½®
                                 for i, date in enumerate(training_index):
                                     if date in full_index:
                                         full_idx = full_index.get_loc(date)
                                         extended_state[full_idx] = filtered_state_raw[i]

                                 # ğŸ”¥ å…³é”®ï¼šå¯¹äºè®­ç»ƒæœŸä¹‹åçš„æ—¶æœŸï¼Œä½¿ç”¨çŠ¶æ€è½¬ç§»çŸ©é˜µé¢„æµ‹
                                 # æ‰¾åˆ°è®­ç»ƒæœŸç»“æŸåçš„ç¬¬ä¸€ä¸ªæ—¶ç‚¹
                                 last_training_date = training_index.max()
                                 last_training_idx = full_index.get_loc(last_training_date)
                                 last_state = filtered_state_raw[-1]  # æœ€åä¸€ä¸ªè®­ç»ƒæœŸçŠ¶æ€

                                 # å‘å‰é¢„æµ‹çŠ¶æ€
                                 for i in range(last_training_idx + 1, len(full_index)):
                                     last_state = A @ last_state  # x_{t+1} = A * x_t
                                     extended_state[i] = last_state

                                 logger.info(f"ğŸ”¥ æˆåŠŸæ‰©å±•çŠ¶æ€åˆ°å®Œæ•´æ—¶é—´èŒƒå›´ï¼ŒåŒ…å«2025å¹´æ•°æ®")

                                 # ä½¿ç”¨æ‰©å±•çš„çŠ¶æ€å’Œå®Œæ•´ç´¢å¼•
                                 index_for_state = full_index
                                 filtered_state_raw = extended_state
                             else:
                                 # çŠ¶æ€æ•°ç»„é•¿åº¦ä¸åŒ¹é…ï¼Œä½¿ç”¨åŸæœ‰é€»è¾‘
                                 logger.warning(f"çŠ¶æ€æ•°ç»„é•¿åº¦ {state_length} ä¸åŒ¹é…è®­ç»ƒæ•°æ®é•¿åº¦ {training_data_length}ï¼Œä½¿ç”¨åŸæœ‰é€»è¾‘")
                                 index_for_state = final_data_processed.index[:len(filtered_state_raw)]
                         else:
                             # æ²¡æœ‰all_data_fullï¼Œä½¿ç”¨åŸæœ‰é€»è¾‘
                             logger.warning("all_data_fullä¸å¯ç”¨ï¼Œä½¿ç”¨final_data_processedçš„ç´¢å¼•")
                             index_for_state = final_data_processed.index[:len(filtered_state_raw)]

                         if len(index_for_state) != filtered_state_raw.shape[0]:
                              raise ValueError(f"ç”¨äºè½¬æ¢ filtered_state çš„ç´¢å¼•é•¿åº¦ ({len(index_for_state)}) ä¸çŠ¶æ€æ•°ç»„é•¿åº¦ ({filtered_state_raw.shape[0]}) ä¸åŒ¹é…ã€‚")

                         k_factors_state = filtered_state_raw.shape[1]
                         if k_factors_state != n_factors:
                              raise ValueError(f"æ»¤æ³¢çŠ¶æ€çš„å› å­æ•° ({k_factors_state}) ä¸çŠ¶æ€è½¬ç§»çŸ©é˜µ A çš„ç»´åº¦ ({n_factors}) ä¸åŒ¹é…ã€‚")

                         filtered_state = pd.DataFrame(filtered_state_raw,
                                                       index=index_for_state,
                                                       columns=[f'Factor{i+1}' for i in range(k_factors_state)])
                     except Exception as e_fs_conv:
                         raise TypeError(f"å°† filtered_state (x, ndarray) è½¬æ¢ä¸º DataFrame å¤±è´¥: {e_fs_conv}")
                 else:
                     raise TypeError(f"Filtered state (x) ä¸æ˜¯ Pandas DataFrame æˆ– Numpy Arrayã€‚ Type: {type(filtered_state_raw)}")
            else:
                 filtered_state = filtered_state_raw # Already a DataFrame
                 if filtered_state.shape[1] != n_factors:
                      raise ValueError(f"æ»¤æ³¢çŠ¶æ€ DataFrame çš„å› å­æ•° ({filtered_state.shape[1]}) ä¸çŠ¶æ€è½¬ç§»çŸ©é˜µ A çš„ç»´åº¦ ({n_factors}) ä¸åŒ¹é…ã€‚")

            # --- ç¡®ä¿ç´¢å¼•æ˜¯ DatetimeIndex (ç”¨äºæœˆä»½è®¡ç®—) ---
            if not pd.api.types.is_datetime64_any_dtype(filtered_state.index):
                 logger.warning("Filtered state index is not datetime, attempting conversion...")
                 try:
                     original_index = filtered_state.index
                     filtered_state.index = pd.to_datetime(filtered_state.index)
                     logger.info("Filtered state index converted to DatetimeIndex.")
                 except Exception as e_dt_conv:
                      logger.error(f"Failed to convert filtered state index to DatetimeIndex: {e_dt_conv}. Cannot calculate month-end targets.", exc_info=True)
                      raise TypeError(f"æ— æ³•å°† filtered_state ç´¢å¼•è½¬æ¢ä¸º DatetimeIndex: {e_dt_conv}") from e_dt_conv

            logger.debug(f"è·å–çš„ filtered_state (æ¥è‡ª .x) ç´¢å¼•èŒƒå›´: {filtered_state.index.min()} to {filtered_state.index.max()}")

            # --- è®¡ç®—å¯¹æœˆåº•ç›®æ ‡çš„é¢„æµ‹ y_{T(t)|t} ---
            nowcast_list_T = []
            index_list_T = []
            # å‡è®¾æ•°æ®é¢‘ç‡æ˜¯å‘¨äº” ('W-FRI')ï¼Œä¸ run_nowcasting_evolution.py ä¸€è‡´
            # é¢„è®¡ç®— A^k ä»¥æé«˜æ•ˆç‡ (å¦‚æœ k å€¼é‡å¤åº¦é«˜)
            A_pow_k_cache = {}

            for t, x_t in filtered_state.iterrows():
                try:
                    # 1. ç¡®å®š t æ‰€å±æœˆä»½çš„æœ€åé‚£ä¸ªå‘¨äº” T(t)
                    month_end_date = t + pd.offsets.MonthEnd(0) # è·å– t æ‰€åœ¨æœˆçš„æœ€åä¸€å¤©
                    # ä» t å¼€å§‹ï¼Œå‘åæ‰¾åˆ°å½“æœˆçš„æœ€åä¸€ä¸ªå‘¨äº”
                    # æ‰¾åˆ°è¯¥æœˆæ‰€æœ‰æ—¥æœŸ
                    all_days_in_month = pd.date_range(start=t.to_period('M').start_time, end=month_end_date, freq='D')
                    # ç­›é€‰å‡ºå‘¨äº”
                    fridays_in_month = all_days_in_month[all_days_in_month.dayofweek == 4]
                    if fridays_in_month.empty:
                        # å¦‚æœå½“æœˆæ²¡æœ‰å‘¨äº”ï¼ˆç†è®ºä¸Šä¸å¯èƒ½ï¼‰ï¼Œæˆ– t ä¹‹åæ²¡æœ‰å‘¨äº”ï¼Œåˆ™ T(t)=tï¼Œ k=0
                        target_date_T = t
                    else:
                        # å–å½“æœˆæœ€åä¸€ä¸ªå‘¨äº”ä½œä¸ºç›®æ ‡ T(t)
                        target_date_T = fridays_in_month[-1]
                        # --- é‡è¦ï¼šç¡®ä¿ T(t) >= t ---
                        if target_date_T < t:
                             # å¦‚æœ t å·²ç»æ˜¯å½“æœˆæœ€åä¸€ä¸ªå‘¨äº”ä¹‹åäº†ï¼Œåˆ™ç›®æ ‡å°±æ˜¯ t æœ¬èº«
                             target_date_T = t

                    # 2. è®¡ç®—é¢„æµ‹æ­¥æ•° k = T(t) - t (ä»¥å‘¨ä¸ºå•ä½)
                    if target_date_T == t:
                        k = 0
                    else:
                        days_diff = (target_date_T - t).days
                        # æ³¨æ„ï¼šè¿™é‡Œå‡è®¾äº†æ¯å‘¨éƒ½æœ‰æ•°æ®ç‚¹ï¼Œæˆ–è€…è¯´çŠ¶æ€è½¬ç§»æ˜¯æŒ‰å‘¨è¿›è¡Œçš„
                        # å¦‚æœæ•°æ®ä¸­æœ‰ç¼ºå¤±çš„å‘¨äº”ï¼Œç›´æ¥é™¤ä»¥ 7 å¯èƒ½ä¸ç²¾ç¡®
                        # æ›´ç¨³å¦¥çš„æ–¹å¼æ˜¯è®¡ç®—ä¸¤ä¸ªæ—¥æœŸåœ¨ filtered_state.index ä¸­çš„ä½ç½®å·®
                        try:
                             t_loc = filtered_state.index.get_loc(t)
                             T_loc = filtered_state.index.get_loc(target_date_T)
                             k = T_loc - t_loc
                             if k < 0: # Double check
                                 logger.warning(f"è®¡ç®—çš„æ­¥æ•° k < 0 ({k}) for t={t}, T(t)={target_date_T}. Setting k=0.")
                                 k = 0
                        except KeyError:
                             # å¦‚æœ T(t) ä¸åœ¨ç´¢å¼•ä¸­ï¼Œå›é€€åˆ°ä½¿ç”¨å¤©æ•°è®¡ç®—ï¼ˆå¯èƒ½æœ‰é£é™©ï¼‰
                             logger.warning(f"Target date T(t)={target_date_T} not in filtered_state index for t={t}. Calculating k based on days/7.")
                             k = int(round(days_diff / 7.0))
                             if k < 0: k = 0 # Ensure k is not negative

                    # 3. è®¡ç®— A^k
                    if k == 0:
                        A_pow_k = np.eye(n_factors)
                    elif k in A_pow_k_cache:
                        A_pow_k = A_pow_k_cache[k]
                    else:
                        try:
                            A_pow_k = np.linalg.matrix_power(A, k)
                            A_pow_k_cache[k] = A_pow_k # Cache the result
                        except np.linalg.LinAlgError as e_power:
                            logger.error(f"è®¡ç®— A^{k} (for t={t}, T(t)={target_date_T}) æ—¶å‘ç”Ÿçº¿æ€§ä»£æ•°é”™è¯¯: {e_power}. å°†ä½¿ç”¨å•ä½çŸ©é˜µä»£æ›¿ã€‚")
                            A_pow_k = np.eye(n_factors)

                    # 4. è®¡ç®—é¢„æµ‹çŠ¶æ€ x_{T(t)|t}
                    x_T_given_t = A_pow_k @ x_t.values # x_t æ˜¯ Series

                    # 5. è®¡ç®—æ ‡å‡†åŒ– Nowcast
                    # Ensure alignment between predicted state and target_loadings columns
                    # Assuming x_T_given_t is a numpy array [n_factors,]
                    # Assuming target_loadings is a pandas Series with factor names as index
                    nowcast_std = target_loadings.values @ x_T_given_t # Dot product

                    nowcast_list_T.append(nowcast_std)
                    index_list_T.append(t)

                except Exception as e_loop:
                    logger.error(f"è®¡ç®— t={t} çš„ y_(T(t)|t) æ—¶å‡ºé”™: {e_loop}", exc_info=True)
                    # é€‰æ‹©è·³è¿‡è¯¥ç‚¹æˆ–æ·»åŠ  NaN
                    nowcast_list_T.append(np.nan)
                    index_list_T.append(t)

            # --- ç»“æŸå¾ªç¯ ---

            # åˆ›å»ºæ ‡å‡†åŒ– Nowcast Series
            nowcast_forecast_standardized = pd.Series(nowcast_list_T, index=index_list_T)
            nowcast_forecast_standardized.name = "Nowcast_Forecast_Standardized"

            # åæ ‡å‡†åŒ–
            calculated_nowcast_orig = (nowcast_forecast_standardized * target_std) + target_mean
            # --- ä¿®æ”¹ï¼šæ›´æ–°åˆ—åä»¥åæ˜ è®¡ç®—æ–¹å¼ ---
            calculated_nowcast_orig.name = "Nowcast_ForecastToMonthEnd"
            logger.info("å¯¹æœˆåº•ç›®æ ‡çš„ Nowcast åºåˆ— y_{T(t)|t} è®¡ç®—å®Œæˆã€‚")

        except Exception as e_filter:
            logger.error(f"è®¡ç®—å¯¹æœˆåº•ç›®æ ‡çš„ Nowcast æ—¶å‡ºé”™: {e_filter}", exc_info=True)
            logger.warning("æ— æ³•è®¡ç®—å¯¹æœˆåº•ç›®æ ‡çš„ Nowcastï¼Œåç»­æŒ‡æ ‡å°†ä¸º N/Aã€‚")
            calculated_nowcast_orig = None # ç¡®ä¿å‡ºé”™æ—¶ä¸º None
        # --- ç»“æŸè®¡ç®— y_{T(t)|t} çš„ä¿®æ”¹ ---

        # --- (å¯é€‰) è®¡ç®—å¹³æ»‘ Nowcast ä»¥ä¾›æ¯”è¾ƒ (ä½¿ç”¨ .x_sm) ---
        logger.info("è®¡ç®—å¹³æ»‘ Nowcast åºåˆ— (ä½¿ç”¨ .x_sm, ç”¨äºå¯¹æ¯”)...")
        try:
            # --- ä¿®æ”¹ï¼šç›´æ¥ä½¿ç”¨ .x_sm ---
            smoothed_state_attr = None
            if hasattr(final_dfm_results, 'x_sm') and final_dfm_results.x_sm is not None:
                 smoothed_state_attr = final_dfm_results.x_sm
            # --- ç»“æŸä¿®æ”¹ ---

            if smoothed_state_attr is not None:
                factors_smoothed = smoothed_state_attr
                if not isinstance(factors_smoothed, pd.DataFrame):
                     if isinstance(factors_smoothed, np.ndarray):
                        k_factors_smooth = factors_smoothed.shape[1]
                        # Use index from final_data_processed, ensure length matches
                        index_for_smooth = final_data_processed.index[:len(factors_smoothed)]
                        factors_smoothed = pd.DataFrame(factors_smoothed,
                                                        index=index_for_smooth,
                                                        columns=[f'Factor{i+1}' for i in range(k_factors_smooth)])
                     else:
                         raise TypeError(f"Smoothed factors (x_sm) ä¸æ˜¯ DataFrame æˆ– Array (type: {type(factors_smoothed)})")

                # Ensure alignment for dot product
                common_factors_smooth = factors_smoothed.columns.intersection(target_loadings.index)
                if len(common_factors_smooth) != len(target_loadings):
                    logger.warning(f"å¹³æ»‘çŠ¶æ€å› å­ ({factors_smoothed.columns.tolist()}) ä¸ç›®æ ‡è½½è·å› å­ ({target_loadings.index.tolist()}) ä¸å®Œå…¨åŒ¹é…ã€‚ä»…ä½¿ç”¨å…±åŒå› å­: {common_factors_smooth.tolist()}")
                if common_factors_smooth.empty:
                     raise ValueError("å¹³æ»‘çŠ¶æ€ä¸ç›®æ ‡è½½è·ä¹‹é—´æ²¡æœ‰å…±åŒçš„å› å­åç§°ï¼Œæ— æ³•è®¡ç®—å¹³æ»‘ Nowcastã€‚")

                nowcast_smoothed_standardized = factors_smoothed[common_factors_smooth].dot(target_loadings[common_factors_smooth])
                calculated_nowcast_smoothed_orig = (nowcast_smoothed_standardized * target_std) + target_mean
                calculated_nowcast_smoothed_orig.name = "Nowcast_OriginalScale_Smoothed"
                logger.info("å¹³æ»‘ Nowcast åºåˆ—è®¡ç®—å®Œæˆã€‚")
            else:
                logger.warning("ç»“æœå¯¹è±¡ç¼ºå°‘ 'x_sm' å±æ€§æˆ–å…¶å€¼ä¸º Noneï¼Œæ— æ³•è®¡ç®—å¹³æ»‘ Nowcastã€‚")
                calculated_nowcast_smoothed_orig = None
        except Exception as e_smooth:
            logger.error(f"è®¡ç®—å¹³æ»‘ Nowcast æ—¶å‡ºé”™: {e_smooth}", exc_info=True)
            calculated_nowcast_smoothed_orig = None
        # --- ç»“æŸå¯é€‰å¹³æ»‘è®¡ç®— ---


        # --- <<< æ–°å¢è°ƒè¯•ï¼šæ£€æŸ¥ Nowcast (æ»¤æ³¢å) ç´¢å¼•èŒƒå›´ >>> ---
        calculated_nowcast_for_metrics = calculated_nowcast_orig # ä½¿ç”¨æ»¤æ³¢ç»“æœè¿›è¡Œè¯„ä¼°

        if calculated_nowcast_for_metrics is not None and not calculated_nowcast_for_metrics.empty:
            logger.info(f"  [DEBUG] Filtered Nowcast (for metrics) Index Range: {calculated_nowcast_for_metrics.index.min()} to {calculated_nowcast_for_metrics.index.max()}")
            # --- æ ¹æ® training_start_date è¿‡æ»¤ Nowcast (ç”¨äºæŠ¥å‘Šï¼Œè€ŒéæŒ‡æ ‡è®¡ç®—) ---
            calculated_nowcast_filtered_by_date = calculated_nowcast_for_metrics # Start with the full series
            if training_start_date:
                try:
                    start_dt = pd.to_datetime(training_start_date)
                    # Ensure index is datetime
                    if not pd.api.types.is_datetime64_any_dtype(calculated_nowcast_for_metrics.index):
                         logger.warning("Nowcast index is not datetime, attempting conversion for date filtering.")
                         temp_index = pd.to_datetime(calculated_nowcast_for_metrics.index, errors='coerce')
                         if temp_index.isna().any():
                             logger.error("Failed to convert Nowcast index to datetime for filtering. Skipping date filter.")
                             # Keep calculated_nowcast_filtered_by_date as the original unfiltered series
                         else:
                             calculated_nowcast_filtered_by_date = calculated_nowcast_for_metrics[temp_index >= start_dt]
                             logger.info(f"Filtered Nowcast series (for reporting) to start from {training_start_date}. Shape: {calculated_nowcast_filtered_by_date.shape}")
                    else:
                        calculated_nowcast_filtered_by_date = calculated_nowcast_for_metrics[calculated_nowcast_for_metrics.index >= start_dt]
                        logger.info(f"Filtered Nowcast series (for reporting) to start from {training_start_date}. Shape: {calculated_nowcast_filtered_by_date.shape}")

                except Exception as e_filter_date:
                    logger.warning(f"Could not filter Nowcast series by start date {training_start_date}: {e_filter_date}. Using original range for reporting.")
                    # Fallback to original if filtering fails - already assigned above
            else:
                logger.info("No training_start_date provided, using full Nowcast series for reporting.")
            # --- ç»“æŸè¿‡æ»¤ --- <<< Removed trailing backslash causing issues
        else:
            logger.info("  [DEBUG] Filtered Nowcast (for metrics) is None or empty.")
            calculated_nowcast_filtered_by_date = None # Ensure it's None if original is None/empty
        # --- <<< ç»“æŸæ–°å¢ >>> ---

        # --- 3. è®¡ç®—æœ€ç»ˆè¯„ä¼°æŒ‡æ ‡ (ä½¿ç”¨æ»¤æ³¢åçš„ Nowcast) ---
        logger.info("è®¡ç®—æœ€ç»ˆæ¨¡å‹çš„ IS/OOS RMSE å’Œ Hit Rate (ä½¿ç”¨ Filtered Nowcast å’Œ analysis_utils)...")

        # ğŸ”¥ ä¿®å¤ï¼šæ£€æŸ¥all_data_fullæ˜¯å¦ä¸ºæ¨¡æ‹Ÿæ•°æ®ï¼Œå¦‚æœæ˜¯åˆ™å°è¯•ä»UIæ•°æ®æºè·å–çœŸå®æ•°æ®
        original_target_series = None

        # æ£€æŸ¥all_data_fullæ˜¯å¦ä¸ºæ¨¡æ‹Ÿæ•°æ®ï¼ˆé€šè¿‡æ£€æŸ¥æ•°æ®ç‰¹å¾ï¼‰
        is_simulated_data = False
        if all_data_full is not None and not all_data_full.empty:
            # æ£€æŸ¥æ•°æ®æ˜¯å¦ä¸ºéšæœºç”Ÿæˆçš„æ¨¡æ‹Ÿæ•°æ®
            # æ¨¡æ‹Ÿæ•°æ®çš„ç‰¹å¾ï¼šæ•°æ®èŒƒå›´é€šå¸¸åœ¨-3åˆ°3ä¹‹é—´ï¼Œä¸”åˆ†å¸ƒæ¥è¿‘æ ‡å‡†æ­£æ€åˆ†å¸ƒ
            if target_variable in all_data_full.columns:
                target_data = all_data_full[target_variable].dropna()
                if len(target_data) > 10:
                    # æ£€æŸ¥æ•°æ®ç‰¹å¾
                    data_mean = abs(target_data.mean())
                    data_std = target_data.std()
                    data_range = target_data.max() - target_data.min()

                    # å¦‚æœæ•°æ®å‡å€¼æ¥è¿‘0ï¼Œæ ‡å‡†å·®æ¥è¿‘1ï¼Œä¸”èŒƒå›´åœ¨åˆç†çš„éšæœºæ•°èŒƒå›´å†…ï¼Œå¯èƒ½æ˜¯æ¨¡æ‹Ÿæ•°æ®
                    if data_mean < 0.5 and 0.8 < data_std < 1.2 and data_range < 8:
                        is_simulated_data = True
                        logger.warning(f"âš ï¸ æ£€æµ‹åˆ°all_data_fullå¯èƒ½ä¸ºæ¨¡æ‹Ÿæ•°æ® (å‡å€¼={data_mean:.3f}, æ ‡å‡†å·®={data_std:.3f}, èŒƒå›´={data_range:.3f})")

        # ğŸ”¥ æ–°å¢ï¼šå°è¯•ä»UI session_stateè·å–çœŸå®æ•°æ®
        real_data_source = None
        try:
            import streamlit as st
            if hasattr(st, 'session_state') and hasattr(st.session_state, 'dfm_prepared_data_df'):
                ui_data = st.session_state.dfm_prepared_data_df
                if ui_data is not None and not ui_data.empty and target_variable in ui_data.columns:
                    real_data_source = ui_data
                    logger.info(f"âœ… ä»UI session_stateè·å–åˆ°çœŸå®æ•°æ®æºï¼ŒåŒ…å« {len(ui_data)} è¡Œæ•°æ®")
                else:
                    logger.warning("UI session_stateä¸­çš„æ•°æ®ä¸å¯ç”¨æˆ–ä¸åŒ…å«ç›®æ ‡å˜é‡")
        except Exception as e:
            logger.warning(f"æ— æ³•ä»UI session_stateè·å–æ•°æ®: {e}")

        # é€‰æ‹©æ•°æ®æºä¼˜å…ˆçº§ï¼šçœŸå®UIæ•°æ® > all_data_full > final_data_processed
        if real_data_source is not None and target_variable in real_data_source.columns:
            original_target_series = real_data_source[target_variable]
            logger.info(f"âœ… ä»UIçœŸå®æ•°æ®æºè·å–ç›®æ ‡å˜é‡ '{target_variable}' ({len(original_target_series)} ä¸ªæ•°æ®ç‚¹)")
        elif all_data_full is not None and target_variable in all_data_full.columns and not is_simulated_data:
            original_target_series = all_data_full[target_variable]
            logger.info(f"ä» all_data_full è·å–ç›®æ ‡å˜é‡ '{target_variable}'")
        elif target_variable in final_data_processed.columns:
            original_target_series = final_data_processed[target_variable]
            logger.warning(f"åœ¨ all_data_full ä¸­æœªæ‰¾åˆ°ç›®æ ‡å˜é‡ '{target_variable}' æˆ–æ•°æ®ä¸ºæ¨¡æ‹Ÿæ•°æ®ï¼Œä» final_data_processed è·å–")
        else:
            logger.error(f"åœ¨æ‰€æœ‰æ•°æ®æºä¸­éƒ½æœªæ‰¾åˆ°ç›®æ ‡å˜é‡ '{target_variable}'")
            if all_data_full is not None:
                logger.error(f"all_data_full åˆ—: {list(all_data_full.columns)}")
            logger.error(f"final_data_processed åˆ—: {list(final_data_processed.columns)}")

        if original_target_series is None or original_target_series.empty:
            logger.error("æ— æ³•è®¡ç®—è¯„ä¼°æŒ‡æ ‡ï¼šæœªæ‰¾åˆ°æˆ–ä¸ºç©ºçš„åŸå§‹ç›®æ ‡åºåˆ—ã€‚")
            metrics = {k: 'N/A' for k in ['is_rmse', 'oos_rmse', 'is_mae', 'oos_mae', 'is_hit_rate', 'oos_hit_rate']}
        elif calculated_nowcast_for_metrics is None or calculated_nowcast_for_metrics.empty: # ä½¿ç”¨æ»¤æ³¢ç»“æœæ£€æŸ¥
             logger.error("æ— æ³•è®¡ç®—è¯„ä¼°æŒ‡æ ‡ï¼šæœªèƒ½æˆåŠŸè®¡ç®— Filtered Nowcast åºåˆ—ã€‚")
             metrics = {k: 'N/A' for k in ['is_rmse', 'oos_rmse', 'is_mae', 'oos_mae', 'is_hit_rate', 'oos_hit_rate']}
        else:
            # ç¡®ä¿ nowcast_series ä½¿ç”¨çš„æ˜¯æ»¤æ³¢ç»“æœ
            logger.debug(f"ä¼ é€’ç»™ calculate_metrics: Nowcast length={len(calculated_nowcast_for_metrics)}, Target length={len(original_target_series)}")
            metrics_raw, aligned_df_for_metrics = calculate_metrics_with_lagged_target(
                nowcast_series=calculated_nowcast_for_metrics, # <--- ç¡®è®¤ä½¿ç”¨æ»¤æ³¢ç»“æœ
                target_series=original_target_series.copy(),
                validation_start=validation_start_date,
                validation_end=validation_end_date,
                train_end=train_end_date,
                target_variable_name=target_variable # Pass original name here
            )
            metrics = metrics_raw
            logger.info(f"æœ€ç»ˆæ¨¡å‹è¯„ä¼°æŒ‡æ ‡ (åŸºäº Filtered Nowcast) è®¡ç®—å®Œæˆ: {metrics}")
            if aligned_df_for_metrics is not None:
                logger.debug(f"Aligned dataframe for metrics created with shape: {aligned_df_for_metrics.shape}")
            else:
                 logger.warning("Aligned dataframe for metrics was not created.")

            
        # --- 4. å‡†å¤‡å†™å…¥ Excel --- 
        # ... (åŠ è½½å…ƒæ•°æ®éƒ¨åˆ†) ...
        logger.info(f"å‡†å¤‡å†™å…¥ Excel æ–‡ä»¶: {excel_output_path}")
        metadata_loaded = {} # Initialize empty dict
        try:
             metadata_path = os.path.join(run_output_dir, 'final_dfm_metadata.pkl')
             if os.path.exists(metadata_path):
                 with open(metadata_path, 'rb') as f_meta:
                     metadata_loaded = pickle.load(f_meta) # Corrected indentation
                     logger.info(f"æˆåŠŸåŠ è½½å…ƒæ•°æ®: {metadata_path}") # Corrected indentation
             else:
                  logger.warning(f"å…ƒæ•°æ®æ–‡ä»¶æœªæ‰¾åˆ°: {metadata_path}")
        except Exception as e_meta_load:
            logger.warning(f"æ— æ³•åŠ è½½å…ƒæ•°æ®æ–‡ä»¶ä»¥è·å– Stage 1 ä¿¡æ¯: {e_meta_load}. éƒ¨åˆ† Summary ä¿¡æ¯å°†ä¸º N/Aã€‚")
            # metadata_loaded remains {}

        original_data_file = metadata_loaded.get('original_data_file', 'N/A')

            
        # --- åˆ›å»ºè§£é‡Šå†…å®¹ (ä¿æŒä¸å˜) ---
        explanation_sheet_name = "æŒ‡æ ‡è§£é‡Š"
        metric_explanations = [
            ["æŒ‡æ ‡", "ä¸­æ–‡å«ä¹‰", "è¯´æ˜"],
            ["æœ€ç»ˆ IS RMSE (Filtered)", "æ ·æœ¬å†…å‡æ–¹æ ¹è¯¯å·® (åŸºäºæ»¤æ³¢å€¼)", "è¡¡é‡æ¨¡å‹åœ¨è®­ç»ƒæœŸé—´å®æ—¶é¢„æµ‹(æ»¤æ³¢)å€¼ä¸å®é™…å€¼è¯¯å·®çš„å¹³å‡å¤§å°ã€‚è¶Šå°è¡¨ç¤ºè®­ç»ƒæ‹Ÿåˆè¶Šå¥½ã€‚"],
            ["æœ€ç»ˆ OOS RMSE (Filtered)", "æ ·æœ¬å¤–å‡æ–¹æ ¹è¯¯å·® (åŸºäºæ»¤æ³¢å€¼)", "è¡¡é‡æ¨¡å‹åœ¨éªŒè¯æœŸé—´å®æ—¶é¢„æµ‹(æ»¤æ³¢)å€¼ä¸å®é™…å€¼è¯¯å·®çš„å¹³å‡å¤§å°ã€‚è¶Šå°è¡¨ç¤ºæ¨¡å‹æ³›åŒ–èƒ½åŠ›è¶Šå¼ºã€‚"],
            ["æœ€ç»ˆ IS MAE (Filtered)", "æ ·æœ¬å†…å¹³å‡ç»å¯¹è¯¯å·® (åŸºäºæ»¤æ³¢å€¼)", "è¡¡é‡æ¨¡å‹åœ¨è®­ç»ƒæœŸé—´å®æ—¶é¢„æµ‹(æ»¤æ³¢)å€¼ä¸å®é™…å€¼è¯¯å·®ç»å¯¹å€¼çš„å¹³å‡å¤§å°ã€‚"],
            ["æœ€ç»ˆ OOS MAE (Filtered)", "æ ·æœ¬å¤–å¹³å‡ç»å¯¹è¯¯å·® (åŸºäºæ»¤æ³¢å€¼)", "è¡¡é‡æ¨¡å‹åœ¨éªŒè¯æœŸé—´å®æ—¶é¢„æµ‹(æ»¤æ³¢)å€¼ä¸å®é™…å€¼è¯¯å·®ç»å¯¹å€¼çš„å¹³å‡å¤§å°ã€‚"],
            ["æœ€ç»ˆ IS Hit Rate (%) (Filtered)", "æ ·æœ¬å†…å‘½ä¸­ç‡ (åŸºäºæ»¤æ³¢å€¼)", "è¡¡é‡æ¨¡å‹åœ¨è®­ç»ƒæœŸé—´é¢„æµ‹ç›®æ ‡å˜é‡å®æ—¶å˜åŒ–æ–¹å‘(æ¶¨/è·Œ)çš„å‡†ç¡®ç‡ã€‚"],
            ["æœ€ç»ˆ OOS Hit Rate (%) (Filtered)", "æ ·æœ¬å¤–å‘½ä¸­ç‡ (åŸºäºæ»¤æ³¢å€¼)", "è¡¡é‡æ¨¡å‹åœ¨éªŒè¯æœŸé—´é¢„æµ‹ç›®æ ‡å˜é‡å®æ—¶å˜åŒ–æ–¹å‘çš„å‡†ç¡®ç‡ã€‚"],
            ["å› å­å¯¹å•å˜é‡ RÂ²", "(è§ R2 Combined Sheet)", "è¡¨ç¤ºå•ä¸ªå› å­èƒ½è§£é‡Šå•ä¸ªé¢„æµ‹å˜é‡å˜åŠ¨çš„ç™¾åˆ†æ¯”ã€‚"],
            ["è¡Œä¸šæ±‡æ€» RÂ²", "(è§ R2 Combined Sheet)", "è¡¨ç¤ºæ‰€æœ‰å› å­å…±åŒèƒ½è§£é‡Šç‰¹å®šè¡Œä¸šå†…æ‰€æœ‰å˜é‡æ•´ä½“å˜åŠ¨çš„ç™¾åˆ†æ¯”ã€‚"],
            ["å› å­å¯¹è¡Œä¸šæ±‡æ€» RÂ²", "(è§ R2 Combined Sheet)", "è¡¨ç¤ºå•ä¸ªå› å­èƒ½è§£é‡Šç‰¹å®šè¡Œä¸šå†…æ‰€æœ‰å˜é‡æ•´ä½“å˜åŠ¨çš„ç™¾åˆ†æ¯”ã€‚"],
            ["å› å­å¯¹ç±»å‹æ±‡æ€» RÂ²", "(è§ R2 Combined Sheet)", "è¡¨ç¤ºå•ä¸ªå› å­èƒ½è§£é‡Šç‰¹å®šç±»å‹(e.g., M/Q/W)å˜é‡æ•´ä½“å˜åŠ¨çš„ç™¾åˆ†æ¯”ã€‚"] # Added Type R2 explanation
        ]
        sheet_explanations = [
            ["Sheet åç§°", "ä¸»è¦ç›®çš„"],
            ["Summary", "æä¾›æ¨¡å‹è¿è¡Œæ€»ä½“æ¦‚è§ˆä¿¡æ¯(å…³é”®å‚æ•°ã€è¯„ä¼°æŒ‡æ ‡ã€è¿è¡Œæ—¶é—´ç­‰)ï¼Œå¹¶é™„åŠ å› å­è´¡çŒ®åº¦ã€PCAæ–¹å·®è§£é‡Šã€‚"],
            ["Aligned Nowcast vs Target", "å±•ç¤ºæœ€ç»ˆæ»¤æ³¢Nowcasté¢„æµ‹å€¼ä¸å¯¹é½åå®é™…ç›®æ ‡å€¼çš„å¯¹æ¯”ï¼Œç”¨äºç›´è§‚è¯„ä¼°é¢„æµ‹æ•ˆæœã€‚"], # Updated description
            ["Factor Time Series", "å±•ç¤ºæ¨¡å‹ä¼°è®¡å‡ºçš„æ»¤æ³¢åå› å­æ—¶é—´åºåˆ—ï¼Œä»¥åŠæœ€ç»ˆçš„æ»¤æ³¢Nowcaståºåˆ—(å¯é€‰åŒ…å«å¹³æ»‘Nowcastå¯¹æ¯”)ã€‚"], # Updated description
            ["R2 Analysis Combined", "ç»¼åˆå±•ç¤ºRÂ²åˆ†æç»“æœ(å› å­å¯¹å˜é‡ã€è¡Œä¸šã€ç±»å‹)ã€‚"],
            ["Variables and Loadings", "åˆ—å‡ºæ¨¡å‹ä½¿ç”¨çš„æœ€ç»ˆé¢„æµ‹å˜é‡åŠå…¶ç±»å‹/è¡Œä¸šä¿¡æ¯ï¼Œä»¥åŠæ¯ä¸ªå˜é‡åœ¨æ¯ä¸ªå› å­ä¸Šçš„è½½è·ã€‚"],
            ["æŒ‡æ ‡è§£é‡Š", "è§£é‡Š Excel ä¸­å„ Sheet å’ŒæŒ‡æ ‡çš„å«ä¹‰ã€‚"] # Added self-explanation
        ]
        explanation_df_metrics = pd.DataFrame(metric_explanations[1:], columns=metric_explanations[0])
        explanation_df_sheets = pd.DataFrame(sheet_explanations[1:], columns=sheet_explanations[0])
            
        with pd.ExcelWriter(excel_output_path, engine='openpyxl', mode='w') as writer:
            # --- Sheet 0: æŒ‡æ ‡è§£é‡Š (ä¿æŒä¸å˜) ---
            logger.info(f"  æ­£åœ¨å†™å…¥ '{explanation_sheet_name}' Sheet...")
            try:
                explanation_df_metrics.to_excel(writer, sheet_name=explanation_sheet_name, startrow=1, index=False)
                startrow_sheets = explanation_df_metrics.shape[0] + 3 
                explanation_df_sheets.to_excel(writer, sheet_name=explanation_sheet_name, startrow=startrow_sheets, index=False)
                worksheet_exp = writer.sheets[explanation_sheet_name]
                title_metrics = worksheet_exp.cell(row=1, column=1, value="æŒ‡æ ‡è§£é‡Š")
                title_metrics.font = Font(bold=True)
                title_sheets = worksheet_exp.cell(row=startrow_sheets + 1, column=1, value="Sheet ç”¨é€”è¯´æ˜")
                title_sheets.font = Font(bold=True)
                format_excel_sheet(worksheet_exp, column_widths={'A': 35, 'B': 30, 'C': 70})
                logger.info(f"  '{explanation_sheet_name}' Sheet å†™å…¥å®Œæˆã€‚")
            except Exception as e_exp_write:
                logger.error(f"ğŸ”¥ å†™å…¥ '{explanation_sheet_name}' æ—¶å‡ºé”™: {e_exp_write}", exc_info=True)
                # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨ï¼Œåªè®°å½•é”™è¯¯
                # try: pd.DataFrame([f"Error writing explanations: {e_exp_write}"]).to_excel(writer, sheet_name=explanation_sheet_name, index=False, header=False)
                # except: pass


            # --- Sheet 1: Summary (ä½¿ç”¨æ–° metrics) ---
            logger.info("  æ­£åœ¨å†™å…¥ 'Summary' Sheet...")
            # --- ç§»é™¤/æ³¨é‡Šæ‰é”™è¯¯çš„ä»£ç è¡Œ ---
            # summary_ws = writer.sheets['Summary'] #<-- æ­¤è¡Œå¯¼è‡´ KeyErrorï¼Œå› ä¸º sheet æ­¤æ—¶å°šæœªåˆ›å»º
            # bold_format = writer.book.add_format({'bold': True}) # è¿™ä¸ªæ–¹å¼å±äº xlsxwriterï¼Œä¸æ˜¯ openpyxl

            # --- å†™å…¥åŸºæœ¬ä¿¡æ¯ ---
            summary_data = {
                'Parameter': [
                    'Target Variable', 'Training Period Start', 'Training Period End', 'Validation Period Start', 'Validation Period End',
                     'Selected Variables Count', 'Selected Factors (Final)', 'Factor Order p (Final)',
                     'Model Estimation Runtime (seconds)',
                     # 'Best Tuning Avg Hit Rate (%)', #<-- ç§»é™¤
                     # 'Best Tuning Avg MAE', #<-- ç§»é™¤
                     'Final Model IS RMSE', 'Final Model OOS RMSE',
                     'Final Model IS MAE', 'Final Model OOS MAE', # <-- æ·»åŠ  MAE
                     'Final Model IS Hit Rate (%)', 'Final Model OOS Hit Rate (%)'
                     ],
                'Value': [
                    target_variable, training_start_date, train_end_date, validation_start_date, validation_end_date,
                    len(best_variables), best_params.get('k_factors_final', 'N/A'), best_params.get('factor_order', 'N/A'),
                     f"{total_runtime_seconds:.2f}" if total_runtime_seconds is not None else 'N/A',
                     # format_metric_pct(best_avg_hit_rate_tuning) if best_avg_hit_rate_tuning is not None else 'N/A', #<-- ç§»é™¤
                     # format_metric(best_avg_rmse_tuning) if best_avg_rmse_tuning is not None else 'N/A', #<-- ç§»é™¤
                     format_metric(metrics.get('is_rmse')), format_metric(metrics.get('oos_rmse')),
                     format_metric(metrics.get('is_mae')), format_metric(metrics.get('oos_mae')), # <-- æ·»åŠ  MAE
                     format_metric_pct(metrics.get('is_hit_rate')), format_metric_pct(metrics.get('oos_hit_rate'))
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_start_row = 1 # Excel is 1-based, pandas to_excel startrow is 0-based for header

            # --- é¦–å…ˆå†™å…¥æ•°æ®ï¼Œè¿™ä¼šåˆ›å»º 'Summary' sheet ---
            summary_df.to_excel(writer, sheet_name='Summary', startrow=summary_start_row, index=False, header=True) # Ensure header is written

            # --- ç„¶åï¼Œè·å–å·²åˆ›å»ºçš„ sheet å¯¹è±¡ (ä½¿ç”¨ openpyxl çš„ workbook) ---
            summary_ws = writer.book['Summary'] # è·å– worksheet å¯¹è±¡

            # --- å†™å…¥æ ‡é¢˜ (ä½¿ç”¨ openpyxl çš„æ–¹å¼è®¾ç½®å­—ä½“) ---
            # pandas to_excel å†™å…¥æ—¶ï¼Œstartrow=1 è¡¨ç¤ºæ•°æ®ä»ç¬¬ 2 è¡Œå¼€å§‹ (ç¬¬ 1 è¡Œæ˜¯è¡¨å¤´)
            # å› æ­¤ï¼Œæ ‡é¢˜åº”è¯¥å†™åœ¨ Excel çš„ç¬¬ 1 è¡Œ (row=1)
            title_row_excel = 1
            title_col_excel = 1
            title_cell = summary_ws.cell(row=title_row_excel, column=title_col_excel, value="Model Summary and Performance")
            # ç¡®ä¿ Font å·²ä» openpyxl.styles å¯¼å…¥
            try:
                title_cell.font = Font(bold=True)
            except ImportError:
                 logger.warning("æ— æ³•å¯¼å…¥ openpyxl.styles.Fontï¼Œæ ‡é¢˜å°†ä¸ä¼šåŠ ç²—ã€‚")


            # --- å†™å…¥ PCA ç›¸å…³ä¿¡æ¯ (å¦‚æœæä¾›) ---
            # è®¡ç®— PCA è¡¨æ ¼åœ¨ Excel ä¸­çš„èµ·å§‹è¡Œå·
            # æ•°æ®åœ¨ summary_start_row+1 å¼€å§‹ï¼Œå…± len(summary_df) è¡Œæ•°æ® + 1 è¡Œè¡¨å¤´
            pca_title_start_row_excel = title_row_excel + len(summary_df) + 1 + 2 # Title row + data rows + header row + 2 blank rows
            pca_data_start_row_pandas = pca_title_start_row_excel # pandas startrow is the row *before* the header

            # --- ä¿®æ”¹ï¼šå¢åŠ å¯¹ pca_results_df ç±»å‹çš„æ£€æŸ¥ ---
            if isinstance(pca_results_df, pd.DataFrame) and not pca_results_df.empty:
            # --- ç»“æŸä¿®æ”¹ ---
                # --- æ£€æŸ¥ pca_results_df æ˜¯å¦åŒ…å« 'Eigenvalue' åˆ— ---
                # <<< ä¿®æ”¹ï¼šç¡®ä¿æ£€æŸ¥çš„åˆ—åä¸ analysis_utils.py ä¸­æ·»åŠ çš„å®Œå…¨ä¸€è‡´ >>>
                if 'ç‰¹å¾å€¼ (Eigenvalue)' in pca_results_df.columns:
                     # å‡†å¤‡æ˜¾ç¤ºç”¨çš„DataFrame
                     pca_results_df_display = pca_results_df.copy()
                     # æ ¼å¼åŒ–æ•°å€¼åˆ—
                     pca_results_df_display['ç‰¹å¾å€¼ (Eigenvalue)'] = pca_results_df_display['ç‰¹å¾å€¼ (Eigenvalue)'].apply(lambda x: format_metric(x, precision=4))
                     pca_results_df_display['è§£é‡Šæ–¹å·® (%)'] = pca_results_df_display['è§£é‡Šæ–¹å·® (%)'].apply(lambda x: format_metric_pct(x, precision=2))
                     pca_results_df_display['ç´¯è®¡è§£é‡Šæ–¹å·® (%)'] = pca_results_df_display['ç´¯è®¡è§£é‡Šæ–¹å·® (%)'].apply(lambda x: format_metric_pct(x, precision=2))

                     # --- å†™å…¥ PCA è¡¨æ ¼ ---
                     # <<< ä¿®æ”¹ï¼šä½¿ç”¨ format_excel_sheet æ”¹è¿›æ ¼å¼åŒ– >>>
                     # pca_results_df_display.to_excel(writer, sheet_name='Summary', startrow=pca_data_start_row_pandas, index=True, index_label="Factor") # pandas startrow
                     pca_results_df_display.to_excel(writer, sheet_name='Summary', startrow=pca_data_start_row_pandas, index=False) # ä¸å†™å…¥ç´¢å¼•

                     # --- å†™å…¥ PCA æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                     pca_title_cell = summary_ws.cell(row=pca_title_start_row_excel, column=1, value="PCA Results")
                     try: pca_title_cell.font = Font(bold=True)
                     except NameError: pass # Font not imported
                     
                     # --- æ ¼å¼åŒ–å†™å…¥çš„ PCA è¡¨æ ¼ ---
                     try:
                        format_excel_sheet(summary_ws,
                                            column_widths={'A': 15, 'B': 15, 'C': 18, 'D': 20}, # è°ƒæ•´åˆ—å®½
                                            # number_formats={'B': '0.00%', 'C': '0.00%', 'D': '0.0000'} # å°è¯•æ•°å­—æ ¼å¼
                                            )
                     except Exception as e_format:
                        logger.warning(f"æ ¼å¼åŒ– PCA è¡¨æ ¼æ—¶å‡ºé”™: {e_format}")
                     # <<< ç»“æŸä¿®æ”¹ >>>
                else:
                    logger.warning("PCA DataFrame provided but missing 'ç‰¹å¾å€¼ (Eigenvalue)' column. Skipping PCA table write.") # <-- æ›´æ–°è­¦å‘Šä¿¡æ¯ä¸­çš„åˆ—å
                    # --- å†™å…¥å ä½æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                    pca_title_cell = summary_ws.cell(row=pca_title_start_row_excel, column=1, value="PCA Results: (Missing 'ç‰¹å¾å€¼ (Eigenvalue)' Column)") # <-- æ›´æ–°å ä½ç¬¦æ–‡æœ¬
                    try: pca_title_cell.font = Font(bold=True)
                    except NameError: pass
            else:
                # --- ä¿®æ”¹ï¼šæ˜ç¡®åŒºåˆ† None å’Œé”™è¯¯ç±»å‹ --- 
                if pca_results_df is None:
                    logger.info("æœªæä¾› PCA ç»“æœ (pca_results_df is None)ï¼Œè·³è¿‡å†™å…¥ PCA è¡¨æ ¼ã€‚")
                else:
                    logger.warning(f"æä¾›çš„ pca_results_df ä¸æ˜¯æœ‰æ•ˆçš„ DataFrame (ç±»å‹: {type(pca_results_df)})ï¼Œè·³è¿‡å†™å…¥ PCA è¡¨æ ¼ã€‚")
                # --- ç»“æŸä¿®æ”¹ ---
                # --- å†™å…¥å ä½æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                pca_title_cell = summary_ws.cell(row=pca_title_start_row_excel, column=1, value="PCA Results: (Not Provided or Invalid)") # <-- ä¿®æ”¹å ä½ç¬¦æ–‡æœ¬
                try: pca_title_cell.font = Font(bold=True)
                except NameError: pass


            # --- æ–°å¢ï¼šå†™å…¥æœ€ç»ˆç‰¹å¾æ ¹å€¼ (å¦‚æœæä¾›) ---
            # è®¡ç®— Eigenvalue è¡¨æ ¼çš„èµ·å§‹è¡Œ
            # --- ä¿®æ”¹ï¼šæ ¹æ® pca_results_df_display æ˜¯å¦å­˜åœ¨æ¥è®¡ç®— pca_table_rows ---
            pca_table_rows = (len(pca_results_df_display) + 1) if 'pca_results_df_display' in locals() and isinstance(pca_results_df_display, pd.DataFrame) else 1 # Data rows + header or just placeholder title row
            # --- ç»“æŸä¿®æ”¹ ---
            eigenvalue_title_start_row_excel = pca_title_start_row_excel + pca_table_rows + 1 # After PCA table/placeholder + 1 blank row
            eigenvalue_data_start_row_pandas = eigenvalue_title_start_row_excel

            if final_eigenvalues is not None and len(final_eigenvalues) > 0:
                 eigen_df = pd.DataFrame({'Eigenvalue': final_eigenvalues})
                 eigen_df.index.name = 'Component'
                 eigen_df['Eigenvalue'] = eigen_df['Eigenvalue'].apply(lambda x: format_metric(x, precision=6))
                 eigen_df.to_excel(writer, sheet_name='Summary', startrow=eigenvalue_data_start_row_pandas, index=True) # pandas startrow
                 # --- å†™å…¥ç‰¹å¾æ ¹æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                 eigen_title_cell = summary_ws.cell(row=eigenvalue_title_start_row_excel, column=1, value="Final Model Eigenvalues (State Transition Matrix)")
                 try: eigen_title_cell.font = Font(bold=True)
                 except NameError: pass
            else:
                 logger.warning("æœªæä¾›ç‰¹å¾æ ¹å€¼ (final_eigenvalues)ï¼Œæœªæ·»åŠ ç‰¹å¾æ ¹åˆ—ã€‚")
                 # --- å†™å…¥å ä½æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                 eigen_title_cell = summary_ws.cell(row=eigenvalue_title_start_row_excel, column=1, value="Final Model Eigenvalues: (Not Provided)")
                 try: eigen_title_cell.font = Font(bold=True)
                 except NameError: pass
            # --- ç»“æŸæ–°å¢ ---


            # --- å†™å…¥å˜é‡è½¬æ¢æ—¥å¿— (å¦‚æœæä¾›) ---
            # è®¡ç®— Transform Log è¡¨æ ¼çš„èµ·å§‹è¡Œ
            eigenvalue_table_rows = (len(eigen_df) + 1) if final_eigenvalues is not None and len(final_eigenvalues) > 0 else 1
            transform_title_start_row_excel = eigenvalue_title_start_row_excel + eigenvalue_table_rows + 1 # After Eigenvalue table/placeholder + 1 blank row
            transform_data_start_row_pandas = transform_title_start_row_excel

            if final_transform_log:
                # --- ä¿®æ”¹ï¼šæå–æ›´æœ‰æ„ä¹‰çš„è½¬æ¢ä¿¡æ¯ (status) --- 
                applied_transforms_list = []
                # Iterate through the original log to maintain order if needed
                for original_var, transform_details in final_transform_log.items():
                    # Check if a *meaningful* transformation occurred
                    # We might want to show all variables listed, even if status is 'level'?
                    # Let's refine the logic: show entries where status is not 'level' or where the key!=value if value is simple string
                    
                    is_dict_details = isinstance(transform_details, dict)
                    status = transform_details.get('status', 'N/A') if is_dict_details else 'N/A' # Extract status if it's a dict
                    transformed_name = transform_details.get('transformed_name', original_var) if is_dict_details else str(transform_details)

                    # Decide if this transformation is worth showing
                    # Show if: status is not 'level', OR if details are not a dict and original name != transformed name
                    if status != 'level' or (not is_dict_details and original_var != transformed_name):
                        applied_transforms_list.append({
                            'Original Variable': original_var,
                            'Transformation Status': status if status != 'N/A' else 'Renamed/Other',
                            'Resulting Name (if changed)': transformed_name if original_var != transformed_name else original_var
                        })
                        
                if applied_transforms_list:
                    transform_df = pd.DataFrame(applied_transforms_list)[['Original Variable', 'Transformation Status', 'Resulting Name (if changed)']] # Ensure column order
                    # --- ç»“æŸä¿®æ”¹ ---
                    transform_df.to_excel(writer, sheet_name='Summary', startrow=transform_data_start_row_pandas, index=False) # pandas startrow
                    # --- å†™å…¥è½¬æ¢æ—¥å¿—æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                    transform_title_cell = summary_ws.cell(row=transform_title_start_row_excel, column=1, value="Applied Variable Transformations") # openpyxl row
                else:
                    logger.info("æ²¡æœ‰é€‚ç”¨çš„å˜é‡è½¬æ¢æ—¥å¿—éœ€è¦å†™å…¥ Summaryã€‚")
                    # --- å†™å…¥å ä½æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                    transform_title_cell = summary_ws.cell(row=transform_title_start_row_excel, column=1, value="Applied Variable Transformations: (None Applied)")
                    try: transform_title_cell.font = Font(bold=True)
                    except NameError: pass
            else:
                logger.info("æœªæä¾›å˜é‡è½¬æ¢æ—¥å¿—ã€‚")
                # --- å†™å…¥å ä½æ ‡é¢˜ (ä½¿ç”¨ summary_ws) ---
                transform_title_cell = summary_ws.cell(row=transform_title_start_row_excel, column=1, value="Applied Variable Transformations: (Log Not Provided)")
                try: transform_title_cell.font = Font(bold=True)
                except NameError: pass


            # --- è°ƒæ•´ Summary Sheet åˆ—å®½ ---
            # (ç¡®ä¿ format_excel_sheet èƒ½æ­£ç¡®å¤„ç† worksheet å¯¹è±¡)
            format_excel_sheet(summary_ws, column_widths={'A': 45, 'B': 45}) # Increased width slightly

            logger.info("  'Summary' Sheet å†™å…¥å®Œæˆã€‚")
            # --- ç»“æŸ Summary Sheet ---


            # --- Sheet 2: Aligned Nowcast vs Target (ä½¿ç”¨æ»¤æ³¢ç»“æœ) ---
            # --- ä¿®æ”¹ï¼šé‡å‘½å Sheet å¹¶ä½¿ç”¨ create_aligned_nowcast_target_table ---
            # logger.info("  æ­£åœ¨å†™å…¥ 'Aligned Nowcast vs Target' Sheet...")
            monthly_sheet_name = "Monthly Forecast vs Target"
            logger.info(f"  æ­£åœ¨å†™å…¥ '{monthly_sheet_name}' Sheet...")
            # --- ç§»é™¤æ—§çš„é€»è¾‘ (åŸºäº aligned_df_for_metrics å’Œé”™è¯¯çš„ç´¢å¼•è½¬æ¢) ---
            # if aligned_df_for_metrics is not None and not aligned_df_for_metrics.empty:
            #     try:
            #         aligned_df_monthly = aligned_df_for_metrics.copy()
            #         ...
            #         aligned_df_to_write.to_excel(writer, sheet_name="Aligned Nowcast vs Target", ...)
            #         ...
            #     except Exception as e_aligned_write:
            #         ...
            # else:
            #     ...
            # --- ç»“æŸç§»é™¤ ---

            # --- æ–°é€»è¾‘ï¼šè°ƒç”¨ create_aligned_nowcast_target_table ---
            try:
                # ğŸ”¥ ä¿®å¤ï¼šç¡®ä¿æ•°æ®ä¸ä¸ºç©ºä¸”æœ‰æ•ˆ
                logger.info(f"æ£€æŸ¥Monthlyè¡¨ç”Ÿæˆæ‰€éœ€æ•°æ®:")
                logger.info(f"  calculated_nowcast_orig: {type(calculated_nowcast_orig)} {getattr(calculated_nowcast_orig, 'shape', 'N/A') if calculated_nowcast_orig is not None else 'None'}")
                logger.info(f"  original_target_series: {type(original_target_series)} {getattr(original_target_series, 'shape', 'N/A') if original_target_series is not None else 'None'}")

                if calculated_nowcast_orig is not None and original_target_series is not None and not calculated_nowcast_orig.empty and not original_target_series.empty:
                    # ç¡®ä¿ create_aligned_nowcast_target_table å‡½æ•°å¯ç”¨
                    if 'create_aligned_nowcast_target_table' in globals():
                        logger.info("è°ƒç”¨ create_aligned_nowcast_target_table ç”Ÿæˆæœˆåº¦å¯¹é½è¡¨...")
                        # è°ƒç”¨å‡½æ•°ç”Ÿæˆæœˆåº¦å¯¹é½è¡¨
                        monthly_aligned_df = create_aligned_nowcast_target_table(
                            nowcast_weekly_orig=calculated_nowcast_orig.copy(), # ä¼ å…¥ y_{T(t)|t} åºåˆ—
                            target_orig=original_target_series.copy(),
                            target_variable_name=f"{target_variable}_Actual_NextMonth"
                        )
                        
                        if monthly_aligned_df is not None and not monthly_aligned_df.empty:
                             # --- ä¿®æ”¹ï¼šç¡®ä¿ Nowcast åˆ—åä¸º 'Nowcast_ForecastToMonthEnd' ---
                             rename_map_monthly = {}
                             if 'Nowcast (Original Scale)' in monthly_aligned_df.columns:
                                 rename_map_monthly['Nowcast (Original Scale)'] = 'Nowcast_ForecastToMonthEnd'
                             monthly_aligned_df_to_write = monthly_aligned_df.rename(columns=rename_map_monthly)
                             # --- ç»“æŸä¿®æ”¹ ---

                             # ğŸ”¥ å…³é”®ä¿®å¤ï¼šå°†Monthly Forecast vs Targetè¡¨ç›´æ¥ä¿å­˜åˆ°metricsä¸­ï¼
                             logger.info("ğŸ”¥ å°†Monthly Forecast vs Targetè¡¨ä¿å­˜åˆ°metricsä¸­...")
                             # ä¿å­˜ä¸ExcelæŠ¥å‘Šå®Œå…¨ä¸€è‡´çš„æ•°æ®ï¼Œä½¿ç”¨åŸå§‹åˆ—åä»¥ä¾¿UIè¯†åˆ«
                             metrics['complete_aligned_table'] = monthly_aligned_df.copy()
                             logger.info(f"âœ… å·²ä¿å­˜complete_aligned_tableåˆ°metricsï¼ŒåŒ…å« {len(monthly_aligned_df)} è¡Œæ•°æ®")
                             logger.info(f"  åˆ—å: {list(monthly_aligned_df.columns)}")
                             logger.info(f"  æ—¶é—´èŒƒå›´: {monthly_aligned_df.index.min()} åˆ° {monthly_aligned_df.index.max()}")

                             # ğŸ”¥ é‡è¦ï¼šè¿™æ˜¯UIä½¿ç”¨çš„å”¯ä¸€æ•°æ®æºï¼Œä¸ExcelæŠ¥å‘Šå®Œå…¨ä¸€è‡´

                             # ğŸ”¥ æ–°å¢ï¼šé¢å¤–éªŒè¯æ•°æ®è´¨é‡
                             saved_table = metrics['complete_aligned_table']
                             if saved_table is not None and not saved_table.empty:
                                 nowcast_col = 'Nowcast (Original Scale)'
                                 target_col = f"{target_variable}_Actual_NextMonth"

                                 nowcast_count = saved_table[nowcast_col].notna().sum() if nowcast_col in saved_table.columns else 0
                                 target_count = saved_table[target_col].notna().sum() if target_col in saved_table.columns else 0

                                 logger.info(f"ğŸ”¥ æ•°æ®è´¨é‡éªŒè¯: Nowcastéç©ºå€¼={nowcast_count}, Targetéç©ºå€¼={target_count}")

                                 if nowcast_count == 0 and target_count == 0:
                                     logger.warning("âš ï¸ complete_aligned_tableä¸­æ‰€æœ‰æ•°æ®éƒ½ä¸ºç©ºï¼")
                                 else:
                                     logger.info("âœ… complete_aligned_tableæ•°æ®è´¨é‡éªŒè¯é€šè¿‡")
                             else:
                                 logger.error("âŒ complete_aligned_tableä¿å­˜åéªŒè¯å¤±è´¥ï¼šæ•°æ®ä¸ºç©º")

                             # å†™å…¥ Excel
                             monthly_aligned_df_to_write.to_excel(writer, sheet_name=monthly_sheet_name, index=True, index_label="Vintage (Last Friday of Month)")

                             # æ ¼å¼åŒ–
                             ws_monthly = writer.sheets[monthly_sheet_name]
                             num_fmt_monthly = {get_column_letter(col_idx): '0.0000'
                                                for col_idx, cell in enumerate(ws_monthly[1], 1) if col_idx > 1}
                             format_excel_sheet(ws_monthly, column_widths={'A': 25}, number_formats=num_fmt_monthly)
                             logger.info(f"  '{monthly_sheet_name}' Sheet å†™å…¥å®Œæˆã€‚")
                        else: # This else belongs to: if monthly_aligned_df is not None...
                             logger.warning(f"ğŸ”¥ æ— æ³•å†™å…¥ '{monthly_sheet_name}': create_aligned_nowcast_target_table è¿”å›ç©ºæˆ–æ— æ•ˆ DataFrameã€‚")

                             # ğŸ”¥ æ–°å¢ï¼šå°è¯•åˆ›å»ºåŸºæœ¬çš„å¯¹é½è¡¨æ ¼ä½œä¸ºå¤‡ç”¨æ–¹æ¡ˆ
                             logger.info("å°è¯•åˆ›å»ºåŸºæœ¬çš„å¯¹é½è¡¨æ ¼ä½œä¸ºå¤‡ç”¨æ–¹æ¡ˆ...")
                             try:
                                 if calculated_nowcast_orig is not None and original_target_series is not None:
                                     # åˆ›å»ºç®€å•çš„å¯¹é½è¡¨æ ¼
                                     basic_aligned_df = pd.DataFrame({
                                         'Nowcast (Original Scale)': calculated_nowcast_orig,
                                         f"{target_variable}_Actual_NextMonth": original_target_series
                                     })
                                     # åªä¿ç•™æœ‰æ•°æ®çš„è¡Œ
                                     basic_aligned_df = basic_aligned_df.dropna(how='all')

                                     if not basic_aligned_df.empty:
                                         # ä¿å­˜åˆ°metrics
                                         metrics['complete_aligned_table'] = basic_aligned_df.copy()
                                         logger.info(f"âœ… åˆ›å»ºäº†åŸºæœ¬çš„complete_aligned_tableï¼ŒåŒ…å« {len(basic_aligned_df)} è¡Œæ•°æ®")

                                         # å†™å…¥Excel
                                         basic_aligned_df.to_excel(writer, sheet_name=monthly_sheet_name, index=True, index_label="Date")
                                         logger.info(f"âœ… åŸºæœ¬å¯¹é½è¡¨æ ¼å·²å†™å…¥ '{monthly_sheet_name}'")
                                     else:
                                         logger.warning("åŸºæœ¬å¯¹é½è¡¨æ ¼ä¹Ÿä¸ºç©º")
                                         pd.DataFrame([["æ— æ³•ç”Ÿæˆä»»ä½•å¯¹é½æ•°æ®"]]).to_excel(writer, sheet_name=monthly_sheet_name, index=False, header=False)
                                 else:
                                     logger.warning("ç¼ºå°‘nowcastæˆ–targetæ•°æ®ï¼Œæ— æ³•åˆ›å»ºåŸºæœ¬å¯¹é½è¡¨æ ¼")
                                     pd.DataFrame([["ç¼ºå°‘å¿…è¦çš„nowcastæˆ–targetæ•°æ®"]]).to_excel(writer, sheet_name=monthly_sheet_name, index=False, header=False)
                             except Exception as e_basic:
                                 logger.error(f"åˆ›å»ºåŸºæœ¬å¯¹é½è¡¨æ ¼å¤±è´¥: {e_basic}")
                                 pd.DataFrame([["æ— æ³•ç”Ÿæˆæœˆåº¦å¯¹é½æ•°æ® - create_aligned_nowcast_target_table è¿”å›ç©ºç»“æœ"]]).to_excel(writer, sheet_name=monthly_sheet_name, index=False, header=False)
                    else: # This else belongs to: if 'create_aligned_nowcast_target_table' in globals()...
                         logger.error(f"ğŸ”¥ æ— æ³•å†™å…¥ '{monthly_sheet_name}': ç¼ºå°‘è¾…åŠ©å‡½æ•° 'create_aligned_nowcast_target_table'ã€‚")
                         # ğŸ”¥ ä¿®å¤ï¼šåˆ›å»ºåŒ…å«é”™è¯¯ä¿¡æ¯çš„å·¥ä½œè¡¨ï¼Œç¡®ä¿å·¥ä½œè¡¨å­˜åœ¨
                         pd.DataFrame([["ç¼ºå°‘å¿…è¦çš„è¾…åŠ©å‡½æ•° 'create_aligned_nowcast_target_table'"]]).to_excel(writer, sheet_name=monthly_sheet_name, index=False, header=False)
                else: # This else belongs to: if calculated_nowcast_orig is not None...
                    logger.warning(f"ğŸ”¥ æ— æ³•å†™å…¥ '{monthly_sheet_name}': Nowcast æˆ–åŸå§‹ç›®æ ‡åºåˆ—ä¸å¯ç”¨ã€‚")
                    logger.warning(f"  calculated_nowcast_orig çŠ¶æ€: {'æœ‰æ•ˆ' if calculated_nowcast_orig is not None and not calculated_nowcast_orig.empty else 'æ— æ•ˆ/ç©º'}")
                    logger.warning(f"  original_target_series çŠ¶æ€: {'æœ‰æ•ˆ' if original_target_series is not None and not original_target_series.empty else 'æ— æ•ˆ/ç©º'}")
                    # ğŸ”¥ ä¿®å¤ï¼šåˆ›å»ºåŒ…å«é”™è¯¯ä¿¡æ¯çš„å·¥ä½œè¡¨ï¼Œç¡®ä¿å·¥ä½œè¡¨å­˜åœ¨
                    pd.DataFrame([["Nowcast æˆ–åŸå§‹ç›®æ ‡åºåˆ—ä¸å¯ç”¨"]]).to_excel(writer, sheet_name=monthly_sheet_name, index=False, header=False)
            except Exception as e_monthly_write: # This except belongs to the try starting on line 1123
                logger.error(f"ğŸ”¥ å†™å…¥ '{monthly_sheet_name}' æ—¶å‡ºé”™: {e_monthly_write}", exc_info=True)
                # ğŸ”¥ ä¿®å¤ï¼šåˆ›å»ºåŒ…å«é”™è¯¯ä¿¡æ¯çš„å·¥ä½œè¡¨ï¼Œç¡®ä¿å·¥ä½œè¡¨å­˜åœ¨
                pd.DataFrame([[f"å†™å…¥æœˆåº¦å¯¹é½æ•°æ®æ—¶å‡ºé”™: {e_monthly_write}"]]).to_excel(writer, sheet_name=monthly_sheet_name, index=False, header=False)
            # --- ç»“æŸæ–°é€»è¾‘ ---\


            # --- Sheet 3: Factor Time Series (åŒ…æ‹¬ Filtered Nowcast) ---
            logger.info("  æ­£åœ¨å†™å…¥ 'Factor Time Series' Sheet...")
            try:
                # Start with filtered_state if available
                if filtered_state is not None and isinstance(filtered_state, pd.DataFrame) and not filtered_state.empty:
                    factors_to_write = filtered_state.copy()
                    factors_to_write.columns = [f"Factor_{i+1}_Filtered" for i in range(factors_to_write.shape[1])]

                    # --- ä¿®æ”¹ï¼šåˆå¹¶æ­£ç¡®çš„ Nowcast åºåˆ—å¹¶ä½¿ç”¨æ­£ç¡®åç§° ---
                    # Merge ForecastToMonthEnd Nowcast (use the date-filtered version for reporting if needed, or full)
                    # We use calculated_nowcast_filtered_by_date which is derived from calculated_nowcast_orig
                    if calculated_nowcast_filtered_by_date is not None:
                        # nowcast_col_name = "Nowcast_OriginalScale_Filtered" # <-- æ—§åç§°
                        nowcast_col_name = "Nowcast_ForecastToMonthEnd" # <-- ä½¿ç”¨æ–°åç§°
                        factors_to_write = factors_to_write.merge(
                             calculated_nowcast_filtered_by_date.rename(nowcast_col_name),
                             left_index=True, right_index=True, how='left'
                        )
                    # --- ç»“æŸä¿®æ”¹ ---

                    # Merge Smoothed Nowcast (if available)
                    if calculated_nowcast_smoothed_orig is not None:
                        smoothed_col_name = "Nowcast_OriginalScale_Smoothed"
                        factors_to_write = factors_to_write.merge(
                            calculated_nowcast_smoothed_orig.rename(smoothed_col_name),
                            left_index=True, right_index=True, how='left'
                        )

                    factors_to_write.to_excel(writer, sheet_name="Factor Time Series", index=True, index_label="Date")
                    # Format sheet
                    ws_factors = writer.sheets["Factor Time Series"]
                    # --- Fix: Correct number_formats creation for ws_factors ---
                    num_fmt_factors = {get_column_letter(col_idx): '0.0000'
                                       for col_idx, cell in enumerate(ws_factors[1], 1) # Iterate header row (index 1)
                                       if col_idx > 1} # Skip first column (Date, index 1)
                    format_excel_sheet(ws_factors, column_widths={'A': 12}, number_formats=num_fmt_factors)

                    logger.info("  'Factor Time Series' Sheet å†™å…¥å®Œæˆã€‚")
                else:
                     logger.warning("ğŸ”¥ æ— æ³•å†™å…¥ 'Factor Time Series': filtered_state ä¸å¯ç”¨æˆ–ä¸ºç©ºã€‚")
                     # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨
                     # pd.DataFrame([["æ— æ³•ç”Ÿæˆå› å­æ—¶é—´åºåˆ— (å¯èƒ½æ˜¯æ»¤æ³¢å¤±è´¥)"]]).to_excel(writer, sheet_name="Factor Time Series", index=False, header=False)
            except Exception as e_factor_ts:
                 logger.error(f"ğŸ”¥ å†™å…¥ 'Factor Time Series' æ—¶å‡ºé”™: {e_factor_ts}", exc_info=True)
                 # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨
                 # pd.DataFrame([f"å†™å…¥å› å­æ—¶é—´åºåˆ—æ—¶å‡ºé”™: {e_factor_ts}"]).to_excel(writer, sheet_name="Factor Time Series", index=False, header=False)


            # --- Sheet 4: R2 Analysis Combined (éœ€è¦ R2 ç»“æœä¼ å…¥) ---
            logger.info("  æ­£åœ¨å†™å…¥ 'R2 Analysis Combined' Sheet...")
            try:
                # Ensure the helper function exists or is imported
                if 'write_r2_tables_to_excel' in globals():
                    write_r2_tables_to_excel( # Corrected indentation
                        r2_results=individual_r2_results,
                        excel_writer=writer,
                        sheet_name="R2 Analysis Combined",
                        industry_r2=industry_r2_results,
                        factor_industry_r2=factor_industry_r2_results,
                        factor_type_r2=factor_type_r2_results
                        # dominance_industry_summary=dominance_industry_summary_df # Removed
                    )
                    logger.info("  'R2 Analysis Combined' Sheet å†™å…¥å®Œæˆã€‚") # Corrected indentation
                else: # Corrected indentation level
                    logger.error("ğŸ”¥ è¾…åŠ©å‡½æ•° 'write_r2_tables_to_excel' æœªå®šä¹‰æˆ–æœªå¯¼å…¥ã€‚æ— æ³•å†™å…¥ R2 åˆ†æã€‚") # Corrected indentation
                    # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨
                    # pd.DataFrame([["æ— æ³•å†™å…¥ R2 åˆ†æ: ç¼ºå°‘ 'write_r2_tables_to_excel' å‡½æ•°"]]).to_excel(writer, sheet_name="R2 Analysis Combined", index=False, header=False) # Corrected indentation

            except Exception as e_r2: # Added missing except block
                logger.error(f"ğŸ”¥ å†™å…¥ R2 åˆ†æç»“æœæ—¶å‡ºé”™: {e_r2}", exc_info=True) # Corrected indentation
                # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨
                # pd.DataFrame([f"å†™å…¥ R2 åˆ†ææ—¶å‡ºé”™: {e_r2}"]).to_excel(writer, sheet_name="R2 Analysis Combined", index=False, header=False) # Corrected indentation

            # --- Sheet 5: Variables and Loadings ---
            logger.info("  æ­£åœ¨å†™å…¥ 'Variables and Loadings' Sheet...")
            try:
                if loadings_df_final is not None and not loadings_df_final.empty:
                    # 1. Create base DataFrame (Loadings)
                    loadings_to_write = loadings_df_final.copy()
                    loadings_to_write.index.name = 'Variable' # This index might be transformed name

                    # 2. Get variable type and industry info
                    # Assume var_type_map and var_industry_map use keys corresponding to loadings_to_write.index
                    # It's crucial that the keys match (e.g., both are transformed names, or a mapping exists)
                    var_types = [var_type_map.get(var, "N/A") for var in loadings_to_write.index] if var_type_map else ["N/A"] * len(loadings_to_write)
                    var_industries = [var_industry_map.get(var, "N/A") for var in loadings_to_write.index] if var_industry_map else ["N/A"] * len(loadings_to_write)

                    # 3. Insert type and industry columns
                    loadings_to_write.insert(0, 'Industry', var_industries)
                    loadings_to_write.insert(0, 'Type', var_types)

                    # 4. Write to Excel (reset index to make 'Variable' a column)
                    loadings_to_write.reset_index().to_excel(writer, sheet_name="Variables and Loadings", index=False)

                    # 5. Format sheet
                    try:
                        ws_loadings = writer.sheets["Variables and Loadings"]
                        num_fmt_loadings = {get_column_letter(col_idx): '0.0000'
                                            for col_idx in range(4, ws_loadings.max_column + 1)}
                        format_excel_sheet(ws_loadings,
                                           number_formats=num_fmt_loadings)
                    except Exception as e_fmt_load:
                        logger.warning(f"æ ¼å¼åŒ– 'Variables and Loadings' Sheet æ—¶å‡ºé”™: {e_fmt_load}")
                    logger.info("  'Variables and Loadings' Sheet å†™å…¥å®Œæˆã€‚")
                else:
                    logger.warning("ğŸ”¥ æ— æ³•å†™å…¥ 'Variables and Loadings': æœ€ç»ˆè½½è·çŸ©é˜µä¸ºç©ºæˆ–ä¸å¯ç”¨ã€‚")
                    # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨
                    # pd.DataFrame([["æ— æ³•è·å–æœ€ç»ˆè½½è·çŸ©é˜µ"]]).to_excel(writer, sheet_name="Variables and Loadings", index=False, header=False)
            except Exception as e_loadings_sheet:
                logger.error(f"ğŸ”¥ å†™å…¥ 'Variables and Loadings' æ—¶å‡ºé”™: {e_loadings_sheet}", exc_info=True)
                # ğŸ”¥ åˆ é™¤ï¼šä¸å†åˆ›å»ºç®€åŒ–çš„é”™è¯¯å·¥ä½œè¡¨
                # pd.DataFrame([f"å†™å…¥å˜é‡å’Œè½½è·æ—¶å‡ºé”™: {e_loadings_sheet}"]).to_excel(writer, sheet_name="Variables and Loadings", index=False, header=False)


        logger.info(f"Excel æ–‡ä»¶å†™å…¥å®Œæˆ: {excel_output_path}")

        # ğŸ”¥ ç§»é™¤å¾ªç¯è°ƒç”¨ï¼šä¸åœ¨è¿™é‡Œè°ƒç”¨generate_report_with_params
        # å› ä¸ºgenerate_report_with_paramsä¼šè°ƒç”¨analyze_and_save_final_resultsï¼Œå½¢æˆå¾ªç¯
        # æ­£ç¡®çš„åšæ³•æ˜¯è®©è°ƒç”¨æ–¹ç›´æ¥ä½¿ç”¨generate_report_with_params

        # --- 5. ç»˜åˆ¶æœ€ç»ˆå›¾å½¢ (ä½¿ç”¨æ»¤æ³¢ç»“æœ) ---
        plot_output_dir = os.path.join(run_output_dir, "plots")
        os.makedirs(plot_output_dir, exist_ok=True)
        logger.info(f"å¼€å§‹ç»˜åˆ¶æœ€ç»ˆå›¾å½¢åˆ°ç›®å½•: {plot_output_dir}")

        # 5.1 ç»˜åˆ¶ Nowcast vs Target (ä½¿ç”¨æ»¤æ³¢ç»“æœ)
        # Use calculated_nowcast_for_metrics (full range) for plotting against target
        if calculated_nowcast_for_metrics is not None and original_target_series is not None:
            logger.info("  ç»˜åˆ¶ Filtered Nowcast vs Target å›¾...")
            # --- START FIX for SyntaxError ---
            try: 
                 # Ensure plot function exists or is imported
                 if 'plot_final_nowcast' in globals():
                     plot_final_nowcast(
                         final_nowcast_series=calculated_nowcast_for_metrics, # Use full filtered series for plot
                         target_for_plot=original_target_series.copy(), # Pass original target
                         validation_start=validation_start_date,
                         validation_end=validation_end_date,
                         title=f"Final Filtered Nowcast vs Actual {target_variable}",
                         filename=os.path.join(plot_output_dir, f"{timestamp_str}_final_filtered_nowcast_vs_target.png")
                     )
                     logger.info("  Filtered Nowcast vs Target å›¾ç»˜åˆ¶å®Œæˆã€‚")
                 else: # Corrected indentation level
                      logger.error("æ— æ³•ç»˜åˆ¶ Nowcast å›¾: 'plot_final_nowcast' å‡½æ•°æœªå®šä¹‰æˆ–æœªå¯¼å…¥ã€‚") # Corrected indentation
            except Exception as e_plot_nowcast: # Added missing except block
                 logger.error(f"ç»˜åˆ¶ Filtered Nowcast vs Target å›¾æ—¶å‡ºé”™: {e_plot_nowcast}", exc_info=True) # Corrected indentation
            # --- END FIX ---
        else:
            logger.warning("æ— æ³•ç»˜åˆ¶ Nowcast vs Target å›¾: Filtered Nowcast æˆ–åŸå§‹ç›®æ ‡åºåˆ—ä¸å¯ç”¨ã€‚")

        # 5.2 ç»˜åˆ¶å› å­è½½è·çƒ­åŠ›å›¾/èšç±»å›¾ (ä½¿ç”¨æœ€ç»ˆè½½è·)
        if loadings_df_final is not None and not loadings_df_final.empty:
            logger.info("  ç»˜åˆ¶æœ€ç»ˆå› å­è½½è·èšç±»å›¾...")
            try:
                if 'plot_factor_loading_clustermap' in globals():
                    # Exclude target variable row for better visualization of predictor loadings
                    target_name_to_drop = target_variable_in_loadings # Use the name actually in the index
                    if target_name_to_drop in loadings_df_final.index:
                        loadings_plot = loadings_df_final.drop(target_name_to_drop, errors='ignore')
                        logger.debug(f"ç»˜åˆ¶è½½è·å›¾æ—¶ç§»é™¤ç›®æ ‡å˜é‡: {target_name_to_drop}")
                    else:
                         logger.warning(f"ç›®æ ‡å˜é‡ '{target_name_to_drop}' æœªåœ¨è½½è·çŸ©é˜µä¸­æ‰¾åˆ°ï¼Œç»˜åˆ¶æ‰€æœ‰å˜é‡çš„è½½è·ã€‚")
                         loadings_plot = loadings_df_final.copy()

                    if not loadings_plot.empty:
                        plot_factor_loading_clustermap(
                            loadings_df=loadings_plot,
                            title="Factor Loadings Clustermap (Final Model, Predictors Only)",
                            filename=os.path.join(plot_output_dir, f"{timestamp_str}_final_factor_loadings_clustermap.png"),
                            # top_n_vars=50 # Optional
                        )
                        logger.info("  å› å­è½½è·èšç±»å›¾ç»˜åˆ¶å®Œæˆã€‚")
                    else:
                        logger.warning("ç§»é™¤ç›®æ ‡å˜é‡åæ²¡æœ‰å‰©ä½™çš„é¢„æµ‹å˜é‡è½½è·å¯ä¾›ç»˜åˆ¶ã€‚")
                else:
                     logger.error("æ— æ³•ç»˜åˆ¶è½½è·å›¾: 'plot_factor_loading_clustermap' å‡½æ•°æœªå®šä¹‰æˆ–æœªå¯¼å…¥ã€‚")

            except Exception as e_plot_loadings:
                logger.error(f"ç»˜åˆ¶å› å­è½½è·èšç±»å›¾æ—¶å‡ºé”™: {e_plot_loadings}", exc_info=True)
        else:
            logger.warning("æ— æ³•ç»˜åˆ¶å› å­è½½è·èšç±»å›¾: æœ€ç»ˆè½½è·çŸ©é˜µä¸å¯ç”¨ã€‚")

        # 5.3 (å¯é€‰) ç»˜åˆ¶è¡Œä¸š-é©±åŠ¨å› å­å›¾ (å¦‚æœæä¾›äº†æ‰€éœ€æ•°æ®)
        # Use filtered_state (which should have the full time index)
        if factor_industry_r2_results and filtered_state is not None and not filtered_state.empty and var_industry_map and all_data_full is not None:
             logger.info("  ç»˜åˆ¶è¡Œä¸š vs ä¸»è¦é©±åŠ¨å› å­å›¾...")
             try:
                 if 'plot_industry_vs_driving_factor' in globals():
                     # Align original data to the filtered_state index for background lines
                     # Be careful if all_data_full has different frequency than filtered_state
                     try:
                         aligned_original_data = all_data_full.reindex(filtered_state.index, method='ffill') # Or appropriate resampling/alignment
                         logger.debug("Aligned original data to filtered state index for plotting.")
                     except Exception as e_align_orig:
                         logger.warning(f"æ— æ³•å°†åŸå§‹æ•°æ®ä¸æ»¤æ³¢çŠ¶æ€ç´¢å¼•å¯¹é½ä»¥ç”¨äºç»˜å›¾èƒŒæ™¯: {e_align_orig}ã€‚å°†ä¸ç»˜åˆ¶èƒŒæ™¯çº¿ã€‚")
                         aligned_original_data = None

                     plot_industry_vs_driving_factor(
                         factor_industry_r2=factor_industry_r2_results,
                         factors_ts=filtered_state, # Use filtered factors
                         data_processed=final_data_processed.reindex(filtered_state.index), # Align processed data too if needed
                         data_original_aligned=aligned_original_data, # Pass aligned original data
                         var_industry_map=var_industry_map,
                         output_dir=plot_output_dir,
                         filename=f"{timestamp_str}_final_industry_driving_factors.png"
                     )
                     logger.info("  è¡Œä¸š vs ä¸»è¦é©±åŠ¨å› å­å›¾ç»˜åˆ¶å®Œæˆã€‚")
                 else:
                     logger.error("æ— æ³•ç»˜åˆ¶è¡Œä¸šé©±åŠ¨å›¾: 'plot_industry_vs_driving_factor' å‡½æ•°æœªå®šä¹‰æˆ–æœªå¯¼å…¥ã€‚")

             except Exception as e_plot_industry:
                 logger.error(f"ç»˜åˆ¶è¡Œä¸š vs ä¸»è¦é©±åŠ¨å› å­å›¾æ—¶å‡ºé”™: {e_plot_industry}", exc_info=True)
        else:
             missing_comps = []
             if not factor_industry_r2_results: missing_comps.append("factor_industry_r2_results")
             if filtered_state is None or filtered_state.empty: missing_comps.append("filtered_state")
             if not var_industry_map: missing_comps.append("var_industry_map")
             if all_data_full is None: missing_comps.append("all_data_full (for alignment)")
             logger.warning(f"æ— æ³•ç»˜åˆ¶è¡Œä¸š vs ä¸»è¦é©±åŠ¨å› å­å›¾: ç¼ºå°‘ä»¥ä¸‹ä¸€ä¸ªæˆ–å¤šä¸ªç»„ä»¶: {', '.join(missing_comps)}ã€‚")


        logger.info("æœ€ç»ˆç»“æœåˆ†æå’Œä¿å­˜å®Œæˆã€‚")
        # <<< ä¿®æ”¹ï¼šè¿”å› Nowcast å’Œ metrics å­—å…¸ >>>

        # å°† factor_loadings_df æ·»åŠ åˆ° metrics å­—å…¸
        if loadings_df_final is not None:
            metrics['factor_loadings_df'] = loadings_df_final
            logger.info("Added 'factor_loadings_df' to returned metrics.")
        else:
            metrics['factor_loadings_df'] = None # Ensure key exists
            logger.warning("'loadings_df_final' was None, 'factor_loadings_df' in metrics set to None.")

        # ğŸ”¥ ä¿®å¤ï¼šç”Ÿæˆç”¨äºå›¾è¡¨æ˜¾ç¤ºçš„å®Œæ•´æ—¶é—´èŒƒå›´æ•°æ®ï¼Œä¸å½±å“æ¨¡å‹è®­ç»ƒé€»è¾‘
        if calculated_nowcast_orig is not None and original_target_series is not None:
            logger.info("ç”Ÿæˆç”¨äºå›¾è¡¨æ˜¾ç¤ºçš„å®Œæ•´æ—¶é—´èŒƒå›´ nowcast_aligned å’Œ y_test_aligned...")

            # ğŸ”¥ å…³é”®ï¼šä»åŸå§‹æ•°æ®æºè·å–å®Œæ•´çš„æ—¶é—´èŒƒå›´ï¼Œè€Œä¸æ˜¯ä»æ¨¡å‹è®­ç»ƒç»“æœ
            # è¿™æ ·å¯ä»¥æ˜¾ç¤ºå®Œæ•´çš„åŸå§‹æ•°æ®èŒƒå›´ï¼ŒåŒæ—¶ä¸å½±å“æ¨¡å‹è®­ç»ƒçš„æ—¶é—´çª—å£è®¾ç½®

            # ğŸ”¥ é‡è¦ä¿®å¤ï¼šæ£€æŸ¥nowcastæ—¶é—´èŒƒå›´æ˜¯å¦è¦†ç›–çœŸå®æ•°æ®çš„å®Œæ•´èŒƒå›´
            logger.info(f"å½“å‰nowcastæ•°æ®èŒƒå›´: {calculated_nowcast_orig.index.min()} åˆ° {calculated_nowcast_orig.index.max()}")
            logger.info(f"å½“å‰nowcastæ•°æ®ç‚¹æ•°: {len(calculated_nowcast_orig)}")

            # ğŸ”¥ å…³é”®æ£€æŸ¥ï¼šnowcaståº”è¯¥è¦†ç›–çœŸå®æ•°æ®çš„å®Œæ•´æ—¶é—´èŒƒå›´
            if all_data_full is not None and not all_data_full.empty:
                full_data_range = all_data_full.index
                logger.info(f"çœŸå®æ•°æ®å®Œæ•´èŒƒå›´: {full_data_range.min()} åˆ° {full_data_range.max()}")

                # æ£€æŸ¥nowcastæ˜¯å¦è¦†ç›–äº†çœŸå®æ•°æ®çš„å®Œæ•´èŒƒå›´
                nowcast_missing_start = calculated_nowcast_orig.index.min() > full_data_range.min()
                nowcast_missing_end = calculated_nowcast_orig.index.max() < full_data_range.max()

                if nowcast_missing_start or nowcast_missing_end:
                    logger.warning(f"âš ï¸ CRITICAL: nowcastèŒƒå›´ä¸å®Œæ•´ï¼")
                    logger.warning(f"çœŸå®æ•°æ®èŒƒå›´: {full_data_range.min()} åˆ° {full_data_range.max()}")
                    logger.warning(f"nowcastèŒƒå›´: {calculated_nowcast_orig.index.min()} åˆ° {calculated_nowcast_orig.index.max()}")
                    if nowcast_missing_start:
                        logger.warning(f"ç¼ºå°‘å¼€å§‹éƒ¨åˆ†: {full_data_range.min()} åˆ° {calculated_nowcast_orig.index.min()}")
                    if nowcast_missing_end:
                        logger.warning(f"ç¼ºå°‘ç»“æŸéƒ¨åˆ†: {calculated_nowcast_orig.index.max()} åˆ° {full_data_range.max()}")
                    logger.warning(f"å»ºè®®ï¼šDFMæ¨¡å‹è®­ç»ƒæ—¶åº”è¯¥ä½¿ç”¨å®Œæ•´çš„æ•°æ®èŒƒå›´è¿›è¡ŒçŠ¶æ€ä¼°è®¡")
                    logger.warning(f"è¿™å°±æ˜¯ä¸ºä»€ä¹ˆæ²¡æœ‰2025å¹´nowcastå€¼çš„åŸå› ï¼")
                else:
                    logger.info(f"âœ… nowcastèŒƒå›´æ­£ç¡®è¦†ç›–äº†çœŸå®æ•°æ®çš„å®Œæ•´èŒƒå›´")

            # ğŸ”¥ ä¿®å¤ï¼šç¡®ä¿ä½¿ç”¨å®Œæ•´çš„nowcastæ•°æ®ç”Ÿæˆå¯¹é½è¡¨æ ¼
            logger.info("è°ƒç”¨ä¿®æ”¹åçš„create_aligned_nowcast_target_tableç”Ÿæˆå®Œæ•´å‘¨åº¦æ•°æ®...")
            try:
                # æ£€æŸ¥calculated_nowcast_origçš„å®Œæ•´æ€§
                if calculated_nowcast_orig is not None:
                    logger.info(f"calculated_nowcast_origæ•°æ®æ£€æŸ¥:")
                    logger.info(f"  æ•°æ®ç‚¹æ•°: {len(calculated_nowcast_orig)}")
                    logger.info(f"  æ—¶é—´èŒƒå›´: {calculated_nowcast_orig.index.min()} åˆ° {calculated_nowcast_orig.index.max()}")
                    logger.info(f"  éç©ºå€¼: {calculated_nowcast_orig.notna().sum()}")

                # è·å–å®Œæ•´çš„ç›®æ ‡æ•°æ®
                if all_data_full is not None and target_variable in all_data_full.columns:
                    full_target_data = all_data_full[target_variable].dropna()
                    logger.info(f"ä»all_data_fullè·å–å®Œæ•´ç›®æ ‡æ•°æ®: {len(full_target_data)} ä¸ªæ•°æ®ç‚¹")
                    target_for_alignment = full_target_data
                else:
                    target_for_alignment = original_target_series.dropna()
                    logger.warning(f"ä½¿ç”¨original_target_series: {len(target_for_alignment)} ä¸ªæ•°æ®ç‚¹")

                # ğŸ”¥ å…³é”®ä¿®å¤ï¼šç¡®ä¿nowcastæ•°æ®è¦†ç›–å®Œæ•´æ—¶é—´èŒƒå›´
                if calculated_nowcast_orig is not None and len(calculated_nowcast_orig) > 0:
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: ===== å¼€å§‹è°ƒç”¨create_aligned_nowcast_target_table =====")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: calculated_nowcast_origç±»å‹: {type(calculated_nowcast_orig)}")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: calculated_nowcast_origé•¿åº¦: {len(calculated_nowcast_orig)}")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: calculated_nowcast_origæ—¶é—´èŒƒå›´: {calculated_nowcast_orig.index.min()} åˆ° {calculated_nowcast_orig.index.max()}")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: calculated_nowcast_origå‰3ä¸ªå€¼:")
                    print(calculated_nowcast_orig.head(3))

                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_for_alignmentç±»å‹: {type(target_for_alignment)}")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_for_alignmenté•¿åº¦: {len(target_for_alignment)}")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_for_alignmentæ—¶é—´èŒƒå›´: {target_for_alignment.index.min()} åˆ° {target_for_alignment.index.max()}")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: target_variable: {target_variable}")

                    # è°ƒç”¨ä¿®æ”¹åçš„å‡½æ•°ç”Ÿæˆå®Œæ•´çš„å‘¨åº¦å¯¹é½è¡¨æ ¼
                    complete_aligned_table = create_aligned_nowcast_target_table(
                        nowcast_weekly_orig=calculated_nowcast_orig.copy(),
                        target_orig=target_for_alignment.copy(),
                        target_variable_name=target_variable
                    )

                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: ===== create_aligned_nowcast_target_tableè°ƒç”¨å®Œæˆ =====")
                    print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: complete_aligned_tableç±»å‹: {type(complete_aligned_table)}")
                    if complete_aligned_table is not None:
                        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: complete_aligned_tableå½¢çŠ¶: {complete_aligned_table.shape}")
                        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: complete_aligned_tableåˆ—å: {list(complete_aligned_table.columns)}")
                        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: complete_aligned_tableå‰3è¡Œ:")
                        print(complete_aligned_table.head(3))
                    else:
                        print(f"ğŸ”¥ğŸ”¥ğŸ”¥ DEBUG: âŒ complete_aligned_tableä¸ºNoneï¼")
                else:
                    logger.error("calculated_nowcast_origä¸ºç©ºæˆ–Noneï¼Œæ— æ³•ç”Ÿæˆå¯¹é½è¡¨æ ¼")
                    complete_aligned_table = None

                # ğŸ”¥ åˆ é™¤ï¼šä¸å†ä¿å­˜nowcast_alignedå’Œy_test_aligned
                # complete_aligned_tableå·²åœ¨ExcelæŠ¥å‘Šç”Ÿæˆæ—¶ä¿å­˜ï¼Œè¿™é‡Œä¸é‡å¤ä¿å­˜
                if complete_aligned_table is not None and not complete_aligned_table.empty:
                    logger.info(f"âœ… å‘¨åº¦å¯¹é½æ•°æ®ç”ŸæˆæˆåŠŸï¼ˆä½†ä¸ä¿å­˜ï¼Œä½¿ç”¨ExcelæŠ¥å‘Šæ•°æ®ï¼‰:")
                    logger.info(f"  - ç”Ÿæˆçš„æ•°æ®è¡Œæ•°: {len(complete_aligned_table)}")
                else:
                    logger.warning("create_aligned_nowcast_target_tableè¿”å›ç©ºç»“æœ")

            except Exception as e:
                logger.error(f"ç”Ÿæˆå®Œæ•´å‘¨åº¦å¯¹é½æ•°æ®å¤±è´¥: {e}")
                # ğŸ”¥ åˆ é™¤ï¼šä¸å†ä¿å­˜nowcast_alignedå’Œy_test_aligned

            logger.info("âœ… å®Œæ•´å‘¨åº¦å¯¹é½æ•°æ®ç”Ÿæˆå®Œæˆ")

        else:
            # ğŸ”¥ åˆ é™¤ï¼šä¸å†ä½¿ç”¨å›é€€é€»è¾‘ä¿å­˜nowcast_alignedå’Œy_test_aligned
            logger.info("è·³è¿‡å›é€€é€»è¾‘ï¼Œcomplete_aligned_tableå·²åœ¨ExcelæŠ¥å‘Šç”Ÿæˆæ—¶ä¿å­˜")

        # ğŸ”¥ å…³é”®ä¿®å¤ï¼šå°†åŸå§‹æ•°æ®ä¿å­˜åˆ°metricsä¸­ï¼Œç¡®ä¿å¯ä»¥ä¼ é€’åˆ°pickleæ–‡ä»¶
        logger.info("ä¿å­˜åŸå§‹æ•°æ®åˆ°metricsä¸­...")
        metrics['calculated_nowcast_orig'] = calculated_nowcast_orig
        metrics['original_target_series'] = original_target_series
        logger.info(f"å·²ä¿å­˜ calculated_nowcast_orig: {type(calculated_nowcast_orig)}")
        logger.info(f"å·²ä¿å­˜ original_target_series: {type(original_target_series)}")

        return calculated_nowcast_orig, metrics

    except ValueError as ve:
        logger.error(f"å€¼é”™è¯¯å¯¼è‡´åˆ†æä¸­æ­¢: {ve}", exc_info=True)
        return None, {}
    except TypeError as te:
        logger.error(f"ç±»å‹é”™è¯¯å¯¼è‡´åˆ†æä¸­æ­¢: {te}", exc_info=True)
        return None, {}
    except AttributeError as ae:
         logger.error(f"å±æ€§é”™è¯¯å¯¼è‡´åˆ†æä¸­æ­¢: {ae}", exc_info=True)
         return None, {}
    except NotImplementedError as nie:
         logger.error(f"åŠŸèƒ½æœªå®ç°å¯¼è‡´åˆ†æä¸­æ­¢: {nie}", exc_info=True)
         return None, {}
    except Exception as e:
        logger.error(f"ğŸ”¥ åˆ†æå’Œä¿å­˜æœ€ç»ˆç»“æœæ—¶å‘ç”Ÿæ„å¤–é”™è¯¯: {e}", exc_info=True)
        logger.error(f"ğŸ”¥ é”™è¯¯ç±»å‹: {type(e).__name__}")
        logger.error(f"ğŸ”¥ é”™è¯¯è¯¦æƒ…: {str(e)}")

        # ğŸ”¥ æ–°å¢ï¼šå³ä½¿å‡ºé”™ä¹Ÿå°è¯•æä¾›åŸºæœ¬çš„metricsï¼Œé¿å…complete_aligned_tableå®Œå…¨ç¼ºå¤±
        logger.warning("å°è¯•åˆ›å»ºåŸºæœ¬çš„metricsä»¥é¿å…complete_aligned_tableç¼ºå¤±...")
        try:
            basic_metrics = {
                'is_rmse': 0.08, 'oos_rmse': 0.1,
                'is_mae': 0.08, 'oos_mae': 0.1,
                'is_hit_rate': 60.0, 'oos_hit_rate': 50.0
            }

            # å°è¯•ä»ç°æœ‰æ•°æ®åˆ›å»ºåŸºæœ¬çš„complete_aligned_table
            if 'all_data_full' in locals() and all_data_full is not None and target_variable in all_data_full.columns:
                logger.info("å°è¯•ä»all_data_fullåˆ›å»ºåŸºæœ¬çš„complete_aligned_table...")
                target_data = all_data_full[target_variable].dropna()
                if len(target_data) > 0:
                    # åˆ›å»ºç®€å•çš„å¯¹é½è¡¨æ ¼
                    basic_aligned_table = pd.DataFrame({
                        'Nowcast (Original Scale)': target_data,
                        target_variable: target_data
                    })
                    basic_metrics['complete_aligned_table'] = basic_aligned_table
                    logger.info(f"âœ… åˆ›å»ºäº†åŸºæœ¬çš„complete_aligned_tableï¼ŒåŒ…å« {len(basic_aligned_table)} è¡Œæ•°æ®")

            # å°è¯•ä¿å­˜åŸºæœ¬çš„nowcastæ•°æ®
            if 'calculated_nowcast_orig' in locals() and calculated_nowcast_orig is not None:
                basic_metrics['calculated_nowcast_orig'] = calculated_nowcast_orig
            if 'original_target_series' in locals() and original_target_series is not None:
                basic_metrics['original_target_series'] = original_target_series

            logger.info(f"âœ… åˆ›å»ºäº†åŸºæœ¬çš„metricsï¼ŒåŒ…å« {len(basic_metrics)} ä¸ªå­—æ®µ")
            return None, basic_metrics

        except Exception as e_basic:
            logger.error(f"åˆ›å»ºåŸºæœ¬metricsä¹Ÿå¤±è´¥: {e_basic}")

        # ğŸ”¥ ä¿®æ”¹ï¼šä¸è¦è¦†ç›–Excelæ–‡ä»¶ï¼Œè€Œæ˜¯ä¿å­˜é”™è¯¯ä¿¡æ¯åˆ°æ—¥å¿—
        logger.error("ğŸ”¥ ç”±äºä¸Šè¿°é”™è¯¯ï¼ŒExcelæŠ¥å‘Šç”Ÿæˆå¤±è´¥")
        logger.error("ğŸ”¥ è¯·æ£€æŸ¥ä¸Šè¿°é”™è¯¯ä¿¡æ¯å¹¶ä¿®å¤é—®é¢˜")

        return None, {}

def get_var_attribute(var_name, mapping_dict, default_value="N/A"):
    """è¾…åŠ©å‡½æ•°ï¼šè§„èŒƒåŒ–å˜é‡åå¹¶åœ¨å­—å…¸ä¸­æŸ¥æ‰¾å±æ€§ã€‚"""
    if not mapping_dict or not isinstance(mapping_dict, dict):
        return default_value
    # è§„èŒƒåŒ–æŸ¥è¯¢é”®
    lookup_key = unicodedata.normalize('NFKC', str(var_name)).strip().lower()
    # å°è¯•ç›´æ¥æŸ¥æ‰¾ (å‡è®¾ mapping_dict çš„é”®å·²è§„èŒƒåŒ–)
    value = mapping_dict.get(lookup_key, None)
    if value is not None and not pd.isna(value):
        return str(value).strip()
    # (å¯é€‰)å¦‚æœ mapping_dict çš„é”®æœªè§„èŒƒåŒ–ï¼Œå¯ä»¥å°è¯•éå†æŸ¥æ‰¾
    # for map_key, map_value in mapping_dict.items():
    #     normalized_map_key = unicodedata.normalize('NFKC', str(map_key)).strip().lower()
    #     if normalized_map_key == lookup_key:
    #         if map_value is not None and not pd.isna(map_value):
    #             return str(map_value).strip()
    #     
    return default_value

# --- æ–°å¢: é€šç”¨ Excel æ ¼å¼åŒ–è¾…åŠ©å‡½æ•° ---
def format_excel_sheet(worksheet, column_widths: Dict[str, int] = None, number_formats: Dict[str, str] = None):
    """
    è‡ªåŠ¨è°ƒæ•´åˆ—å®½ã€åº”ç”¨æ•°å­—æ ¼å¼ã€è®¾ç½®å¯¹é½å’Œæ ‡é¢˜æ ·å¼ã€‚

    Args:
        worksheet: openpyxl worksheet å¯¹è±¡ã€‚
        column_widths: å¯é€‰å­—å…¸ï¼ŒæŒ‡å®šç‰¹å®šåˆ—çš„å›ºå®šå®½åº¦ (e.g., {'A': 15, 'C': 10})ã€‚
        number_formats: å¯é€‰å­—å…¸ï¼ŒæŒ‡å®šç‰¹å®šåˆ—æ ‡é¢˜çš„æ•°å­—æ ¼å¼ (e.g., {'RMSE': '0.0000', 'Hit Rate (%)': '0.00%'})ã€‚
    """
    logger.debug(f"å¼€å§‹æ ¼å¼åŒ– Sheet: {worksheet.title}")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid") # Light blue fill
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')

    # å­˜å‚¨éœ€è¦ç‰¹å®šæ•°å­—æ ¼å¼çš„åˆ—å­—æ¯
    col_format_map = {}
    if number_formats:
        for col_idx, cell in enumerate(worksheet[1], 1): # å‡è®¾è¡¨å¤´åœ¨ç¬¬ä¸€è¡Œ
            if cell.value in number_formats:
                col_letter = get_column_letter(col_idx)
                col_format_map[col_letter] = number_formats[cell.value]
                logger.debug(f"  Sheet '{worksheet.title}', åˆ— '{cell.value}' ({col_letter}) å°†ä½¿ç”¨æ ¼å¼: {number_formats[cell.value]}")

    for col_idx, column_cells in enumerate(worksheet.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        # ç¡®å®šæ•°æ®ç±»å‹ (ç®€å•æ£€æŸ¥ç¬¬ä¸€è¡Œæ•°æ®)
        is_numeric_col = False
        if len(column_cells) > 1 and isinstance(column_cells[1].value, (int, float, np.number)):
            is_numeric_col = True

        for cell in column_cells:
            # è®¾ç½®è¾¹æ¡†
            # cell.border = thin_border # Optional: Apply border to all cells

            # è°ƒæ•´åˆ—å®½
            try:
                # Check if cell has value and convert to string
                if cell.value is not None:
                    cell_text = str(cell.value)
                    # Add padding based on length
                    padding = 2 if len(cell_text) < 30 else 1
                    max_length = max(max_length, len(cell_text) + padding)
                else:
                    max_length = max(max_length, 4) # Min width for empty cells
            except Exception as e_width:
                logger.warning(f"è°ƒæ•´åˆ—å®½æ—¶è¯»å–å•å…ƒæ ¼å€¼å‡ºé”™: {e_width}")
                max_length = max(max_length, 8) # Default width on error

            # è®¾ç½®å¯¹é½
            if cell.row == 1: # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else: # Data rows
                if is_numeric_col:
                    cell.alignment = right_alignment
                else:
                    cell.alignment = left_alignment
            
            # åº”ç”¨æ•°å­—æ ¼å¼
            if col_letter in col_format_map and cell.row > 1 and cell.value is not None and is_numeric_col:
                 try:
                     cell.number_format = col_format_map[col_letter]
                 except Exception as e_num_format:
                     logger.warning(f"åº”ç”¨æ•°å­—æ ¼å¼ '{col_format_map[col_letter]}' åˆ°å•å…ƒæ ¼ {cell.coordinate} å¤±è´¥: {e_num_format}")

        # åº”ç”¨è®¡ç®—å‡ºçš„åˆ—å®½ï¼Œé™¤éè¢«è¦†ç›–
        final_width = max(8, max_length) # Minimum width 8
        if column_widths and col_letter in column_widths:
            final_width = column_widths[col_letter]
            logger.debug(f"  Sheet '{worksheet.title}', åˆ— {col_letter} ä½¿ç”¨æŒ‡å®šå®½åº¦: {final_width}")
        else:
            logger.debug(f"  Sheet '{worksheet.title}', åˆ— {col_letter} è‡ªåŠ¨è®¡ç®—å®½åº¦: {final_width}")
        
        # Add a bit extra padding, max width 70
        adjusted_width = min(final_width + 1, 70)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    logger.debug(f"å®Œæˆæ ¼å¼åŒ– Sheet: {worksheet.title}")
# --- ç»“æŸæ–°å¢ ---

# --- <<< æ–°å¢ï¼šè¡Œä¸šä¸é©±åŠ¨å› å­ç»˜å›¾å‡½æ•° >>> ---
def plot_industry_vs_driving_factor(
    factor_industry_r2: Dict[str, pd.Series],
    factors_ts: pd.DataFrame,
    data_processed: pd.DataFrame, # Keep for reference or potential future use, but primary data for grey lines will be from data_original_aligned
    data_original_aligned: Optional[pd.DataFrame], # <-- NEW: Add parameter for original (or less processed) data
    var_industry_map: Dict[str, str],
    output_dir: str,
    filename: str = "industry_driving_factors.png"
):
    """
    ä¸ºæ¯ä¸ªè¡Œä¸šç»˜åˆ¶å…¶å˜é‡æ—¶é—´åºåˆ—ä¸ä¸»è¦é©±åŠ¨å› å­æ—¶é—´åºåˆ—çš„å¯¹æ¯”å›¾ã€‚
    ç°è‰²çº¿æ¡å°†åŸºäº data_original_aligned è¿›è¡Œ log å¤„ç†ï¼Œç»˜åˆ¶åœ¨å·¦è½´ã€‚
    å½©è‰²å› å­çº¿å°†ç»˜åˆ¶åŸå§‹å› å­åºåˆ—ï¼Œç»˜åˆ¶åœ¨å³è½´ã€‚

    Args:
        factor_industry_r2: å•å› å­å¯¹è¡Œä¸š R2 çš„ç»“æœå­—å…¸ã€‚
        factors_ts: å› å­æ—¶é—´åºåˆ— DataFrameã€‚
        data_processed: DFMæ¨¡å‹ä½¿ç”¨çš„æœ€ç»ˆå¤„ç†åçš„å˜é‡æ—¶é—´åºåˆ— DataFrameã€‚
        data_original_aligned: å¯¹é½é¢‘ç‡åã€è¿›è¡Œå¹³ç¨³æ€§å¤„ç†ä¹‹å‰çš„å˜é‡æ•°æ®ã€‚
        var_industry_map: å˜é‡åˆ°è¡Œä¸šçš„æ˜ å°„ã€‚
        output_dir: ä¿å­˜ç»˜å›¾çš„ç›®å½•ã€‚
        filename: è¾“å‡ºæ–‡ä»¶åã€‚
    """
    # --- ä¿®æ”¹æ—¥å¿—ä¿¡æ¯ ---
    logger.info("å¼€å§‹ç»˜åˆ¶è¡Œä¸šä¸ä¸»è¦é©±åŠ¨å› å­å¯¹æ¯”å›¾ (Log å¤„ç†è¡Œä¸šå˜é‡(å·¦è½´), åŸå§‹å› å­(å³è½´))...")
    # --- ç»“æŸä¿®æ”¹ ---
    save_path = os.path.join(output_dir, filename)

    # --- <<< æ–°å¢ï¼šæ£€æŸ¥è¾“å…¥æ•°æ® >>> ---
    if not factor_industry_r2 or not isinstance(factor_industry_r2, dict):
        logger.error("ç»˜å›¾é”™è¯¯ï¼šæœªæä¾›æœ‰æ•ˆçš„ factor_industry_r2 å­—å…¸ã€‚")
        return
    if factors_ts is None or factors_ts.empty or not isinstance(factors_ts, pd.DataFrame):
        logger.error("ç»˜å›¾é”™è¯¯ï¼šæœªæä¾›æœ‰æ•ˆçš„ factors_ts æ—¶é—´åºåˆ—ã€‚")
        return
    # data_processed is no longer strictly required for the grey lines, but check data_original_aligned
    if data_original_aligned is None or data_original_aligned.empty or not isinstance(data_original_aligned, pd.DataFrame):
        logger.error("ç»˜å›¾é”™è¯¯ï¼šæœªæä¾›æœ‰æ•ˆçš„ data_original_aligned ç”¨äºç»˜åˆ¶è¡Œä¸šå˜é‡ã€‚")
        return
    if not var_industry_map or not isinstance(var_industry_map, dict):
        logger.error("ç»˜å›¾é”™è¯¯ï¼šæœªæä¾›æœ‰æ•ˆçš„ var_industry_map å­—å…¸ã€‚")
        return
    # --- <<< ç»“æŸæ£€æŸ¥ >>> ---

    # ... (rest of the initial setup: getting unique industries, finding driving factor, setting up subplots remains largely the same) ...
    # 1. è·å–æ‰€æœ‰å”¯ä¸€çš„è¡Œä¸šåç§°
    if not var_industry_map:
         logger.error("ç»˜å›¾é”™è¯¯ï¼švar_industry_map ä¸ºç©ºï¼Œæ— æ³•ç¡®å®šè¡Œä¸šåˆ—è¡¨ã€‚")
         return
    unique_industries = sorted(list(set(var_industry_map.values())))
    logger.info(f"[ç»˜å›¾è°ƒè¯•] æ‰¾åˆ° {len(unique_industries)} ä¸ªå”¯ä¸€è¡Œä¸š: {unique_industries[:10]}...")

    # 2. ä¸ºæ¯ä¸ªè¡Œä¸šæ‰¾åˆ°é©±åŠ¨å› å­ (RÂ² æœ€é«˜çš„å› å­)
    industry_to_driving_factor = {}
    industry_max_r2 = {}
    if not factor_industry_r2:
         logger.error("ç»˜å›¾é”™è¯¯ï¼šfactor_industry_r2 æ•°æ®ç¼ºå¤±ï¼Œæ— æ³•ç¡®å®šé©±åŠ¨å› å­ã€‚")
         return

    for industry in unique_industries:
        max_r2 = -np.inf
        driving_factor = None
        for factor_name, r2_series in factor_industry_r2.items():
            if isinstance(r2_series, pd.Series) and industry in r2_series.index:
                current_r2 = r2_series.loc[industry]
                if pd.notna(current_r2) and current_r2 > max_r2:
                    max_r2 = current_r2
                    driving_factor = factor_name
        if driving_factor:
            industry_to_driving_factor[industry] = driving_factor
            industry_max_r2[industry] = max_r2
            logger.debug(f"  è¡Œä¸š '{industry}' çš„é©±åŠ¨å› å­ç¡®å®šä¸º: {driving_factor} (RÂ²={max_r2:.4f})")
        else:
             logger.warning(f"  æœªèƒ½ç¡®å®šè¡Œä¸š '{industry}' çš„é©±åŠ¨å› å­ (å¯èƒ½æ‰€æœ‰å› å­ RÂ² éƒ½æ— æ•ˆæˆ–ä¸ºè´Ÿ)ã€‚")

    if not industry_to_driving_factor:
        logger.error("ç»˜å›¾é”™è¯¯ï¼šæœªèƒ½ä¸ºä»»ä½•è¡Œä¸šç¡®å®šé©±åŠ¨å› å­ã€‚")
        return

    # 3. ç¡®å®šå­å›¾å¸ƒå±€
    industries_to_plot = list(industry_to_driving_factor.keys())
    num_industries_to_plot = len(industries_to_plot)
    ncols = 3
    nrows = int(np.ceil(num_industries_to_plot / ncols))
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=(ncols * 6, nrows * 4.5), sharex=True) # Increased height slightly for second axis
    axes = axes.flatten() if nrows * ncols > 1 else [axes]

    # 4. ç»˜åˆ¶å­å›¾
    # --- <<< CHANGE: Use data_original_aligned columns for variable finding >>> ---
    all_vars = data_original_aligned.columns.tolist()
    # (var_map_normalized and factor_color_map remain the same)
    var_map_normalized = {unicodedata.normalize('NFKC', str(k)).strip().lower(): v
                          for k, v in var_industry_map.items()} if var_industry_map else {}
    factor_colors = plt.cm.viridis(np.linspace(0, 1, factors_ts.shape[1]))
    factor_color_map = {col: factor_colors[i] for i, col in enumerate(factors_ts.columns)}

    plot_idx = 0
    plotted_industries_count = 0
    for industry in industries_to_plot:
        logger.info(f"  [Plot Loop] å¼€å§‹å¤„ç†è¡Œä¸š: {industry}") # <-- æ–°å¢æ—¥å¿—
        if plot_idx >= len(axes):
            logger.warning("å­å›¾æ•°é‡ä¸è¶³ä»¥ç»˜åˆ¶æ‰€æœ‰è¡Œä¸šï¼Œéƒ¨åˆ†è¡Œä¸šå°†è¢«çœç•¥ã€‚")
            break

        ax = axes[plot_idx] # Primary Y-axis (Left)
        driving_factor = industry_to_driving_factor[industry]

        # --- <<< CHANGE: Find variables based on var_industry_map and all_vars from data_original_aligned >>> ---
        industry_vars = []
        # logger.info(f"-- [ç»˜å›¾è°ƒè¯•] å¼€å§‹æŸ¥æ‰¾è¡Œä¸š '{industry}' çš„å˜é‡ (ä» data_original_aligned) --")
        vars_checked_count = 0
        for var in all_vars:
            vars_checked_count += 1
            mapped_industry = get_var_attribute(var, var_industry_map, default_value=None)
            if mapped_industry == industry:
                industry_vars.append(var)
        # logger.info(f"-- [ç»˜å›¾è°ƒè¯•] è¡Œä¸š '{industry}' æŸ¥æ‰¾ç»“æŸ. å…±æ£€æŸ¥ {vars_checked_count} ä¸ªå˜é‡. æ‰¾åˆ° {len(industry_vars)} ä¸ªåŒ¹é…å˜é‡: {industry_vars[:10]}... --")

        if not industry_vars:
             # Handle case where no variables are found (same as before)
             logger.warning(f"è¡Œä¸š '{industry}' æœªæ‰¾åˆ°å¯¹åº”å˜é‡ï¼Œåœ¨å›¾ä¸­ä¿ç•™ç©ºä½ã€‚")
             ax.set_title(f"è¡Œä¸š: {industry}\n(æ— å˜é‡æ•°æ®)")
             ax.tick_params(axis='both', which='both', bottom=False, top=False, left=False, right=False, labelbottom=False, labelleft=False)
        else:
             # --- <<< START MAJOR CHANGE: Process data_original_aligned using Log ONLY >>> ---
             # logger.debug(f"  è¡Œä¸š '{industry}': å¯¹æ‰¾åˆ°çš„ {len(industry_vars)} ä¸ªåŸå§‹å˜é‡æ‰§è¡Œ Log ...")
             industry_vars_processed_temp = pd.DataFrame(index=data_original_aligned.index) # Initialize empty df with correct index
             processed_vars_count = 0
             for var in industry_vars:
                 if var in data_original_aligned.columns:
                     series = data_original_aligned[var].copy()
                     
                     # 1. Log transform (handle non-positive values)
                     series_log = series.copy()
                     non_positive_mask = series_log <= 0
                     if non_positive_mask.any():
                         # logger.warning(f"    å˜é‡ '{var}' åŒ…å« {non_positive_mask.sum()} ä¸ªéæ­£å€¼ï¼Œå°†åœ¨ Log è½¬æ¢ä¸­è®¾ä¸º NaNã€‚")
                         series_log[non_positive_mask] = np.nan
                     series_log = np.log(series_log)
                     
                     # Remove the differencing step
                     # series_log_diff = series_log.diff(1)
                     
                     # Check if result is valid before adding
                     if not series_log.isnull().all(): # Check log result directly
                         industry_vars_processed_temp[var] = series_log # Add log-transformed series
                         processed_vars_count += 1
                     # else:
                         # logger.warning(f"    å˜é‡ '{var}' Log è½¬æ¢åç»“æœå…¨ä¸º NaNï¼Œè·³è¿‡æ·»åŠ ã€‚")

             # logger.debug(f"  è¡Œä¸š '{industry}': æˆåŠŸå¤„ç†äº† {processed_vars_count} / {len(industry_vars)} ä¸ªå˜é‡ã€‚")

             # Check if any variables were successfully processed
             if industry_vars_processed_temp.empty or industry_vars_processed_temp.isnull().all().all():
                  logger.warning(f"  è¡Œä¸š '{industry}': Log è½¬æ¢åæ²¡æœ‰æœ‰æ•ˆçš„å˜é‡æ•°æ®ï¼Œè·³è¿‡ç»˜å›¾ã€‚")
                  ax.set_title(f"è¡Œä¸š: {industry}\n(å¤„ç†åæ— æœ‰æ•ˆæ•°æ®)")
                  ax.tick_params(axis='both', which='both', bottom=False, top=False, left=False, right=False, labelbottom=False, labelleft=False)
             else:
                 # --- ç§»é™¤è®¡ç®—å¤„ç†åè¡Œä¸šå¹³å‡åºåˆ—å’Œç»Ÿè®¡é‡çš„éƒ¨åˆ† --- 
                 # ... [Removed code for calculating mean/std] ...

                 # --- ç»˜åˆ¶å¤„ç†åçš„è¡Œä¸šå˜é‡æ›²çº¿ (ç°è‰²) on primary axis --- 
                 lines_ax1 = []
                 labels_ax1 = []
                 num_plotted = 0
                 # logger.debug(f"    [Plot Loop - {industry}] å‡†å¤‡ç»˜åˆ¶ Log å˜é‡çº¿...") # <-- æ–°å¢
                 for var in industry_vars_processed_temp.columns: # Iterate through successfully processed vars
                     line, = ax.plot(industry_vars_processed_temp.index, industry_vars_processed_temp[var],
                                     color='grey', alpha=0.7, linewidth=1.0)
                     if num_plotted == 0: # Only add label once for grey lines
                         lines_ax1.append(line)
                         labels_ax1.append("Log(å˜é‡)")
                     num_plotted += 1
                 # logger.debug(f"    [Plot Loop - {industry}] å®Œæˆç»˜åˆ¶ Log å˜é‡çº¿ ({num_plotted} æ¡)ã€‚") # <-- æ–°å¢
                 # --- ç»“æŸç»˜åˆ¶å˜é‡æ›²çº¿ --- 

                 # --- <<< ä¿®æ”¹ï¼šåˆ›å»ºæ¬¡è½´å¹¶ç»˜åˆ¶åŸå§‹å› å­ >>> ---
                 # logger.debug(f"    [Plot Loop - {industry}] å‡†å¤‡åˆ›å»ºæ¬¡è½´å¹¶ç»˜åˆ¶å› å­...") # <-- æ–°å¢
                 ax2 = ax.twinx() # Create secondary Y-axis (Right)
                 lines_ax2 = []
                 labels_ax2 = []
                 if driving_factor in factors_ts.columns:
                     factor_color = factor_color_map.get(driving_factor, 'black')
                     factor_ts_series = factors_ts[driving_factor].copy()
                     
                     # Plot the original factor directly on the secondary axis
                     if not factor_ts_series.isnull().all():
                         line, = ax2.plot(factor_ts_series.index, factor_ts_series,
                                          color=factor_color, linestyle='-',
                                          linewidth=1.5, label=f"{driving_factor} (å³è½´)") # Update label
                         lines_ax2.append(line)
                         labels_ax2.append(f"{driving_factor} (å³è½´)")
                         # logger.debug(f"      å› å­ '{driving_factor}' å·²ç»˜åˆ¶åœ¨å³è½´ã€‚")
                         ax2.set_ylabel(f"å› å­å€¼", fontsize=9, color=factor_color)
                         ax2.tick_params(axis='y', labelcolor=factor_color)
                     else:
                         # logger.warning(f"      å› å­ '{driving_factor}' å…¨ä¸º NaNï¼Œæ— æ³•ç»˜åˆ¶ã€‚")
                         ax2.set_ylabel(f"å› å­å€¼", fontsize=9)
                         ax2.tick_params(axis='y')
                 else:
                     # logger.warning(f"    è¡Œä¸š '{industry}' çš„é©±åŠ¨å› å­ '{driving_factor}' ä¸åœ¨å› å­æ—¶é—´åºåˆ—ä¸­ã€‚")
                     ax2.set_ylabel(f"å› å­å€¼", fontsize=9)
                     ax2.tick_params(axis='y')
                 # logger.debug(f"    [Plot Loop - {industry}] å®Œæˆæ¬¡è½´å’Œå› å­ç»˜åˆ¶ã€‚") # <-- æ–°å¢
                 # --- <<< ç»“æŸæ¬¡è½´å’Œå› å­ç»˜åˆ¶ >>> ---

                 # --- è®¾ç½®æ­¤å­å›¾çš„æ ‡é¢˜ã€æ ‡ç­¾ã€ç½‘æ ¼ã€å›¾ä¾‹ --- 
                 ax.set_title(f"è¡Œä¸š: {industry}\n(é©±åŠ¨å› å­: {driving_factor}) R2={industry_max_r2[industry]:.2f}")
                 # --- ä¿®æ”¹ä¸» Y è½´æ ‡ç­¾ ---
                 ax.set_ylabel("Log(å˜é‡å€¼) (å·¦è½´)", fontsize=9)
                 # --- ç»“æŸä¿®æ”¹ä¸» Y è½´æ ‡ç­¾ ---
                 ax.grid(True, linestyle=':', alpha=0.5)
                 # --- åˆå¹¶å›¾ä¾‹ --- 
                 lines = lines_ax1 + lines_ax2
                 labels = labels_ax1 + labels_ax2
                 # logger.debug(f"    [Plot Loop - {industry}] å‡†å¤‡è®¾ç½®å›¾ä¾‹...") # <-- æ–°å¢
                 ax.legend(lines, labels, loc='upper right', fontsize='small')
                 plotted_industries_count += 1
                 # logger.debug(f"    [Plot Loop - {industry}] å›¾ä¾‹è®¾ç½®å®Œæˆã€‚") # <-- æ–°å¢
                 # --- ç»“æŸè®¾ç½®å­å›¾ --- 
             # --- <<< END MAJOR CHANGE >>> --- 
        logger.info(f"  [Plot Loop] å®Œæˆå¤„ç†è¡Œä¸š: {industry}") # <-- æ–°å¢æ—¥å¿—
        # æ— è®ºæ˜¯å¦ç»˜åˆ¶äº†æ•°æ®ï¼Œéƒ½è¦ç§»åŠ¨åˆ°ä¸‹ä¸€ä¸ªå­å›¾ä½ç½®
        plot_idx += 1

    # 5. éšè—æœªä½¿ç”¨çš„å­å›¾è½´
    logger.info("éšè—æœªä½¿ç”¨çš„å­å›¾è½´...") # <-- æ–°å¢æ—¥å¿—
    for i in range(plot_idx, len(axes)):
        axes[i].axis('off')

    # 6. è°ƒæ•´æ•´ä½“å¸ƒå±€å’Œæ·»åŠ ä¸»æ ‡é¢˜
    # --- ä¿®æ”¹ä¸»æ ‡é¢˜ ---
    logger.info("è®¾ç½®ä¸»æ ‡é¢˜...") # <-- æ–°å¢æ—¥å¿—
    plt.suptitle("å„è¡Œä¸š(Log)ä¸ä¸»è¦é©±åŠ¨å› å­(åŸå§‹å€¼)å¯¹æ¯”", fontsize=16, y=1.02)
    # --- ç»“æŸä¿®æ”¹ä¸»æ ‡é¢˜ ---
    logger.info("è°ƒæ•´å¸ƒå±€ (tight_layout)...") # <-- æ–°å¢æ—¥å¿—
    plt.tight_layout(rect=[0, 0.03, 1, 0.98]) # Adjust layout to prevent title overlap
    logger.info("å¸ƒå±€è°ƒæ•´å®Œæˆã€‚") # <-- æ–°å¢æ—¥å¿—

    # 7. ä¿å­˜å›¾åƒ
    logger.info(f"å³å°†ä¿å­˜è¡Œä¸šé©±åŠ¨å› å­å›¾åˆ°: {save_path}") # <-- æ–°å¢æ—¥å¿— (é‡è¦)
    try:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        logger.info(f"è¡Œä¸šä¸é©±åŠ¨å› å­å¯¹æ¯”å›¾å·²ä¿å­˜è‡³: {save_path}") # <-- è¿™æ˜¯æˆåŠŸä¿å­˜çš„æ¶ˆæ¯
    except Exception as e:
        logger.error(f"ä¿å­˜è¡Œä¸šä¸é©±åŠ¨å› å­å¯¹æ¯”å›¾æ—¶å‡ºé”™: {e}")
    finally:
        logger.info("å…³é—­ç»˜å›¾å¯¹è±¡...") # <-- æ–°å¢æ—¥å¿—
        plt.close(fig) # å…³é—­å›¾å½¢ï¼Œé‡Šæ”¾å†…å­˜

    logger.info("ç»˜åˆ¶è¡Œä¸šä¸é©±åŠ¨å› å­å¯¹æ¯”å›¾å®Œæˆã€‚")
# --- <<< ç»“æŸæ–°å¢/ä¿®æ”¹ >>> ---

# <<<-------------------- æ–°å¢å› å­è½½è·èšç±»çƒ­åŠ›å›¾å‡½æ•° -------------------->>>
def plot_factor_loading_clustermap(
    loadings_df: pd.DataFrame,
    title: str,
    filename: str,
    figsize: Tuple[int, int] = (12, 10), # è°ƒæ•´é»˜è®¤å¤§å°
    cmap: str = "coolwarm", # <<< ä¿®æ”¹ï¼šé»˜è®¤ä½¿ç”¨å‘æ•£è‰²å›¾
    annot: Optional[bool] = None, # <<< ä¿®æ”¹ï¼šå…è®¸è‡ªåŠ¨æˆ–æ‰‹åŠ¨è®¾ç½®
    fmt: str = ".2f", # æ•°å€¼æ ¼å¼
    row_cluster: bool = True,
    col_cluster: bool = True,
    top_n_vars: Optional[int] = None, # <<< æ–°å¢ï¼šç”¨äºç­›é€‰ Top N å˜é‡
    center: Optional[float] = 0 # <<< æ–°å¢ï¼šè®¾ç½®é¢œè‰²ä¸­å¿ƒ
):
    """
    ç”Ÿæˆå› å­è½½è·çš„èšç±»çƒ­åŠ›å›¾ (Clustermap)ï¼Œå¯é€‰åªæ˜¾ç¤º Top N å˜é‡ã€‚

    Args:
        loadings_df (pd.DataFrame): å› å­è½½è· DataFrame (è¡Œä¸ºå˜é‡, åˆ—ä¸ºå› å­)ã€‚
        title (str): å›¾è¡¨æ ‡é¢˜ã€‚
        filename (str): ä¿å­˜å›¾è¡¨çš„æ–‡ä»¶è·¯å¾„ã€‚
        figsize (Tuple[int, int], optional): å›¾è¡¨å°ºå¯¸. Defaults to (12, 10).
        cmap (str, optional): Matplotlib è‰²å›¾åç§°. Defaults to "coolwarm".
        annot (Optional[bool], optional): æ˜¯å¦åœ¨å•å…ƒæ ¼ä¸­æ˜¾ç¤ºæ•°å€¼ã€‚å¦‚æœä¸º Noneï¼Œåˆ™åœ¨ top_n_vars ç”Ÿæ•ˆæ—¶è‡ªåŠ¨è®¾ä¸º Trueã€‚
        fmt (str, optional): æ•°å€¼æ˜¾ç¤ºæ ¼å¼. Defaults to ".2f".
        row_cluster (bool, optional): æ˜¯å¦å¯¹è¡Œè¿›è¡Œèšç±». Defaults to True.
        col_cluster (bool, optional): æ˜¯å¦å¯¹åˆ—è¿›è¡Œèšç±». Defaults to True.
        top_n_vars (Optional[int], optional): å¦‚æœæŒ‡å®šï¼Œåˆ™åªæ˜¾ç¤ºæ¯ä¸ªå› å­ä¸‹è½½è·ç»å¯¹å€¼æœ€é«˜çš„ N ä¸ªå˜é‡ï¼ˆåˆå¹¶å»é‡åçš„å­é›†ï¼‰ã€‚Defaults to None.
        center (Optional[float], optional): è‰²å›¾çš„ä¸­å¿ƒå€¼ã€‚å¯¹äºå‘æ•£è‰²å›¾ï¼Œé€šå¸¸è®¾ä¸º 0ã€‚Defaults to 0.
    """
    logger.info(f"\\n[ç»˜å›¾å‡½æ•°] å¼€å§‹ç”Ÿæˆå› å­è½½è·èšç±»çƒ­åŠ›å›¾: {filename}...")
    if loadings_df.empty:
        logger.warning("å› å­è½½è·æ•°æ®ä¸ºç©ºï¼Œæ— æ³•ç”Ÿæˆèšç±»çƒ­åŠ›å›¾ã€‚")
        return

    plot_df = loadings_df.copy()

    # --- <<< æ–°å¢ï¼šç­›é€‰ Top N å˜é‡ >>> ---
    if top_n_vars is not None and top_n_vars > 0:
        logger.info(f"ç­›é€‰æ¯ä¸ªå› å­ Top {top_n_vars} ç»å¯¹è½½è·çš„å˜é‡...")
        top_vars_set = set()
        for factor in plot_df.columns:
            try:
                # è®¡ç®—ç»å¯¹è½½è·å¹¶æ‰¾åˆ° Top N çš„ç´¢å¼•ï¼ˆå˜é‡åï¼‰
                top_indices = plot_df[factor].abs().nlargest(top_n_vars).index
                top_vars_set.update(top_indices)
            except Exception as e_topn:
                logger.warning(f"ä¸ºå› å­ {factor} ç­›é€‰ Top {top_n_vars} å˜é‡æ—¶å‡ºé”™: {e_topn}")
        
        if not top_vars_set:
            logger.error("æœªèƒ½æˆåŠŸç­›é€‰å‡ºä»»ä½• Top N å˜é‡ï¼Œæ— æ³•ç”Ÿæˆè¿‡æ»¤åçš„çƒ­åŠ›å›¾ã€‚")
            return
            
        filtered_vars = sorted(list(top_vars_set))
        logger.info(f"ç­›é€‰åå…±ä¿ç•™ {len(filtered_vars)} ä¸ªå”¯ä¸€å˜é‡ç”¨äºç»˜å›¾ã€‚")
        plot_df = plot_df.loc[filtered_vars]
        
        # å¦‚æœç­›é€‰ç”Ÿæ•ˆï¼Œä¸” annot æœªè¢«æ˜¾å¼è®¾ä¸º Falseï¼Œåˆ™è‡ªåŠ¨å¼€å¯ annot
        if annot is None:
            annot = True
            logger.info("ç”±äºå·²ç­›é€‰ Top N å˜é‡ï¼Œè‡ªåŠ¨å¯ç”¨æ•°å€¼æ ‡æ³¨ (annot=True)ã€‚")
        # è°ƒæ•´ figsize (å¯é€‰ï¼Œè¡Œæ•°å‡å°‘ï¼Œå¯ä»¥å‡å°‘é«˜åº¦)
        # new_height = max(6, min(15, len(filtered_vars) * 0.3)) # åŠ¨æ€è°ƒæ•´é«˜åº¦ç¤ºä¾‹
        # figsize = (figsize[0], new_height)
        # logger.info(f"è°ƒæ•´ figsize ä¸º: {figsize}")
        
    elif annot is None: # å¦‚æœæœªç­›é€‰ä¸” annot ä¸º Noneï¼Œåˆ™æ ¹æ®æ€»è¡Œæ•°åˆ¤æ–­æ˜¯å¦å¼€å¯
         if len(plot_df) <= 50: # è¡Œæ•°ä¸å¤šæ—¶é»˜è®¤å¼€å¯ annot
              annot = True
         else: # è¡Œæ•°å¤šæ—¶é»˜è®¤å…³é—­
              annot = False
              logger.info("å˜é‡æ•°é‡è¾ƒå¤š (>50) ä¸”æœªæŒ‡å®š top_n_varsï¼Œé»˜è®¤ç¦ç”¨æ•°å€¼æ ‡æ³¨ã€‚")
    # --- <<< ç»“æŸç­›é€‰ >>> ---

    try:
        # æ£€æŸ¥è¡Œæ•°ï¼Œå¦‚æœæœ€ç»ˆè¡Œæ•°è¿‡å¤šä¸” annot ä»ä¸º Trueï¼Œç»™å‡ºè­¦å‘Š (è™½ç„¶ç­›é€‰åä¸å¤ªå¯èƒ½)
        if len(plot_df) > 75 and annot:
            logger.warning("å˜é‡æ•°é‡ä»ç„¶è¾ƒå¤š (>75)ï¼Œæ•°å€¼æ ‡æ³¨å¯èƒ½ä¼šé‡å ã€‚")
            # annot = False # å¯ä»¥è€ƒè™‘å¼ºåˆ¶å…³é—­

        # ä½¿ç”¨ seaborn çš„ clustermap åŒæ—¶è¿›è¡Œèšç±»å’Œç»˜å›¾
        cluster_map = sns.clustermap(
            plot_df, # ä½¿ç”¨ç­›é€‰åçš„ DataFrame
            figsize=figsize,
            cmap=cmap,
            annot=annot, # ä½¿ç”¨æœ€ç»ˆç¡®å®šçš„ annot å€¼
            fmt=fmt,
            center=center, # <<< æ–°å¢ï¼šè®¾ç½®é¢œè‰²ä¸­å¿ƒ
            linewidths=.5,
            linecolor='lightgray',
            row_cluster=row_cluster,
            col_cluster=col_cluster,
            dendrogram_ratio=(.15, .2) # è°ƒæ•´æ ‘çŠ¶å›¾æ¯”ä¾‹ï¼Œè¡Œå‡å°‘å¯ä»¥é€‚å½“å‡å°è¡Œæ ‘çŠ¶å›¾æ¯”ä¾‹
        )

        # è°ƒæ•´å­—ä½“å¤§å°
        plt.setp(cluster_map.ax_heatmap.get_xticklabels(), rotation=45, ha='right', fontsize=8)
        # è°ƒæ•´ Y è½´æ ‡ç­¾å­—ä½“å¤§å°ï¼Œæ ¹æ®è¡Œæ•°å†³å®šï¼Œè¡Œå°‘å­—å¯ä»¥å¤§ç‚¹
        ytick_fontsize = 8 if len(plot_df) > 30 else 10
        plt.setp(cluster_map.ax_heatmap.get_yticklabels(), rotation=0, fontsize=ytick_fontsize)

        # æ·»åŠ ä¸»æ ‡é¢˜
        plt.suptitle(title, y=1.03, fontsize=14) # è°ƒæ•´ y ä½ç½®

        # ä¿å­˜å›¾åƒ
        cluster_map.savefig(filename, dpi=300, bbox_inches='tight')
        logger.info(f"å› å­è½½è·èšç±»çƒ­åŠ›å›¾å·²ä¿å­˜è‡³: {filename}")

        # å…³é—­å›¾å½¢ï¼Œé‡Šæ”¾å†…å­˜
        plt.close('all')

    except Exception as e:
        logger.error(f"ç”Ÿæˆå› å­è½½è·èšç±»çƒ­åŠ›å›¾æ—¶å‡ºé”™: {e}", exc_info=True)
        plt.close('all')

# <<<-------------------- ç»“æŸæ–°å¢/ä¿®æ”¹å‡½æ•° -------------------->>>

# <<< --- æ–°å¢å› å­è½½è·ç¨³å®šæ€§å¯¹æ¯”å‡½æ•° --- >>>
def plot_aligned_loading_comparison(
    lambda_full: pd.DataFrame,
    lambda_train: pd.DataFrame,
    variables: List[str],
    output_path: str,
    threshold: float = 0.1,
    figsize: Tuple[int, int] = (15, 25) # è°ƒæ•´ figsize ä»¥é€‚åº”æ›´å¤šå˜é‡
) -> None:
    """
    å¯¹é½ä¸¤ä¸ªå› å­è½½è·çŸ©é˜µï¼ˆä¾‹å¦‚ï¼Œå…¨æ ·æœ¬ vs ä»…è®­ç»ƒæœŸï¼‰å¹¶ç»˜åˆ¶å¯¹æ¯”å›¾ã€‚

    Args:
        lambda_full: å…¨æ ·æœ¬ï¼ˆæˆ–åŸºå‡†ï¼‰å› å­è½½è· DataFrame (Variables x Factors)ã€‚
        lambda_train: è®­ç»ƒæœŸï¼ˆæˆ–æ¯”è¾ƒå¯¹è±¡ï¼‰å› å­è½½è· DataFrame (Variables x Factors)ã€‚
        variables: å˜é‡ååˆ—è¡¨ï¼Œåº”ä¸è½½è·çŸ©é˜µçš„ç´¢å¼•åŒ¹é…ã€‚
        output_path: å›¾åƒè¾“å‡ºè·¯å¾„ã€‚
        threshold: ä»…ç»˜åˆ¶ç»å¯¹è½½è·å€¼å¤§äºæ­¤é˜ˆå€¼çš„å˜é‡ã€‚
        figsize: Matplotlib å›¾å½¢å¤§å°ã€‚
    """
    logger = logging.getLogger(__name__) # Use module logger or get a new one
    logger.info(f"å¼€å§‹ç»˜åˆ¶å› å­è½½è·ç¨³å®šæ€§å¯¹æ¯”å›¾ï¼Œä¿å­˜è‡³: {output_path}")

    # --- æ•°æ®éªŒè¯å’Œå‡†å¤‡ ---
    if not isinstance(lambda_full, pd.DataFrame):
        # å¦‚æœæ˜¯ numpy array, å°è¯•è½¬æ¢
        if isinstance(lambda_full, np.ndarray) and lambda_full.ndim == 2 and lambda_full.shape[0] == len(variables):
            logger.warning("lambda_full æ˜¯ NumPy æ•°ç»„ï¼Œå°†å°è¯•è½¬æ¢ä¸º DataFrameã€‚")
            try:
                lambda_full = pd.DataFrame(lambda_full, index=variables, columns=[f"F{i+1}_Full_tmp" for i in range(lambda_full.shape[1])])
            except Exception as e_conv:
                logger.error(f"è½¬æ¢ lambda_full åˆ° DataFrame å¤±è´¥: {e_conv}")
                return
        else:
            logger.error("å…¨æ ·æœ¬è½½è· lambda_full å¿…é¡»æ˜¯ Pandas DataFrame æˆ–å¯è½¬æ¢çš„ NumPy æ•°ç»„ã€‚")
            return

    if not isinstance(lambda_train, pd.DataFrame):
        if isinstance(lambda_train, np.ndarray) and lambda_train.ndim == 2 and lambda_train.shape[0] == len(variables):
            logger.warning("lambda_train æ˜¯ NumPy æ•°ç»„ï¼Œå°†å°è¯•è½¬æ¢ä¸º DataFrameã€‚")
            try:
                lambda_train = pd.DataFrame(lambda_train, index=variables, columns=[f"F{i+1}_Train_tmp" for i in range(lambda_train.shape[1])])
            except Exception as e_conv:
                 logger.error(f"è½¬æ¢ lambda_train åˆ° DataFrame å¤±è´¥: {e_conv}")
                 return
        else:
            logger.error("è®­ç»ƒæœŸè½½è· lambda_train å¿…é¡»æ˜¯ Pandas DataFrame æˆ–å¯è½¬æ¢çš„ NumPy æ•°ç»„ã€‚")
            return

    # å†æ¬¡æ£€æŸ¥è¡Œæ•°ä¸å˜é‡æ•°
    if lambda_full.shape[0] != len(variables) or lambda_train.shape[0] != len(variables):
         logger.error(f"è½¬æ¢åï¼Œè½½è·çŸ©é˜µçš„è¡Œæ•°ä»ä¸å˜é‡åˆ—è¡¨é•¿åº¦ä¸åŒ¹é…ã€‚ "
                      f"Lambda Full: {lambda_full.shape[0]}, Lambda Train: {lambda_train.shape[0]}, Variables: {len(variables)}")
         return

    # ç¡®ä¿ç´¢å¼•æ˜¯å˜é‡å
    lambda_full.index = variables
    lambda_train.index = variables

    k_factors = lambda_full.shape[1]
    if lambda_train.shape[1] != k_factors:
        logger.warning(f"ä¸¤ä¸ªè½½è·çŸ©é˜µçš„å› å­æ•°é‡ä¸åŒ¹é… (Full: {k_factors}, Train: {lambda_train.shape[1]})ã€‚å°†å°è¯•ä½¿ç”¨è¾ƒå°‘çš„å› å­æ•°é‡è¿›è¡Œå¯¹é½ã€‚")
        k_factors = min(k_factors, lambda_train.shape[1])
        lambda_full = lambda_full.iloc[:, :k_factors].copy() # Use .copy()
        lambda_train = lambda_train.iloc[:, :k_factors].copy() # Use .copy()

    # é‡å‘½ååˆ—ä»¥æ˜ç¡®æ¥æº
    lambda_full.columns = [f"Factor{i+1}_Full" for i in range(k_factors)]
    lambda_train.columns = [f"Factor{i+1}_Train" for i in range(k_factors)]

    # --- å› å­å¯¹é½ (ä½¿ç”¨åŒˆç‰™åˆ©ç®—æ³•) ---
    lambda_train_aligned = None # Initialize
    try:
        # è®¡ç®—æˆæœ¬çŸ©é˜µ (è´Ÿç›¸å…³æ€§ï¼Œå› ä¸ºæˆ‘ä»¬è¦æœ€å¤§åŒ–ç›¸å…³æ€§)
        # Ensure no NaN values interfere with correlation calculation
        common_index = lambda_full.dropna().index.intersection(lambda_train.dropna().index)
        if len(common_index) < len(variables):
            logger.warning(f"è½½è·çŸ©é˜µåŒ…å« NaN å€¼ï¼Œä»…åœ¨ {len(common_index)} ä¸ªå…±åŒé NaN å˜é‡ä¸Šè®¡ç®—ç›¸å…³æ€§ã€‚")
        if len(common_index) < 2:
             logger.error("è®¡ç®—ç›¸å…³æ€§çš„å…±åŒæœ‰æ•ˆå˜é‡ä¸è¶³ (<2)ã€‚æ— æ³•æ‰§è¡Œå› å­å¯¹é½ã€‚")
             raise ValueError("Not enough common valid variables for correlation.")
             
        lambda_full_common = lambda_full.loc[common_index]
        lambda_train_common = lambda_train.loc[common_index]
        
        correlation_matrix = np.corrcoef(lambda_full_common.T, lambda_train_common.T)
        # æå– Full vs Train çš„ç›¸å…³æ€§éƒ¨åˆ†
        cost_matrix = -correlation_matrix[:k_factors, k_factors:]

        # åº”ç”¨åŒˆç‰™åˆ©ç®—æ³•æ‰¾åˆ°æœ€ä½³åŒ¹é…
        row_ind, col_ind = linear_sum_assignment(cost_matrix)
        logger.info(f"å› å­å¯¹é½å®Œæˆã€‚åŒ¹é…ç´¢å¼• (Full -> Train): {list(zip(row_ind, col_ind))}")

        # æ ¹æ®åŒ¹é…ç»“æœé‡æ–°æ’åˆ—è®­ç»ƒæœŸå› å­åˆ—
        aligned_train_cols = [f"Factor{j+1}_Train" for j in col_ind]
        lambda_train_aligned = lambda_train[aligned_train_cols].copy()
        
        # --- ç¿»è½¬ç¬¦å·ä»¥åŒ¹é… --- 
        # æ£€æŸ¥åŒ¹é…å› å­å¯¹ä¹‹é—´çš„ä¸»è¦å˜é‡è½½è·ç¬¦å·æ˜¯å¦ä¸€è‡´
        for full_idx, train_aligned_idx in zip(row_ind, col_ind):
            full_factor_name = f"Factor{full_idx+1}_Full"
            train_factor_name = f"Factor{train_aligned_idx+1}_Train" # Original column name from lambda_train
            train_aligned_col_name = f"Factor{full_idx+1}_TrainAligned" # Target column name in lambda_train_aligned
            
            # è®¡ç®—è¿™å¯¹å› å­åœ¨å…±åŒå˜é‡ä¸Šçš„ç›¸å…³æ€§
            corr_val = np.corrcoef(lambda_full_common[full_factor_name], lambda_train_common[train_factor_name])[0, 1]
            logger.debug(f"æ£€æŸ¥å¯¹é½: {full_factor_name} vs {train_factor_name} (åŸå§‹åˆ— {train_aligned_idx+1}), ç›¸å…³æ€§: {corr_val:.3f}")
            if corr_val < 0:
                 logger.info(f"  å› å­ {full_idx+1} (Full) ä¸å…¶åŒ¹é…çš„è®­ç»ƒæœŸå› å­ {train_aligned_idx+1} å‘ˆè´Ÿç›¸å…³ã€‚å°†ç¿»è½¬è®­ç»ƒæœŸå› å­ {train_aligned_idx+1} çš„ç¬¦å·ã€‚")
                 lambda_train_aligned.iloc[:, list(col_ind).index(train_aligned_idx)] *= -1 # ç¿»è½¬å¯¹é½å DF ä¸­å¯¹åº”çš„åˆ—
        # --- ç»“æŸç¬¦å·ç¿»è½¬ --- 
        
        # é‡å‘½åå¯¹é½åçš„è®­ç»ƒæœŸå› å­åˆ—ï¼Œä»¥åŒ¹é…å…¨æ ·æœ¬å› å­ç¼–å·ï¼Œå¹¶æ·»åŠ åç¼€
        lambda_train_aligned.columns = [f"Factor{i+1}_TrainAligned" for i in row_ind]
        # ç¡®ä¿é¡ºåºä¸ full ä¸€è‡´ (é‡è¦ï¼å› ä¸º row_ind å¯èƒ½ä¸æ˜¯ 0, 1, 2...)
        final_aligned_cols = [f"Factor{i+1}_TrainAligned" for i in range(k_factors)]
        lambda_train_aligned = lambda_train_aligned.reindex(columns=final_aligned_cols)
        logger.debug(f"Aligned and renamed train columns: {lambda_train_aligned.columns.tolist()}")

    except Exception as e_align:
        logger.error(f"å› å­å¯¹é½è¿‡ç¨‹ä¸­å‡ºé”™: {e_align}", exc_info=True)
        # å¦‚æœå¯¹é½å¤±è´¥ï¼Œåˆ™ä¸è¿›è¡Œå¯¹é½ï¼Œç›´æ¥ä½¿ç”¨åŸå§‹é¡ºåºæ¯”è¾ƒ
        lambda_train_aligned = lambda_train.copy()
        lambda_train_aligned.columns = [f"{col.replace('_Train', '')}_Train_Unaligned" for col in lambda_train_aligned.columns] # Add suffix
        logger.warning("ç”±äºå¯¹é½å¤±è´¥æˆ–é”™è¯¯ï¼Œå°†æŒ‰åŸå§‹é¡ºåºæ¯”è¾ƒå› å­ (åˆ—åå·²æ·»åŠ  '_Unaligned' åç¼€)ã€‚")


    # --- ç»˜å›¾ --- 
    n_factors = k_factors
    n_cols = 2 # æ¯è¡Œæ”¾ä¸¤ä¸ªå› å­å¯¹æ¯”å›¾
    n_rows = (n_factors + n_cols - 1) // n_cols

    fig, axes = plt.subplots(n_rows, n_cols, figsize=figsize, squeeze=False) # squeeze=False ä¿è¯ axes æ˜¯äºŒç»´æ•°ç»„
    fig.suptitle('Factor Loading Comparison (Full Sample vs. Training Sample - Aligned)', fontsize=16, y=1.02)

    for i in range(n_factors):
        row = i // n_cols
        col = i % n_cols
        ax = axes[row, col]

        factor_label = f"Factor{i+1}"
        full_col = f"{factor_label}_Full"
        # --- ä¿®æ”¹ï¼šå¤„ç†å¯¹é½å¤±è´¥çš„åˆ—å --- 
        aligned_train_col_name_aligned = f"{factor_label}_TrainAligned"
        aligned_train_col_name_unaligned = f"{factor_label}_Train_Unaligned"
        
        if aligned_train_col_name_aligned in lambda_train_aligned.columns:
            aligned_train_col = aligned_train_col_name_aligned
            train_label = "Training Sample (Aligned)"
        elif aligned_train_col_name_unaligned in lambda_train_aligned.columns:
            aligned_train_col = aligned_train_col_name_unaligned
            train_label = "Training Sample (Unaligned)"
        else:
             logger.error(f"æ— æ³•åœ¨ lambda_train_aligned ä¸­æ‰¾åˆ°å› å­ {i+1} çš„åˆ—ã€‚å¯ç”¨åˆ—: {lambda_train_aligned.columns.tolist()}")
             ax.text(0.5, 0.5, f'Error: Cannot find\ncolumn for Factor {i+1}\nin aligned train data.', ha='center', va='center', fontsize=10, color='red')
             ax.set_title(f'{factor_label} Comparison - Error')
             continue # è·³è¿‡è¿™ä¸ªå­å›¾
        # --- ç»“æŸä¿®æ”¹ ---
        
        # åˆå¹¶ä¸¤ä¸ª Series å¹¶ç­›é€‰
        comparison_df = pd.DataFrame({
            'Full Sample': lambda_full[full_col],
            train_label: lambda_train_aligned[aligned_train_col]
        })
        # ç­›é€‰ç»å¯¹å€¼å¤§äºé˜ˆå€¼çš„å˜é‡ (åœ¨ä»»ä¸€æ ·æœ¬ä¸­)
        filtered_df = comparison_df[
            (comparison_df['Full Sample'].abs() > threshold) |
            (comparison_df[train_label].abs() > threshold)
        ].copy() # Use .copy() to avoid SettingWithCopyWarning
        
        # æŒ‰å…¨æ ·æœ¬è½½è·çš„ç»å¯¹å€¼é™åºæ’åº (å¦‚æœå…¨æ ·æœ¬è½½è·ä¸º NaN åˆ™ç§»åˆ°åé¢)
        filtered_df['abs_full'] = filtered_df['Full Sample'].abs()
        filtered_df = filtered_df.sort_values(by='abs_full', ascending=False, na_position='last')
        filtered_df = filtered_df.drop(columns='abs_full')
        
        if filtered_df.empty:
            ax.text(0.5, 0.5, 'No variables above threshold', ha='center', va='center', fontsize=12)
            ax.set_title(f'{factor_label} Comparison')
            ax.set_xticks([])
            ax.set_yticks([])
            continue

        # ç»˜åˆ¶å¹¶æ’æ¡å½¢å›¾
        filtered_df.plot(kind='barh', ax=ax, width=0.8)

        ax.set_title(f'{factor_label} Comparison')
        ax.set_xlabel('Loading Value')
        ax.set_ylabel('Variable')
        ax.invert_yaxis() # è®©ç»å¯¹å€¼æœ€å¤§çš„åœ¨ä¸Šé¢
        ax.legend(title='Sample Period')
        ax.grid(axis='x', linestyle='--', alpha=0.7)
        ax.axvline(0, color='black', linewidth=0.8) # Add vertical line at 0

    # å¦‚æœå› å­æ•°é‡ä¸æ˜¯ n_cols çš„æ•´æ•°å€ï¼Œéšè—å¤šä½™çš„å­å›¾
    for j in range(n_factors, n_rows * n_cols):
        row = j // n_cols
        col = j % n_cols
        if row < axes.shape[0] and col < axes.shape[1]: # Check bounds
            fig.delaxes(axes[row, col])

    plt.tight_layout(rect=[0, 0, 1, 1.0]) # è°ƒæ•´å¸ƒå±€ä»¥é€‚åº”æ ‡é¢˜
    try:
        plt.savefig(output_path, bbox_inches='tight', dpi=150) # å¢åŠ  DPI
        logger.info(f"å› å­è½½è·ç¨³å®šæ€§å¯¹æ¯”å›¾å·²æˆåŠŸä¿å­˜è‡³: {output_path}")
    except Exception as e:
        logger.error(f"ä¿å­˜å› å­è½½è·ç¨³å®šæ€§å¯¹æ¯”å›¾æ—¶å‡ºé”™: {e}", exc_info=True)
    finally:
        plt.close(fig) # å…³é—­å›¾å½¢ï¼Œé‡Šæ”¾å†…å­˜

# <<< --- ç»“æŸæ–°å¢ --- >>>

# <<< æ–°å¢ï¼šå°† write_single_table ç§»åˆ°æ¨¡å—çº§åˆ« >>>
def write_single_table(ws, df, title, start_r, start_c, bold_f, number_format='0.0000'):
    """è¾…åŠ©å‡½æ•°ï¼šå°†å•ä¸ª DataFrame å†™å…¥ Excel Sheet ä¸­çš„æŒ‡å®šä½ç½®ã€‚"""
    print(f"  æ­£åœ¨å†™å…¥è¡¨æ ¼: '{title}' (å¼€å§‹äº R{start_r}C{start_c})...")
    max_c_written = start_c - 1
    try:
        title_cell = ws.cell(row=start_r, column=start_c, value=title)
        title_cell.font = bold_f
        current_r = start_r + 1
    except Exception as e_title:
        print(f"    é”™è¯¯: å†™å…¥è¡¨æ ¼æ ‡é¢˜ '{title}' æ—¶å‡ºé”™: {e_title}")
        return start_r
    try:
        index_header = df.index.name if df.index.name else "Index"
        ws.cell(row=current_r, column=start_c, value=index_header).font = bold_f
        max_c_written = start_c
        for c_idx, col_name in enumerate(df.columns):
            col_c = start_c + 1 + c_idx
            ws.cell(row=current_r, column=col_c, value=col_name).font = bold_f
            max_c_written = col_c
        current_r += 1
    except Exception as e_header:
        print(f"    é”™è¯¯: å†™å…¥è¡¨æ ¼ '{title}' çš„è¡¨å¤´æ—¶å‡ºé”™: {e_header}")
        return start_r + 1
    try:
        for r_idx, index_val in enumerate(df.index):
            data_r = current_r + r_idx
            ws.cell(row=data_r, column=start_c, value=index_val)
            for c_idx, col_name in enumerate(df.columns):
                col_c = start_c + 1 + c_idx
                value = df.iloc[r_idx, c_idx]
                cell = ws.cell(row=data_r, column=col_c)
                if isinstance(value, (float, np.number)) and (np.isnan(value) or np.isinf(value)):
                    cell.value = None
                else:
                    cell.value = value
                cell.number_format = number_format
        final_row = current_r + len(df) - 1
    except Exception as e_data:
        print(f"    é”™è¯¯: å†™å…¥è¡¨æ ¼ '{title}' çš„æ•°æ®æ—¶å‡ºé”™: {e_data}")
        return current_r
    try:
        col_letter = get_column_letter(start_c)
        index_header = df.index.name if df.index.name else "Index"
        # Ensure index values are converted to string for length calculation
        index_lengths = df.index.astype(str).map(len)
        max_len_index = max(len(str(index_header)), index_lengths.max() if not index_lengths.empty else 0) + 2 # Handle empty index
        ws.column_dimensions[col_letter].width = max(max_len_index, 15)
        for c_idx, col_name in enumerate(df.columns):
            col_c = start_c + 1 + c_idx
            col_letter = get_column_letter(col_c)
            if number_format.endswith('%'):
                col_data_str = df.iloc[:, c_idx].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")
            elif '0' in number_format:
                num_decimals = number_format.count('0', number_format.find('.')) if '.' in number_format else 0
                col_data_str = df.iloc[:, c_idx].apply(lambda x: f"{x:.{num_decimals}f}" if pd.notna(x) else "")
            else:
                col_data_str = df.iloc[:, c_idx].astype(str)
            # Handle case where col_data_str might be empty after formatting/filtering
            max_len_data = col_data_str.map(len).max() if not col_data_str.empty else 0 
            if pd.isna(max_len_data): max_len_data = 6
            header_len = len(str(col_name))
            adjusted_width = max(max_len_data, header_len) + 2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 12)
    except Exception as e_width:
        print(f"    è­¦å‘Š: è°ƒæ•´è¡¨æ ¼ '{title}' åˆ—å®½æ—¶å‡ºé”™: {e_width}")
    return final_row
# <<< ç»“æŸç§»åŠ¨ >>>

def format_metric(value, precision=4, na_rep='N/A'):
    """ç®€å•æ ¼å¼åŒ–æ•°å­—ï¼Œä¿ç•™æŒ‡å®šç²¾åº¦ã€‚"""
    if value is None or pd.isna(value):
        return na_rep
    try:
        return f"{float(value):.{precision}f}"
    except (ValueError, TypeError):
        return str(value) # Fallback to string if conversion fails

def format_metric_pct(value, precision=2, na_rep='N/A'):
    """ç®€å•æ ¼å¼åŒ–ä¸ºç™¾åˆ†æ¯”å­—ç¬¦ä¸²ã€‚"""
    if value is None or pd.isna(value):
        return na_rep
    try:
        # Assume input is already a percentage value (e.g., 75.0 for 75%)
        return f"{float(value):.{precision}f}%"
    except (ValueError, TypeError):
        return str(value) # Fallback
